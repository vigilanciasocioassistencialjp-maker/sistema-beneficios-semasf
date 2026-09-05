from flask import Flask, render_template, request, redirect, url_for, send_file, flash, get_flashed_messages, session, jsonify
import re as _re

def _strip_html(text):
    """Remove tags HTML e converte entidades básicas para texto puro (para o PDF)."""
    if not text:
        return ''
    text = _re.sub(r'<br\s*/?>', '\n', text, flags=_re.IGNORECASE)
    text = _re.sub(r'</p>', '\n', text, flags=_re.IGNORECASE)
    text = _re.sub(r'<[^>]+>', '', text)
    text = text.replace('&nbsp;', ' ').replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>').replace('&quot;', '"')
    return _re.sub(r'\n{3,}', '\n\n', text).strip()
import bleach
from flask_wtf.csrf import CSRFProtect, generate_csrf
from flask_login import LoginManager, login_user, login_required, logout_user, current_user
from usuarios import Usuario, carregar_usuario
from banco import criar_banco, get_db_connection, _devolver_conexao, DBLogHandler
import storage_fotos
import os
import json
from datetime import datetime, timezone, timedelta
from collections import defaultdict
import bcrypt
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from reportlab.lib.utils import ImageReader
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
import io
import zipfile
import qrcode
import secrets
import logging
from logging.handlers import RotatingFileHandler
from cryptography.fernet import Fernet
import hashlib
import base64
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
import atexit

# =====================================================
# CONFIGURAÇÃO DO APP
# =====================================================

app = Flask(__name__)

# 🔒 Secret key — obrigatória no ambiente (fallback aleatório derrubaria
# todas as sessões a cada restart do Render)
app.secret_key = os.environ.get('SECRET_KEY')
if not app.secret_key:
    raise RuntimeError(
        "SECRET_KEY não definida no ambiente. Gere uma com:\n"
        '  python -c "import secrets; print(secrets.token_hex(32))"\n'
        "e configure a variável SECRET_KEY no Render (ou num arquivo .env local)."
    )

# 🛡️ Proteção CSRF global
csrf = CSRFProtect(app)

# ⏱️ Sessão expira após 8 horas de inatividade
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=8)
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = bool(os.environ.get('RENDER'))

# 🔑 Token CSRF deve durar tanto quanto a sessão. Sem isso, o token embutido
# na página expira em 1h (padrão do Flask-WTF) mesmo com a sessão ainda
# ativa, e formulários enviados após esse tempo falham com 400 "CSRF token
# has expired" mesmo que o usuário continue logado.
app.config['WTF_CSRF_TIME_LIMIT'] = int(app.config['PERMANENT_SESSION_LIFETIME'].total_seconds())

# 📷 Limite de tamanho do corpo da requisição (fotos das atividades) —
# evita que um upload gigante seja lido inteiro na memória antes mesmo do
# Pillow validar/comprimir. 10 fotos x ~5MB de folga cada.
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

# 🕐 Fuso horário de Rondônia (UTC-4)
FUSO_RONDONIA = timezone(timedelta(hours=-4))

# 🏷️ Identificador da versão em produção: o Render define RENDER_GIT_COMMIT
# automaticamente a cada deploy (estável entre reinícios/hibernações do
# serviço, o que evita falsos avisos de "nova versão"). Localmente usa a
# hora de inicialização do processo.
APP_VERSAO = (os.environ.get('RENDER_GIT_COMMIT', '')[:8]
              or 'dev-' + datetime.now(FUSO_RONDONIA).strftime('%d%m%Y-%H%M'))

#===============================================
# 🔐 CHAVE DE CRIPTOGRAFIA - FIXA NO RENDER
#===============================================

CHAVE_CRIPTO = os.environ.get('CHAVE_CRIPTO')

# Sem a chave o app NÃO pode subir: gerar uma nova a cada boot tornaria
# todos os CPFs já criptografados permanentemente ilegíveis, e imprimir
# a chave no console vazaria o segredo nos logs da plataforma.
if not CHAVE_CRIPTO:
    raise RuntimeError(
        "CHAVE_CRIPTO não definida no ambiente. Sem ela os CPFs armazenados "
        "não podem ser descriptografados. Gere uma com:\n"
        '  python -c "from cryptography.fernet import Fernet; print(Fernet.generate_key().decode())"\n'
        "e configure a variável CHAVE_CRIPTO no Render (use SEMPRE a mesma chave)."
    )

# Criar o Fernet
fernet = Fernet(CHAVE_CRIPTO.encode())

# =====================================================
# SANITIZAÇÃO DE HTML (parecer técnico vem do editor Quill)
# =====================================================

# Somente as tags de formatação que o Quill produz; tudo o mais
# (script, style, atributos de evento etc.) é removido.
_TAGS_PERMITIDAS = ['p', 'br', 'strong', 'b', 'em', 'i', 'u', 's',
                    'ol', 'ul', 'li', 'a', 'span', 'blockquote']
_ATRIBUTOS_PERMITIDOS = {'a': ['href']}

def sanitizar_html(texto):
    """Remove tags/atributos perigosos do HTML do editor rich-text."""
    if not texto:
        return texto
    return bleach.clean(texto, tags=_TAGS_PERMITIDAS,
                        attributes=_ATRIBUTOS_PERMITIDOS, strip=True)

# =====================================================
# FUNÇÕES DE CPF
# =====================================================

def hash_cpf(cpf):
    """Gera hash SHA256 para buscas (sempre igual para o mesmo CPF)"""
    cpf_limpo = ''.join(filter(str.isdigit, str(cpf)))
    return hashlib.sha256(cpf_limpo.encode()).hexdigest()

def criptografar_cpf(cpf_limpo):
    """Criptografa o CPF para salvar no banco"""
    if not cpf_limpo:
        return ''
    cpf_limpo = ''.join(filter(str.isdigit, str(cpf_limpo)))
    if len(cpf_limpo) == 11:
        return fernet.encrypt(cpf_limpo.encode()).decode()
    return cpf_limpo

def descriptografar_cpf(cpf_cripto):
    """Descriptografa o CPF para mostrar na tela"""
    if not cpf_cripto:
        return ''
    
    cpf_str = str(cpf_cripto)
    
    # Se já for um CPF em texto plano (11 dígitos), retorna ele
    cpf_limpo = ''.join(filter(str.isdigit, cpf_str))
    if len(cpf_limpo) == 11:
        return cpf_limpo
    
    # Tenta descriptografar
    try:
        valor = fernet.decrypt(cpf_str.encode()).decode()
        if len(valor) == 11 and valor.isdigit():
            return valor
        return cpf_str
    except Exception as e:
        print(f"⚠️ Erro ao descriptografar CPF: {e}")
        return cpf_str

def formatar_cpf(cpf):
    """Formata CPF: 12345678901 → 123.456.789-01"""
    if not cpf:
        return ''
    cpf_limpo = ''.join(filter(str.isdigit, str(cpf)))
    if len(cpf_limpo) == 11:
        return f"{cpf_limpo[:3]}.{cpf_limpo[3:6]}.{cpf_limpo[6:9]}-{cpf_limpo[9:]}"
    return str(cpf)

def validar_cpf(cpf):
    """Valida CPF pelos dígitos verificadores"""
    cpf = ''.join(filter(str.isdigit, str(cpf)))
    if len(cpf) != 11 or cpf == cpf[0] * 11:
        return False
    
    # Primeiro dígito
    soma = sum(int(cpf[i]) * (10 - i) for i in range(9))
    resto = (soma * 10) % 11
    if resto == 10:
        resto = 0
    if resto != int(cpf[9]):
        return False
    
    # Segundo dígito
    soma = sum(int(cpf[i]) * (11 - i) for i in range(10))
    resto = (soma * 10) % 11
    if resto == 10:
        resto = 0
    if resto != int(cpf[10]):
        return False
    
    return True

# 🛡️ Logs
if not os.path.exists('logs'):
    os.makedirs('logs')

handler = RotatingFileHandler('logs/auditoria.log', maxBytes=10000000, backupCount=5)
handler.setFormatter(logging.Formatter('%(asctime)s | %(levelname)s | %(message)s', datefmt='%d/%m/%Y %H:%M:%S'))
logger = logging.getLogger('auditoria')
logger.addHandler(handler)
# 🗄️ O arquivo acima some a cada deploy (sem disco persistente no Render) —
# este handler grava o mesmo log na tabela auditoria_log, que sobrevive.
logger.addHandler(DBLogHandler())
logger.setLevel(logging.INFO)

# Log de acesso a cada request (before_request) fica só no arquivo: é alto
# volume e baixo valor de auditoria (toda página vista, inclusive polling de
# /api/versao a cada 5min por aba aberta) — gravar isso no banco também
# inflaria a tabela de auditoria sem necessidade.
logger_acessos = logging.getLogger('acessos')
logger_acessos.addHandler(handler)
logger_acessos.setLevel(logging.INFO)
logger_acessos.propagate = False

# 🛡️ Força bruta
tentativas_login = defaultdict(list)
MAX_TENTATIVAS = 5
BLOQUEIO_MINUTOS = 15

# 🛡️ Força bruta - recuperação de senha
tentativas_recuperacao = defaultdict(list)
MAX_TENTATIVAS_RECUPERACAO = 5
RESET_TOKEN_VALIDADE_MINUTOS = 60

# 🔧 Criar banco
criar_banco()

# =====================================================
# BACKUP AUTOMATICO POR E-MAIL
# =====================================================

EMAIL_REMETENTE    = os.environ.get('EMAIL_REMETENTE', 'sistema.cestas.semasf@gmail.com')
EMAIL_DESTINATARIO = os.environ.get('EMAIL_DESTINATARIO', 'sistema.cestas.semasf@gmail.com')
BREVO_API_KEY      = os.environ.get('BREVO_API_KEY', '')
BREVO_API_URL      = 'https://api.brevo.com/v3/smtp/email'

# Todas as tabelas do banco, na ordem de dependencia (pais antes de filhos).
# Usada tanto para gerar o backup quanto, na restauracao, para apagar/inserir
# na ordem certa (insercao nessa ordem, remocao na ordem inversa).
TABELAS_BACKUP = [
    'unidades', 'cras_bairros', 'servicos', 'configuracoes',
    'usuarios', 'solicitacoes', 'atividades_fotos', 'fotos_atividade',
    'notificacoes', 'historico_edicoes', 'auditoria_log',
]

def gerar_backup_json():
    conn = get_db()
    cursor = conn.cursor()
    tabelas = {}
    for tabela in TABELAS_BACKUP:
        cursor.execute(f"SELECT * FROM {tabela}")
        colunas = [d[0] for d in cursor.description]
        tabelas[tabela] = [dict(zip(colunas, row)) for row in cursor.fetchall()]
    conn.close()
    return {
        'versao_formato': 1,
        'gerado_em': datetime.now(FUSO_RONDONIA).strftime('%d/%m/%Y %H:%M:%S'),
        'tabelas': tabelas
    }

def enviar_backup_email():
    if not BREVO_API_KEY:
        print("Backup ignorado: BREVO_API_KEY nao configurada no Render.")
        return False
    try:
        agora = datetime.now(FUSO_RONDONIA)
        nome_base = f"backup_semasf_{agora.strftime('%Y%m%d_%H%M%S')}"
        nome_arquivo = f"{nome_base}.zip"

        # Gerar e compactar backup (Brevo nao aceita anexos .gz, apenas .zip)
        dados = gerar_backup_json()
        conteudo = json.dumps(dados, ensure_ascii=False, indent=2, default=str).encode('utf-8')
        buffer_zip = io.BytesIO()
        with zipfile.ZipFile(buffer_zip, mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(f"{nome_base}.json", conteudo)
        buffer_zip.seek(0)
        tamanho_kb = round(buffer_zip.getbuffer().nbytes / 1024, 1)

        linhas_por_tabela = "\n".join(
            f"  - {tabela}: {len(dados['tabelas'][tabela])}" for tabela in TABELAS_BACKUP
        )
        corpo = (
            f"Backup automatico do Sistema de Cestas Basicas - SEMASF Ji-Parana\n\n"
            f"Data/Hora: {agora.strftime('%d/%m/%Y as %H:%M:%S')} (horario de Rondonia)\n"
            f"Tamanho: {tamanho_kb} KB (compactado)\n\n"
            f"Registros por tabela:\n{linhas_por_tabela}\n\n"
            f"Este e-mail e gerado automaticamente todo dia as 8h.\n"
            f"Guarde os ultimos 30 e-mails para manter 30 dias de historico.\n\n"
            f"-- Sistema SEMASF"
        )

        # Enviar via API HTTP do Brevo (porta 443, nao bloqueada pelo Render)
        resposta = requests.post(
            BREVO_API_URL,
            headers={'api-key': BREVO_API_KEY, 'Content-Type': 'application/json', 'Accept': 'application/json'},
            json={
                'sender': {'name': 'Sistema SEMASF', 'email': EMAIL_REMETENTE},
                'to': [{'email': EMAIL_DESTINATARIO}],
                'subject': f"[SEMASF] Backup automatico - {agora.strftime('%d/%m/%Y')}",
                'textContent': corpo,
                'attachment': [{
                    'content': base64.b64encode(buffer_zip.read()).decode(),
                    'name': nome_arquivo
                }]
            },
            timeout=30
        )
        if resposta.status_code >= 300:
            raise Exception(f"Brevo retornou {resposta.status_code}: {resposta.text}")

        print(f"Backup enviado via Brevo: {nome_arquivo} ({tamanho_kb} KB)")
        logger.info(f"Backup automatico enviado: {nome_arquivo} ({tamanho_kb} KB, {len(dados['tabelas']['solicitacoes'])} solicitacoes)")

    except Exception as e:
        print(f"Erro no backup automatico: {e}")
        logger.error(f"Erro no backup automatico: {e}")
        raise

def enviar_email_recuperacao_senha(destinatario, nome, link):
    """Envia o link de redefinição de senha via API do Brevo. Retorna True/False."""
    if not BREVO_API_KEY:
        logger.error("Recuperação de senha: BREVO_API_KEY não configurada.")
        return False
    try:
        corpo = (
            f"Olá, {nome}!\n\n"
            f"Recebemos uma solicitação para redefinir a senha da sua conta no "
            f"Sistema de Benefícios Eventuais - SEMASF Ji-Paraná.\n\n"
            f"Para criar uma nova senha, acesse o link abaixo (válido por "
            f"{RESET_TOKEN_VALIDADE_MINUTOS} minutos):\n{link}\n\n"
            f"Se você não solicitou essa alteração, apenas ignore este e-mail "
            f"e sua senha atual continuará funcionando normalmente.\n\n"
            f"-- Sistema SEMASF"
        )
        resposta = requests.post(
            BREVO_API_URL,
            headers={'api-key': BREVO_API_KEY, 'Content-Type': 'application/json', 'Accept': 'application/json'},
            json={
                'sender': {'name': 'Sistema SEMASF', 'email': EMAIL_REMETENTE},
                'to': [{'email': destinatario}],
                'subject': '[SEMASF] Redefinição de senha',
                'textContent': corpo
            },
            timeout=30
        )
        if resposta.status_code >= 300:
            raise Exception(f"Brevo retornou {resposta.status_code}: {resposta.text}")

        logger.info(f"E-mail de recuperação de senha enviado para {destinatario}")
        return True
    except Exception as e:
        logger.error(f"Erro ao enviar e-mail de recuperação de senha para {destinatario}: {e}")
        return False

scheduler = BackgroundScheduler(timezone='America/Porto_Velho')
scheduler.add_job(
    func=enviar_backup_email,
    trigger=CronTrigger(hour=8, minute=0),
    id='backup_diario',
    name='Backup diario por e-mail',
    replace_existing=True,
    misfire_grace_time=3600  # se perder as 8h, executa em até 1h depois
)
scheduler.start()
atexit.register(lambda: scheduler.shutdown())
print("Agendador de backup iniciado (todo dia às 8h, horario de Rondonia)")

# =====================================================
# HTTPS
# ===================================================

def pagina_inicial(perfil):
    """Rota para onde cada perfil deve ir após login/troca de senha/e-mail."""
    if perfil in ('admin', 'gestor'):
        return 'dashboard'
    return 'solicitacoes'

@app.before_request
def before_request():
    if os.environ.get('RENDER') and not request.is_secure:
        url = request.url.replace('http://', 'https://', 1)
        return redirect(url, 301)
    if request.endpoint and request.endpoint != 'static':
        usuario = current_user.id if current_user.is_authenticated else 'não autenticado'
        logger_acessos.info(f"Acesso: {request.method} {request.path} | IP: {request.remote_addr} | Usuário: {usuario}")

@app.after_request
def adicionar_headers_no_cache(response):
    """Impede o navegador de cachear páginas HTML, garantindo que os
    técnicos sempre recebam a versão mais recente após atualizações."""
    if response.content_type and 'text/html' in response.content_type:
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response

class _PooledConn:
    """Wrapper que devolve a conexão ao pool quando .close() é chamado."""
    def __init__(self, conn):
        self._conn = conn

    def cursor(self, *args, **kwargs):
        return self._conn.cursor(*args, **kwargs)

    def commit(self):
        self._conn.commit()

    def rollback(self):
        self._conn.rollback()

    def close(self):
        try:
            self._conn.rollback()
        except Exception:
            pass
        _devolver_conexao(self._conn)

    def __getattr__(self, name):
        return getattr(self._conn, name)


def get_db():
    return _PooledConn(get_db_connection())

# =====================================================
# LOGIN
# =====================================================

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"
login_manager.login_message = None

@login_manager.user_loader
def load_user(user_id):
    return carregar_usuario(user_id)

@app.context_processor
def inject_versao():
    return {'app_versao': APP_VERSAO}

@app.route("/api/versao")
def api_versao():
    """Consultada pelas abas abertas para detectar se houve novo deploy."""
    return jsonify({'versao': APP_VERSAO})

@app.context_processor
def inject_notificacoes_count():
    if current_user.is_authenticated:
        try:
            conn = get_db()
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM notificacoes WHERE destinatario=%s AND lida=FALSE", (current_user.id,))
            count = cur.fetchone()[0]
            conn.close()
            mostrar_modal = session.pop('mostrar_modal_notif', False) and count > 0
            return {'notificacoes_nao_lidas': count, 'mostrar_modal_notif': mostrar_modal}
        except Exception as e:
            logger.error(f"inject_notificacoes_count: falha ao consultar notificações de {current_user.id}: {e}")
            return {'notificacoes_nao_lidas': 0, 'mostrar_modal_notif': False}
    return {'notificacoes_nao_lidas': 0, 'mostrar_modal_notif': False}

@app.template_filter('fromjson')
def fromjson_filter(value):
    try:
        return json.loads(value) if value else []
    except Exception as e:
        logger.error(f"fromjson_filter: falha ao decodificar JSON ({e}): {str(value)[:200]}")
        return []

@app.template_filter('formatar_data')
def formatar_data(data):
    if not data: return ''
    try:
        if isinstance(data, str) and '-' in data:
            partes = data.split('-')
            if len(partes) == 3: return f"{partes[2]}/{partes[1]}/{partes[0]}"
    except Exception as e:
        logger.error(f"formatar_data: falha ao formatar '{data}': {e}")
    return data

# Nomes completos dos serviços (cadastro/filtros/registros) continuam intactos
# — este mapeamento só encurta o texto exibido nos cards da página Fotos
# Quadrimestral, que ficavam confusos com nomes de serviço muito extensos.
NOMES_CURTOS_SERVICO = {
    'Instituição de Acolhimento Adélia Francisca Santana': 'Instituição Adélia Francisca',
    'Instituição de Acolhimento Girassol': 'Instituição Girassol',
    'Programa de Promoção do Acesso ao Mundo do Trabalho – ACESSUAS Trabalho': 'Acessuas Trabalho',
    'Programa de Fortalecimento do Atendimento do Cadastro Único no Sistema Único de Assistência Social – PROCAD-SUAS': 'PROCAD-SUAS',
    'Serviço de Acolhimento Familiar em Família Acolhedora para Crianças e Adolescentes': 'Família Acolhedora - Crianças e Adolescentes',
    'Serviço de Acolhimento Familiar em Família Acolhedora para Pessoas Idosas': 'Família Acolhedora - Pessoa Idosa',
}

@app.template_filter('nome_curto_servico')
def nome_curto_servico(nome):
    return NOMES_CURTOS_SERVICO.get(nome, nome)

@app.route("/login", methods=["GET", "POST"])
def login():
    erro = None
    ip = request.remote_addr
    agora = datetime.now(FUSO_RONDONIA)
    tentativas_login[ip] = [t for t in tentativas_login[ip] if t > agora - timedelta(minutes=BLOQUEIO_MINUTOS)]
    
    if len(tentativas_login[ip]) >= MAX_TENTATIVAS:
        minutos = BLOQUEIO_MINUTOS - int((agora - tentativas_login[ip][0]).total_seconds() / 60)
        return render_template("login.html", erro=f"⛔ Bloqueado! Aguarde {minutos} min.", bloqueado=True)
    
    if request.method == "POST":
        usuario = request.form["usuario"]
        senha = request.form["senha"]
        conexao = get_db()
        cursor = conexao.cursor()
        cursor.execute("SELECT usuario, senha, perfil, primeiro_acesso, cras, nome, email, acesso_atividades, unidade_id FROM usuarios WHERE usuario = %s", (usuario,))
        dados = cursor.fetchone()
        cursor.close()
        conexao.close()

        if dados and bcrypt.checkpw(senha.encode('utf-8'), dados[1].encode('utf-8')):
            user = Usuario(dados[0], dados[2], dados[4] if len(dados) > 4 else None,
                            dados[5] if len(dados) > 5 else dados[0], dados[7] if len(dados) > 7 else False,
                            dados[8] if len(dados) > 8 else None)
            login_user(user, remember=False)
            session.permanent = True
            session['mostrar_modal_notif'] = True
            if ip in tentativas_login: del tentativas_login[ip]
            logger.info(f"Login: {dados[0]}")
            if dados[3] == 1:
                return redirect(url_for("trocar_senha", primeiro_acesso=True))
            if not dados[6]:
                return redirect(url_for("definir_email"))
            return redirect(url_for(pagina_inicial(dados[2])))
        else:
            tentativas_login[ip].append(agora)
            erro = "❌ Usuário ou senha incorretos!"
    
    return render_template("login.html", erro=erro, bloqueado=False)

# =====================================================
# ESQUECI / REDEFINIR SENHA
# =====================================================

@app.route("/esqueci_senha", methods=["GET", "POST"])
def esqueci_senha():
    erro = sucesso = None
    if request.method == "POST":
        ip = request.remote_addr
        agora = datetime.now(FUSO_RONDONIA)
        tentativas_recuperacao[ip] = [t for t in tentativas_recuperacao[ip] if t > agora - timedelta(minutes=BLOQUEIO_MINUTOS)]

        if len(tentativas_recuperacao[ip]) >= MAX_TENTATIVAS_RECUPERACAO:
            erro = f"⛔ Muitas tentativas. Aguarde {BLOQUEIO_MINUTOS} min."
        else:
            tentativas_recuperacao[ip].append(agora)
            usuario_login = request.form.get("usuario", "").strip()
            conexao = get_db()
            cursor = conexao.cursor()
            cursor.execute("SELECT usuario, nome, email FROM usuarios WHERE usuario = %s", (usuario_login,))
            dados = cursor.fetchone()
            if dados and dados[2]:
                token_bruto = secrets.token_urlsafe(32)
                token_hash = hashlib.sha256(token_bruto.encode()).hexdigest()
                expira = agora + timedelta(minutes=RESET_TOKEN_VALIDADE_MINUTOS)
                cursor.execute(
                    "UPDATE usuarios SET reset_token = %s, reset_token_expira = %s WHERE usuario = %s",
                    (token_hash, expira, dados[0])
                )
                conexao.commit()
                link = url_for("redefinir_senha", token=token_bruto, _external=True)
                enviar_email_recuperacao_senha(dados[2], dados[1], link)
                logger.info(f"Recuperação de senha solicitada: {dados[0]}")
            conexao.close()
            # Mensagem genérica: não revela se o usuário existe ou tem e-mail cadastrado
            sucesso = "✅ Se o usuário existir e tiver um e-mail cadastrado, enviamos um link de redefinição de senha."
    return render_template("esqueci_senha.html", erro=erro, sucesso=sucesso)

@app.route("/redefinir_senha/<token>", methods=["GET", "POST"])
def redefinir_senha(token):
    erro = sucesso = None
    token_hash = hashlib.sha256(token.encode()).hexdigest()
    conexao = get_db()
    cursor = conexao.cursor()
    cursor.execute("SELECT usuario, reset_token_expira FROM usuarios WHERE reset_token = %s", (token_hash,))
    dados = cursor.fetchone()
    conexao.close()

    agora = datetime.now(FUSO_RONDONIA)
    token_valido = bool(dados) and dados[1] is not None and dados[1] > agora

    if not token_valido:
        return render_template("redefinir_senha.html", erro="⛔ Link inválido ou expirado. Solicite um novo.",
                               sucesso=None, token_valido=False)

    if request.method == "POST":
        nova = request.form.get("nova_senha", "")
        confirma = request.form.get("confirmar_senha", "")
        if len(nova) < 6: erro = "Mínimo 6 caracteres!"
        elif nova != confirma: erro = "Senhas não conferem!"
        elif not any(c.isupper() for c in nova): erro = "Precisa de maiúscula!"
        elif not any(c.isdigit() for c in nova): erro = "Precisa de número!"
        else:
            hash_nova = bcrypt.hashpw(nova.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            conexao = get_db()
            cursor = conexao.cursor()
            cursor.execute(
                "UPDATE usuarios SET senha = %s, primeiro_acesso = 0, reset_token = NULL, reset_token_expira = NULL WHERE usuario = %s",
                (hash_nova, dados[0])
            )
            conexao.commit()
            conexao.close()
            logger.info(f"Senha redefinida via recuperação por e-mail: {dados[0]}")
            sucesso = "✅ Senha redefinida com sucesso! Você já pode fazer login."

    return render_template("redefinir_senha.html", erro=erro, sucesso=sucesso, token_valido=True)

# =====================================================
# TROCAR SENHA
# =====================================================

@app.route("/trocar_senha", methods=["GET", "POST"])
@login_required
def trocar_senha():
    erro = sucesso = None
    if request.method == "POST":
        nova = request.form.get("nova_senha", "")
        confirma = request.form.get("confirmar_senha", "")
        if len(nova) < 6: erro = "Mínimo 6 caracteres!"
        elif nova != confirma: erro = "Senhas não conferem!"
        elif not any(c.isupper() for c in nova): erro = "Precisa de maiúscula!"
        elif not any(c.isdigit() for c in nova): erro = "Precisa de número!"
        else:
            hash_nova = bcrypt.hashpw(nova.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            conexao = get_db()
            cursor = conexao.cursor()
            cursor.execute("UPDATE usuarios SET senha = %s, primeiro_acesso = 0 WHERE usuario = %s", (hash_nova, current_user.id))
            conexao.commit()
            conexao.close()
            logger.info(f"Senha alterada: {current_user.id}")
            sucesso = "✅ Senha alterada!"
            if request.args.get('primeiro_acesso'):
                return redirect(url_for(pagina_inicial(current_user.perfil)))
    return render_template("trocar_senha.html", erro=erro, sucesso=sucesso)

# =====================================================
# DEFINIR EMAIL (obrigatório para habilitar recuperação de senha)
# =====================================================

@app.route("/definir_email", methods=["GET", "POST"])
@login_required
def definir_email():
    erro = None
    if request.method == "POST":
        email = request.form.get("email", "").strip()
        if not email or '@' not in email or '.' not in email.split('@')[-1]:
            erro = "Informe um e-mail válido."
        else:
            conexao = get_db()
            cursor = conexao.cursor()
            cursor.execute("UPDATE usuarios SET email = %s WHERE usuario = %s", (email, current_user.id))
            conexao.commit()
            conexao.close()
            logger.info(f"E-mail cadastrado: {current_user.id}")
            flash("✅ E-mail cadastrado com sucesso!", "success")
            return redirect(url_for(pagina_inicial(current_user.perfil)))
    return render_template("definir_email.html", erro=erro)

# =====================================================
# LOGOUT
# =====================================================

@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))

# =====================================================
# API VERIFICAR CPF (COM HASH)
# =====================================================

@app.route("/api/verificar_cpf/<cpf>")
@login_required
def verificar_cpf(cpf):
    """Verifica validade e histórico de cestas do CPF"""
    try:
        cpf_limpo = ''.join(filter(str.isdigit, str(cpf)))
        if len(cpf_limpo) != 11 or not validar_cpf(cpf_limpo):
            return jsonify({'valido': False, 'erro': 'CPF inválido'})
        
        hash_busca = hash_cpf(cpf_limpo)
        conexao = get_db()
        cursor = conexao.cursor()
        
        # Total de cestas entregues
        cursor.execute("SELECT COUNT(*), MAX(data_entrega) FROM solicitacoes WHERE cpf_hash = %s AND status = 'Entregue'", (hash_busca,))
        r = cursor.fetchone()
        total_recebido = r[0] or 0
        ultima_entrega = r[1]
        
        # Pendente
        cursor.execute("SELECT COUNT(*) FROM solicitacoes WHERE cpf_hash = %s AND status = 'Cadastrada'", (hash_busca,))
        pendente = cursor.fetchone()[0] > 0

        # Nome
        cursor.execute("SELECT nome FROM solicitacoes WHERE cpf_hash = %s ORDER BY id DESC LIMIT 1", (hash_busca,))
        nome_row = cursor.fetchone()
        nome = nome_row[0] if nome_row else None

        # Entregas anteriores (id + data) para o front-end exibir links clicáveis
        cursor.execute("""
            SELECT id, data_entrega FROM solicitacoes
            WHERE cpf_hash = %s AND status = 'Entregue'
            ORDER BY id DESC LIMIT 10
        """, (hash_busca,))
        entregas = []
        for ent_id, ent_data in cursor.fetchall():
            data_fmt = str(ent_data) if ent_data else ''
            if '-' in data_fmt:
                partes = data_fmt[:10].split('-')
                if len(partes) == 3:
                    data_fmt = f"{partes[2]}/{partes[1]}/{partes[0]}"
            entregas.append({'id': ent_id, 'data_entrega': data_fmt})

        conexao.close()
        
        # Dias desde última entrega
        dias = None
        if ultima_entrega:
            try:
                data_ultima = datetime.strptime(str(ultima_entrega)[:10], '%Y-%m-%d' if '-' in str(ultima_entrega) else '%d/%m/%Y')
                dias = (datetime.now() - data_ultima).days
            except Exception as e:
                logger.error(f"verificar_cpf: falha ao calcular dias desde última entrega ('{ultima_entrega}'): {e}")
        
        # Alerta (dias == 0 = entrega hoje, o caso mais crítico — por isso
        # a comparação explícita com None em vez de truthiness)
        if dias is not None and dias < 90:
            alerta = 'vermelho'
        elif total_recebido > 0:
            alerta = 'amarelo'
        else:
            alerta = 'verde'
        
        return jsonify({
            'valido': True,
            'total_recebido': total_recebido,
            'dias_desde_ultima': dias,
            'pendente': pendente,
            'alerta': alerta,
            'nome': nome,
            'entregas': entregas
        })
    except Exception as e:
        logger.error(f"Erro verificar CPF: {e}")
        return jsonify({'erro': str(e)}), 500

# =====================================================
# HELPERS - BAIRROS/CRAS (lidos do banco)
# =====================================================

def pode_acessar_solicitacao(cras_solicitacao):
    """Mesma regra de visibilidade da listagem: perfis com visão global
    acessam tudo; técnico de CRAS só acessa registros da própria unidade."""
    if current_user.perfil in ['admin', 'gestor', 'creas', 'cras_volante']:
        return True
    return (cras_solicitacao or '') == (current_user.cras or '')

def get_bairros_por_cras():
    """Retorna dict {cras: [bairros]} ordenado, lido da tabela cras_bairros."""
    conn = get_db()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT cras, bairro FROM cras_bairros ORDER BY cras, bairro")
        rows = cursor.fetchall()
    finally:
        conn.close()
    resultado = {}
    for cras, bairro in rows:
        resultado.setdefault(cras, []).append(bairro)
    return resultado

def get_lista_cras():
    """Retorna lista de CRAS distintos, ordenados."""
    conn = get_db()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT cras FROM cras_bairros ORDER BY cras")
        rows = cursor.fetchall()
    finally:
        conn.close()
    return [r[0] for r in rows]

def get_lista_servicos():
    """Retorna lista de serviços cadastrados na tabela servicos, ordenados.
    Não inclui CREAS/EQUIPE VOLANTE — esses continuam fixos, ligados a
    perfis com comportamento próprio."""
    conn = get_db()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT nome FROM servicos ORDER BY nome")
        rows = cursor.fetchall()
    finally:
        conn.close()
    return [r[0] for r in rows]

# =====================================================
# UNIDADES — fonte única de verdade pra unidade de lotação (fase 1 de
# uma refatoração maior; ver plano em .claude/plans). Ainda não usadas
# em nenhuma rota — só disponíveis pra próxima etapa da migração.
# =====================================================

def get_todas_unidades():
    """Retorna todas as unidades (id, nome, categoria, tem_territorio,
    servicos_ofertados), ordenadas por categoria e nome."""
    conn = get_db()
    try:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, nome, categoria, tem_territorio, servicos_ofertados "
            "FROM unidades ORDER BY categoria, nome"
        )
        rows = cursor.fetchall()
    finally:
        conn.close()
    return rows

def get_unidade_por_id(unidade_id):
    """Retorna (id, nome, categoria, tem_territorio, servicos_ofertados)
    da unidade com o id dado, ou None."""
    if unidade_id is None:
        return None
    conn = get_db()
    try:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, nome, categoria, tem_territorio, servicos_ofertados "
            "FROM unidades WHERE id = %s",
            (unidade_id,)
        )
        return cursor.fetchone()
    finally:
        conn.close()

def get_unidade_por_nome(nome):
    """Retorna (id, nome, categoria, tem_territorio, servicos_ofertados)
    da unidade com o nome dado, ou None."""
    if not nome:
        return None
    conn = get_db()
    try:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, nome, categoria, tem_territorio, servicos_ofertados "
            "FROM unidades WHERE nome = %s",
            (nome,)
        )
        return cursor.fetchone()
    finally:
        conn.close()

# =====================================================
# PÁGINA PRINCIPAL - NOVA SOLICITAÇÃO
# =====================================================

@app.route("/", methods=["GET", "POST"])
@login_required
def inicio():
    if request.method == "POST":
        cpf = request.form["cpf"]
        cpf_limpo = ''.join(filter(str.isdigit, cpf))
        
        if not validar_cpf(cpf_limpo):
            flash('❌ CPF inválido!', 'danger')
            return render_template("index.html", sucesso=False, bairros_por_cras=get_bairros_por_cras())
        
        cpf_cripto = criptografar_cpf(cpf_limpo)
        cpf_hash = hash_cpf(cpf_limpo)

        # 🛡️ Trava contra duplo-clique / reenvio: se este mesmo usuário
        # acabou de submeter uma solicitação para o mesmo CPF há poucos
        # segundos, é quase certamente um clique duplo ou uma tentativa
        # repetida por lentidão de rede — não grava de novo.
        agora_dt = datetime.now(FUSO_RONDONIA)
        ultimo_cpf_hash = session.get('ultimo_cpf_hash')
        ultimo_cpf_ts   = session.get('ultimo_cpf_ts')
        if ultimo_cpf_hash == cpf_hash and ultimo_cpf_ts:
            try:
                segundos_desde = (agora_dt - datetime.fromisoformat(ultimo_cpf_ts)).total_seconds()
            except Exception:
                segundos_desde = 999
            if segundos_desde < 20:
                logger.info(f"Envio duplicado bloqueado (clique duplo): CPF já enviado há {segundos_desde:.1f}s por {current_user.id}")
                flash('⚠️ Esta solicitação já foi registrada há poucos segundos. Evitamos um cadastro duplicado — confira na lista de solicitações.', 'warning')
                return redirect(url_for('solicitacoes'))
        session['ultimo_cpf_hash'] = cpf_hash
        session['ultimo_cpf_ts']   = agora_dt.isoformat()

        # Dados do formulário
        nome = request.form.get("nome", "")
        data_nasc = request.form.get("data_nascimento", "")
        telefone = request.form.get("telefone", "")
        email = request.form.get("email", "")
        endereco = request.form.get("endereco", "")
        numero = request.form.get("numero", "")
        complemento = request.form.get("complemento", "")
        bairro = request.form.get("bairro", "")
        cep = request.form.get("cep", "")
        referencia = request.form.get("referencia", "")
        cras = request.form.get("cras", "")
        data_escuta = request.form.get("data_escuta", "")
        
        # Membros da família
        membros_nomes = request.form.getlist("membro_nome[]")
        membros_idades = request.form.getlist("membro_idade[]")
        membros_vinculos = request.form.getlist("membro_vinculo[]")
        composicao = []
        for i in range(len(membros_nomes)):
            if membros_nomes[i].strip():
                composicao.append({
                    'nome': membros_nomes[i],
                    'idade': membros_idades[i] if i < len(membros_idades) else '',
                    'vinculo': membros_vinculos[i] if i < len(membros_vinculos) else ''
                })
        composicao_json = json.dumps(composicao, ensure_ascii=False)
        total_pessoas = len(composicao)
        
        # Renda
        renda_bruta = float(request.form.get("renda_bruta", 0) or 0)
        renda_per_capita = float(request.form.get("renda_per_capita", 0) or 0)
        beneficios = request.form.get("beneficios", "")
        
        # Vulnerabilidades
        vulnerabilidades = request.form.getlist("vulnerabilidade")
        vulnerabilidade_text = ", ".join(vulnerabilidades) if vulnerabilidades else ""
        
        # Serviços SUAS
        servicos = request.form.getlist("servicos_suas")
        servicos_text = ", ".join(servicos) if servicos else ""
        
        # Parecer (sanitizado: HTML do Quill pode carregar script malicioso)
        parecer = sanitizar_html(request.form.get("parecer", ""))

        # Exceção Art. 64 - concessão fora dos critérios ordinários
        excecao_art64 = request.form.get("excecao_art64") == "1"

        # Novos campos
        valor_bolsa_familia = 0.0
        if beneficios == "Bolsa Família":
            try:
                vbf = request.form.get("valor_bolsa_familia", "0").replace('.','').replace(',','.')
                valor_bolsa_familia = float(vbf or 0)
            except (ValueError, AttributeError) as e:
                logger.error(f"Cadastro: valor do Bolsa Família inválido ('{request.form.get('valor_bolsa_familia')}') por {current_user.id}: {e}")
                flash('❌ Valor do Bolsa Família inválido. Digite apenas números (ex: 600,00).', 'danger')
                return render_template("index.html", sucesso=False, bairros_por_cras=get_bairros_por_cras())
        visita_domiciliar = request.form.get("visita_domiciliar") == "1"

        # Validação server-side: renda per capita vs critério legal
        salario_minimo = get_salario_minimo()
        limite_rpc = salario_minimo / 4
        if renda_per_capita > limite_rpc and not excecao_art64:
            flash(f'❌ Renda per capita (R$ {renda_per_capita:.2f}) ultrapassa o limite legal de R$ {limite_rpc:.2f} (1/4 do salário mínimo). Marque a exceção do Art. 64 para prosseguir.', 'danger')
            return render_template("index.html", sucesso=False, bairros_por_cras=get_bairros_por_cras())
        
        tecnico = current_user.id
        data_solic = datetime.now(FUSO_RONDONIA).strftime("%d/%m/%Y %H:%M:%S")
        
        conexao = get_db()
        cursor = conexao.cursor()
        cursor.execute("""
            INSERT INTO solicitacoes (tecnico, cpf, cpf_hash, nome, data_nascimento, telefone, email, endereco, numero, complemento, bairro, cep, referencia, cras, data_escuta, total_pessoas, composicao_familiar, renda_bruta, renda_per_capita, beneficios, vulnerabilidade, servicos_suas, parecer, status, data_solicitacao, excecao_art64, valor_bolsa_familia, visita_domiciliar)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (tecnico, cpf_cripto, cpf_hash, nome, data_nasc, telefone, email, endereco, numero, complemento, bairro, cep, referencia, cras, data_escuta, total_pessoas, composicao_json, renda_bruta, renda_per_capita, beneficios, vulnerabilidade_text, servicos_text, parecer, 'Cadastrada', data_solic, excecao_art64, valor_bolsa_familia, visita_domiciliar))
        conexao.commit()
        conexao.close()
        
        logger.info(f"Solicitação cadastrada: {nome}")
        flash('✅ Solicitação cadastrada!', 'success')
        return redirect(url_for("solicitacoes"))
    bairros_por_cras = get_bairros_por_cras()
    return render_template("index.html", sucesso=False,
                           bairros_por_cras=bairros_por_cras)

# =====================================================
# LISTAR SOLICITAÇÕES (CPF DESCRIPTOGRAFADO)
# =====================================================

# Colunas em que a listagem de Solicitações pode ser ordenada ao clicar no
# cabeçalho — chave usada na URL (?ordenar_por=) mapeada para a expressão SQL
# real. data_solicitacao é TEXT em formato 'DD/MM/AAAA HH:MM:SS', então a
# expressão reordena os pedaços para AAAAMMDD... antes de comparar, senão a
# ordenação ficaria alfabética (dia) em vez de cronológica.
COLUNAS_ORDENAVEIS_SOLICITACOES = {
    'id': 's.id',
    'tecnico': 's.tecnico',
    'nome': 's.nome',
    'bairro': 's.bairro',
    'unidade': 's.cras',
    'data_solicitacao': (
        "(SUBSTRING(s.data_solicitacao, 7, 4) || SUBSTRING(s.data_solicitacao, 4, 2) "
        "|| SUBSTRING(s.data_solicitacao, 1, 2) || SUBSTRING(s.data_solicitacao, 12, 8))"
    ),
    'status': 's.status',
}

# Rótulo de unidade a partir do perfil do técnico (u) / cras da solicitação
# (s) — hoje esse mesmo CASE está duplicado (com divergências) em pelo
# menos 3 lugares (dashboard, relatório, PDF). Ainda não usada em nenhuma
# query — fica pronta pra próxima etapa da migração de unidades substituir
# as cópias existentes. Requer JOIN "usuarios u ON s.tecnico = u.usuario"
# (ou alias equivalente) e as tabelas terem os aliases "u"/"s".
UNIDADE_CASE_SQL = (
    "CASE "
    "WHEN u.perfil = 'creas' THEN 'CREAS' "
    "WHEN u.perfil = 'cras_volante' THEN 'EQUIPE VOLANTE' "
    "WHEN u.perfil IN ('admin', 'gestor') THEN 'ADMINISTRAÇÃO' "
    "ELSE COALESCE(NULLIF(u.cras, ''), s.cras, 'Não informado') "
    "END"
)

def _construir_filtro_solicitacoes(busca_nome, busca_status, busca_unidade,
                                    busca_tecnico_escuta, busca_tecnico_entrega,
                                    busca_tecnico_qualquer, periodo_inicio, periodo_fim,
                                    current_user, lista_cras_atual, incluir_status=True):
    """Monta (filtros, params, join_usuario) pros filtros da página de
    Solicitações — reaproveitado pela listagem, pelos cards por setor e pelo
    PDF. incluir_status=False ignora o filtro de Status (usado pelos cards,
    que sempre quebram em Cadastrada/Entregue/Ausente, independente do
    status escolhido na tela)."""
    filtros = []
    params  = []
    join_usuario = ""

    if current_user.perfil not in ['admin', 'gestor', 'creas', 'cras_volante']:
        filtros.append("s.cras = %s")
        params.append(current_user.cras)

    if busca_nome:
        filtros.append("s.nome ILIKE %s")
        params.append(f"%{busca_nome}%")

    # Sem filtro de status escolhido, mostra só Cadastrada + Ausente (ainda
    # pendentes de ação) — Entregue/Cancelada só aparecem se o técnico
    # selecionar explicitamente no filtro (inclusive escolhendo "Todos").
    # Visita Domiciliar é um marcador à parte, sem relação com esse padrão.
    if incluir_status:
        if busca_status == 'Visita':
            filtros.append("s.visita_domiciliar = TRUE")
        elif busca_status == 'Pendentes':
            filtros.append("s.status IN ('Cadastrada', 'Ausente')")
        elif busca_status:
            filtros.append("s.status = %s")
            params.append(busca_status)

    if busca_tecnico_escuta:
        filtros.append("s.tecnico = %s")
        params.append(busca_tecnico_escuta)

    if busca_tecnico_entrega:
        filtros.append("s.tecnico_entrega = %s")
        params.append(busca_tecnico_entrega)

    if busca_tecnico_qualquer:
        filtros.append("(s.tecnico = %s OR s.tecnico_entrega = %s)")
        params.extend([busca_tecnico_qualquer, busca_tecnico_qualquer])

    if busca_unidade:
        if busca_unidade == 'ADMINISTRAÇÃO':
            join_usuario = "LEFT JOIN usuarios u ON s.tecnico = u.usuario"
            filtros.append("u.perfil IN ('admin', 'gestor')")
        elif busca_unidade in lista_cras_atual or busca_unidade in ('CREAS', 'EQUIPE VOLANTE'):
            # Filtra pela unidade RESPONSÁVEL PELA ENTREGA (território), não
            # pelo perfil de quem registrou a escuta — mesma regra de
            # /lista_entrega, ver _filtro_unidade_entrega()
            join_entrega, cond_entrega, params_entrega = _filtro_unidade_entrega(busca_unidade)
            join_usuario = join_entrega
            filtros.append(cond_entrega)
            params.extend(params_entrega)
        else:
            join_usuario = "LEFT JOIN usuarios u ON s.tecnico = u.usuario"
            filtros.append("COALESCE(u.cras, s.cras) = %s")
            params.append(busca_unidade)

    cond_periodo, p_periodo = _periodo_sql_dia(periodo_inicio, periodo_fim)
    if cond_periodo:
        filtros.append(cond_periodo)
        params.extend(p_periodo)

    return filtros, params, join_usuario


def _contagem_setores(filtros_base, params_base, join_usuario, current_user, busca_unidade):
    """Cadastrada/Entregue/Ausente por unidade RESPONSÁVEL PELA ENTREGA (mesma
    regra do _filtro_unidade_entrega), pros cards no topo da listagem de
    Solicitações. filtros_base/params_base/join_usuario vêm de
    _construir_filtro_solicitacoes(..., incluir_status=False) — ou seja, já
    respeitam nome/técnico/período/unidade, mas nunca o Status.

    Se busca_unidade estiver definido, ou o perfil do usuário for restrito a
    uma unidade só, retorna 1 card (a unidade já filtrada). Senão, agrupa
    por unidade usando um CASE que reproduz a mesma regra de roteamento do
    _filtro_unidade_entrega (bairro rural/Volante → CREAS pelo perfil do
    técnico → CRAS do técnico → s.cras como fallback)."""
    where = ("WHERE " + " AND ".join(filtros_base)) if filtros_base else ""
    and_status = (" AND " if where else " WHERE ") + "s.status IN ('Cadastrada', 'Entregue', 'Ausente')"

    conexao = get_db()
    cursor  = conexao.cursor()

    unico_setor = bool(busca_unidade) or current_user.perfil not in ['admin', 'gestor', 'creas', 'cras_volante']
    chave_status = {'Cadastrada': 'cadastradas', 'Entregue': 'entregues', 'Ausente': 'ausentes'}

    if unico_setor:
        rotulo = busca_unidade or current_user.cras or 'Sem unidade'
        cursor.execute(f"""
            SELECT s.status, COUNT(*) FROM solicitacoes s {join_usuario}
            {where}{and_status}
            GROUP BY s.status
        """, params_base)
        contagens = {row[0]: row[1] for row in cursor.fetchall()}
        setores = [{
            'nome': rotulo,
            'cadastradas': contagens.get('Cadastrada', 0),
            'entregues':   contagens.get('Entregue', 0),
            'ausentes':    contagens.get('Ausente', 0),
        }]
    else:
        cursor.execute(f"""
            SELECT
              CASE
                WHEN COALESCE(s.bairro, '') IN (SELECT bairro FROM cras_bairros WHERE entrega_volante = TRUE)
                     OR s.cras = 'EQUIPE VOLANTE' THEN 'EQUIPE VOLANTE'
                WHEN u_setor.perfil = 'creas' THEN 'CREAS'
                WHEN u_setor.perfil = 'cras'  THEN u_setor.cras
                ELSE s.cras
              END AS unidade, s.status, COUNT(*)
            FROM solicitacoes s LEFT JOIN usuarios u_setor ON s.tecnico = u_setor.usuario
            {where}{and_status}
            GROUP BY 1, s.status
        """, params_base)
        agregados = {}
        for unidade, status, qtd in cursor.fetchall():
            unidade = unidade or 'Não informado'
            agregados.setdefault(unidade, {'cadastradas': 0, 'entregues': 0, 'ausentes': 0})
            agregados[unidade][chave_status[status]] = qtd
        setores = [{'nome': nome, **valores} for nome, valores in sorted(agregados.items())]

    conexao.close()
    for s in setores:
        s['total'] = s['cadastradas'] + s['entregues'] + s['ausentes']
    return setores


def _formatar_data_entrega(valor):
    """Converte data_entrega de 'YYYY-MM-DD...' para 'DD/MM/YYYY'; vazio se não entregue."""
    data_fmt = str(valor)[:10] if valor else ''
    if '-' in data_fmt:
        partes = data_fmt.split('-')
        if len(partes) == 3:
            data_fmt = f"{partes[2]}/{partes[1]}/{partes[0]}"
    return data_fmt

@app.route("/solicitacoes")
@app.route("/solicitacoes/<int:pagina>")
@login_required
def solicitacoes(pagina=1):
    registros_por_pagina = 100
    offset = (pagina - 1) * registros_por_pagina

    ordenar_por = request.args.get('ordenar_por', 'data_solicitacao').strip()
    ordenar_dir = request.args.get('ordenar_dir', 'asc').strip().lower()
    if ordenar_por not in COLUNAS_ORDENAVEIS_SOLICITACOES:
        ordenar_por = 'data_solicitacao'
    if ordenar_dir not in ('asc', 'desc'):
        ordenar_dir = 'asc'
    order_by_sql = f"{COLUNAS_ORDENAVEIS_SOLICITACOES[ordenar_por]} {ordenar_dir.upper()}, s.id {ordenar_dir.upper()}"

    busca_nome              = request.args.get('busca_nome', '').strip()
    busca_status            = request.args.get('busca_status', 'Pendentes').strip()
    busca_cpf               = request.args.get('busca_cpf', '').strip()
    busca_unidade           = request.args.get('busca_unidade', '').strip()
    busca_tecnico_escuta    = request.args.get('busca_tecnico_escuta', '').strip()
    busca_tecnico_entrega   = request.args.get('busca_tecnico_entrega', '').strip()
    busca_tecnico_qualquer  = request.args.get('busca_tecnico_qualquer', '').strip()
    periodo_inicio          = request.args.get('periodo_inicio', '').strip()
    periodo_fim             = request.args.get('periodo_fim', '').strip()

    lista_cras_atual = get_lista_cras()

    filtros, params, join_usuario = _construir_filtro_solicitacoes(
        busca_nome, busca_status, busca_unidade, busca_tecnico_escuta,
        busca_tecnico_entrega, busca_tecnico_qualquer, periodo_inicio, periodo_fim,
        current_user, lista_cras_atual, incluir_status=True)
    where = ("WHERE " + " AND ".join(filtros)) if filtros else ""

    conexao = get_db()
    cursor  = conexao.cursor()

    # Listas para dropdowns
    cursor.execute("SELECT usuario, nome FROM usuarios ORDER BY nome")
    lista_tecnicos = cursor.fetchall()
    lista_unidades = lista_cras_atual + ['CREAS', 'EQUIPE VOLANTE'] + get_lista_servicos()

    base_query = f"FROM solicitacoes s {join_usuario} LEFT JOIN cras_bairros cb ON s.bairro = cb.bairro {where}"

    # Busca por CPF: descriptografa em Python
    if busca_cpf:
        cpf_limpo_busca = _re.sub(r'\D', '', busca_cpf)
        cursor.execute(
            f"SELECT s.id, s.tecnico, s.nome, s.cpf, s.bairro, s.cras, s.data_solicitacao, s.status, s.visita_domiciliar, "
            f"COALESCE(cb.entrega_volante, FALSE), s.data_entrega "
            f"{base_query} ORDER BY {order_by_sql}",
            params
        )
        todos = cursor.fetchall()
        dados_filtrados = []
        for row in todos:
            row = list(row)
            cpf_desc = descriptografar_cpf(row[3]) if row[3] else ''
            if cpf_limpo_busca in cpf_desc:
                row[3] = formatar_cpf(cpf_desc)
                row[10] = _formatar_data_entrega(row[10])
                dados_filtrados.append(tuple(row))
        total_registros = len(dados_filtrados)
        dados = dados_filtrados[(pagina - 1) * registros_por_pagina: pagina * registros_por_pagina]
    else:
        cursor.execute(f"SELECT COUNT(*) {base_query}", params)
        total_registros = cursor.fetchone()[0]
        cursor.execute(
            f"SELECT s.id, s.tecnico, s.nome, s.cpf, s.bairro, s.cras, s.data_solicitacao, s.status, s.visita_domiciliar, "
            f"COALESCE(cb.entrega_volante, FALSE), s.data_entrega "
            f"{base_query} ORDER BY {order_by_sql} LIMIT %s OFFSET %s",
            params + [registros_por_pagina, offset]
        )
        dados_raw = cursor.fetchall()
        dados = []
        for row in dados_raw:
            row = list(row)
            if row[3]:
                row[3] = formatar_cpf(descriptografar_cpf(row[3]))
            row[10] = _formatar_data_entrega(row[10])
            dados.append(tuple(row))

    conexao.close()

    total_paginas = max(1, (total_registros + registros_por_pagina - 1) // registros_por_pagina)

    # Cards por setor: ignoram o Status ativo (sempre quebram nos 3 valores).
    # Não calculados durante busca por CPF — nesse caso o filtro já restringe
    # a uma pessoa só, e a quebra por setor não agrega valor.
    setores = None
    if not busca_cpf:
        filtros_setores, params_setores, join_setores = _construir_filtro_solicitacoes(
            busca_nome, busca_status, busca_unidade, busca_tecnico_escuta,
            busca_tecnico_entrega, busca_tecnico_qualquer, periodo_inicio, periodo_fim,
            current_user, lista_cras_atual, incluir_status=False)
        setores = _contagem_setores(filtros_setores, params_setores, join_setores, current_user, busca_unidade)

    return render_template(
        "solicitacoes.html",
        solicitacoes=dados,
        numero_inicial=offset,
        setores=setores,
        user_perfil=current_user.perfil,
        datetime=datetime,
        pagina_atual=pagina,
        total_paginas=total_paginas,
        total_registros=total_registros,
        ordenar_por=ordenar_por,
        ordenar_dir=ordenar_dir,
        busca_nome=busca_nome,
        busca_status=busca_status,
        busca_cpf=busca_cpf,
        busca_unidade=busca_unidade,
        busca_tecnico_escuta=busca_tecnico_escuta,
        busca_tecnico_entrega=busca_tecnico_entrega,
        busca_tecnico_qualquer=busca_tecnico_qualquer,
        periodo_inicio=periodo_inicio,
        periodo_fim=periodo_fim,
        lista_tecnicos=lista_tecnicos,
        lista_unidades=lista_unidades,
        current_user=current_user
    )


@app.route("/solicitacoes/pdf")
@login_required
def solicitacoes_pdf():
    """PDF com os cards por setor + a lista completa do filtro ativo (não só
    a página atual) — pra levar pra reunião de equipe. Mesmos filtros da
    tela de Solicitações, recebidos na querystring."""
    ordenar_por = request.args.get('ordenar_por', 'data_solicitacao').strip()
    ordenar_dir = request.args.get('ordenar_dir', 'asc').strip().lower()
    if ordenar_por not in COLUNAS_ORDENAVEIS_SOLICITACOES:
        ordenar_por = 'data_solicitacao'
    if ordenar_dir not in ('asc', 'desc'):
        ordenar_dir = 'asc'
    order_by_sql = f"{COLUNAS_ORDENAVEIS_SOLICITACOES[ordenar_por]} {ordenar_dir.upper()}, s.id {ordenar_dir.upper()}"

    busca_nome              = request.args.get('busca_nome', '').strip()
    busca_status            = request.args.get('busca_status', 'Pendentes').strip()
    busca_cpf               = request.args.get('busca_cpf', '').strip()
    busca_unidade           = request.args.get('busca_unidade', '').strip()
    busca_tecnico_escuta    = request.args.get('busca_tecnico_escuta', '').strip()
    busca_tecnico_entrega   = request.args.get('busca_tecnico_entrega', '').strip()
    busca_tecnico_qualquer  = request.args.get('busca_tecnico_qualquer', '').strip()
    periodo_inicio          = request.args.get('periodo_inicio', '').strip()
    periodo_fim             = request.args.get('periodo_fim', '').strip()

    lista_cras_atual = get_lista_cras()
    filtros, params, join_usuario = _construir_filtro_solicitacoes(
        busca_nome, busca_status, busca_unidade, busca_tecnico_escuta,
        busca_tecnico_entrega, busca_tecnico_qualquer, periodo_inicio, periodo_fim,
        current_user, lista_cras_atual, incluir_status=True)
    where = ("WHERE " + " AND ".join(filtros)) if filtros else ""
    base_query = f"FROM solicitacoes s {join_usuario} {where}"

    conexao = get_db()
    cursor  = conexao.cursor()

    # Segurança: sem CPF a busca não tem um teto natural, então limita o PDF
    # a um número razoável de linhas (evita gerar um documento gigante sem
    # querer quando nenhum filtro de período/unidade está ativo).
    LIMITE_LINHAS = 3000
    truncado = False
    campos = "s.id, s.nome, s.cpf, s.bairro, s.cras, s.data_solicitacao, s.status"

    if busca_cpf:
        # Busca por CPF: descriptografa em Python, sem limite (é sempre uma
        # pessoa só, mesmo padrão da listagem principal).
        cursor.execute(f"SELECT {campos} {base_query} ORDER BY {order_by_sql}", params)
        cpf_limpo_busca = _re.sub(r'\D', '', busca_cpf)
        linhas = []
        for row in cursor.fetchall():
            row = list(row)
            cpf_desc = descriptografar_cpf(row[2]) if row[2] else ''
            if cpf_limpo_busca in cpf_desc:
                row[2] = formatar_cpf(cpf_desc)
                linhas.append(row)
    else:
        cursor.execute(
            f"SELECT {campos} {base_query} ORDER BY {order_by_sql} LIMIT %s",
            params + [LIMITE_LINHAS + 1]
        )
        linhas_raw = cursor.fetchall()
        truncado = len(linhas_raw) > LIMITE_LINHAS
        linhas = []
        for row in linhas_raw[:LIMITE_LINHAS]:
            row = list(row)
            if row[2]:
                row[2] = formatar_cpf(descriptografar_cpf(row[2]))
            linhas.append(row)

    if not linhas:
        conexao.close()
        flash("Nenhuma solicitação encontrada para esse filtro.", "warning")
        return redirect(url_for("solicitacoes"))

    setores = None
    if not busca_cpf:
        filtros_setores, params_setores, join_setores = _construir_filtro_solicitacoes(
            busca_nome, busca_status, busca_unidade, busca_tecnico_escuta,
            busca_tecnico_entrega, busca_tecnico_qualquer, periodo_inicio, periodo_fim,
            current_user, lista_cras_atual, incluir_status=False)
        setores = _contagem_setores(filtros_setores, params_setores, join_setores, current_user, busca_unidade)

    conexao.close()

    hoje = datetime.now(FUSO_RONDONIA)
    PAGE_W, PAGE_H = landscape(A4)

    partes_filtro = []
    if periodo_inicio or periodo_fim:
        partes_filtro.append(f"PERÍODO: {periodo_inicio or '...'} a {periodo_fim or '...'}")
    if busca_status:
        partes_filtro.append(f"STATUS: {busca_status}")
    if busca_unidade:
        partes_filtro.append(f"UNIDADE: {busca_unidade}")
    subtitulo_filtro = "   |   ".join(partes_filtro) if partes_filtro else "TODAS AS SOLICITAÇÕES"

    def _cabecalho_solicitacoes(canvas_obj, doc):
        canvas_obj.saveState()
        logo_path = os.path.join(os.path.dirname(__file__), 'static', 'img', 'logo_prefeitura.png')
        if os.path.exists(logo_path):
            try:
                canvas_obj.drawImage(ImageReader(logo_path), 1*cm, PAGE_H - 2.2*cm,
                                     width=4.5*cm, height=1.5*cm,
                                     preserveAspectRatio=True, mask='auto')
            except Exception:
                pass
        canvas_obj.setFont("Helvetica-Bold", 12)
        canvas_obj.drawCentredString(PAGE_W/2, PAGE_H - 1.1*cm, "PREFEITURA MUNICIPAL DE JI-PARANÁ")
        canvas_obj.setFont("Helvetica", 9)
        canvas_obj.drawCentredString(PAGE_W/2, PAGE_H - 1.55*cm,
                                     "Secretaria Municipal de Assistência Social e Família - SEMASF")
        canvas_obj.setFont("Helvetica-Bold", 11)
        canvas_obj.drawCentredString(PAGE_W/2, PAGE_H - 2.15*cm, "LISTA DE SOLICITAÇÕES DE CESTA BÁSICA")
        canvas_obj.setFont("Helvetica-Bold", 9)
        canvas_obj.drawCentredString(PAGE_W/2, PAGE_H - 2.6*cm,
                                     f"{subtitulo_filtro}   |   TOTAL: {len(linhas)}")
        canvas_obj.setLineWidth(0.8)
        canvas_obj.line(1*cm, PAGE_H - 2.8*cm, PAGE_W - 1*cm, PAGE_H - 2.8*cm)

        canvas_obj.setFont("Helvetica", 6.5)
        canvas_obj.setFillColorRGB(0.35, 0.35, 0.35)
        canvas_obj.drawString(1*cm, 0.8*cm,
            f"Documento gerado pelo Sistema de Cestas Básicas SEMASF em {hoje.strftime('%d/%m/%Y às %H:%M')} (horário de Rondônia)")
        canvas_obj.restoreState()

    estilo_celula = ParagraphStyle('celulaSol', fontName='Helvetica', fontSize=7.5, leading=9)
    elementos = []

    if setores:
        dados_setores = [['Unidade', 'Cadastradas', 'Entregues', 'Ausentes', 'Total']]
        for st in setores:
            dados_setores.append([st['nome'], str(st['cadastradas']), str(st['entregues']),
                                   str(st['ausentes']), str(st['total'])])
        tabela_setores = Table(dados_setores, colWidths=[8*cm, 3.5*cm, 3.5*cm, 3.5*cm, 3.5*cm])
        tabela_setores.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0), (-1, 0), colors.HexColor('#1B2F5E')),
            ('TEXTCOLOR',     (0, 0), (-1, 0), colors.white),
            ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE',      (0, 0), (-1, -1), 8),
            ('ALIGN',         (1, 0), (-1, -1), 'CENTER'),
            ('GRID',          (0, 0), (-1, -1), 0.5, colors.black),
            ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F4F8')]),
            ('TOPPADDING',    (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        elementos.append(tabela_setores)
        elementos.append(Spacer(1, 0.6*cm))

    cabecalho = ['Nº', 'ID', 'Nome', 'CPF', 'Bairro', 'Unidade', 'Data Solicitação', 'Status']
    dados_tabela = [cabecalho]
    for i, (id_, nome, cpf, bairro, cras, data_sol, status) in enumerate(linhas, 1):
        dados_tabela.append([
            str(i),
            f"#{id_}",
            Paragraph(nome or '—', estilo_celula),
            cpf or '—',
            Paragraph(bairro or '—', estilo_celula),
            Paragraph(cras or '—', estilo_celula),
            (data_sol or '—')[:10],
            status or '—',
        ])
    larguras = [1.1*cm, 1.6*cm, 6*cm, 3*cm, 3.5*cm, 3.5*cm, 3*cm, 2.4*cm]
    tabela = Table(dados_tabela, colWidths=larguras, repeatRows=1)
    tabela.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0), colors.HexColor('#1B2F5E')),
        ('TEXTCOLOR',     (0, 0), (-1, 0), colors.white),
        ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, 0), 7.5),
        ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',      (0, 1), (-1, -1), 7.5),
        ('ALIGN',         (0, 0), (1, -1), 'CENTER'),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID',          (0, 0), (-1, -1), 0.5, colors.black),
        ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F4F8')]),
        ('TOPPADDING',    (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    elementos.append(tabela)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=1*cm, rightMargin=1*cm,
        topMargin=3.1*cm, bottomMargin=1.4*cm,
    )
    doc.build(elementos, onFirstPage=_cabecalho_solicitacoes, onLaterPages=_cabecalho_solicitacoes,
              canvasmaker=_CanvasNumerado)
    buffer.seek(0)

    if truncado:
        flash(f"⚠️ O filtro retornou mais de {LIMITE_LINHAS} solicitações; o PDF foi limitado às "
              f"{LIMITE_LINHAS} primeiras. Refine o período ou a unidade para ver a lista completa.", "warning")

    logger.info(f"PDF de Solicitações gerado por {current_user.id} ({len(linhas)} registros)")
    nome_arquivo = f"solicitacoes_{hoje.strftime('%Y%m%d_%H%M')}.pdf"
    return send_file(buffer, as_attachment=True, download_name=nome_arquivo, mimetype='application/pdf')


# =====================================================
# VER SOLICITAÇÕES
# =====================================================

@app.route("/ver_solicitacao/<int:id>")
@login_required
def ver_solicitacao(id):
    conexao = get_db()
    cursor = conexao.cursor()
    cursor.execute("""
        SELECT 
            s.id, s.nome, s.cpf, s.data_nascimento, s.telefone, 
            s.data_solicitacao, s.email, s.endereco, s.numero, 
            s.complemento, s.bairro, s.cep, s.referencia, s.cras, 
            s.renda_bruta, s.renda_per_capita, s.beneficios, 
            s.vulnerabilidade, s.data_entrega, s.tecnico_entrega,
            s.parecer, s.status, s.tecnico, s.composicao_familiar,
            s.servicos_suas, u_escuta.nome as tecnico_escuta_nome,
            u_entrega.nome as tecnico_entrega_nome,
            s.num_tentativas,
            s.valor_bolsa_familia,
            s.visita_domiciliar,
            COALESCE(cb.entrega_volante, FALSE)
        FROM solicitacoes s
        LEFT JOIN usuarios u_escuta ON s.tecnico = u_escuta.usuario
        LEFT JOIN usuarios u_entrega ON s.tecnico_entrega = u_entrega.usuario
        LEFT JOIN cras_bairros cb ON s.bairro = cb.bairro
        WHERE s.id = %s
    """, (id,))
    s = cursor.fetchone()
    if not s:
        conexao.close()
        return "Não encontrada", 404

    # s[13] = cras da solicitação
    if not pode_acessar_solicitacao(s[13]):
        conexao.close()
        flash('❌ Você não tem permissão para acessar solicitações de outra unidade.', 'danger')
        return redirect(url_for('solicitacoes'))

    s = list(s)
    if s[2]:
        s[2] = formatar_cpf(descriptografar_cpf(s[2]))
    # Sanitiza o parecer (s[20]) — o template exibe com |safe, e registros
    # antigos podem ter sido salvos antes da sanitização no cadastro
    s[20] = sanitizar_html(s[20])
    if s[3]:
        try:
            if '-' in str(s[3]):
                partes = str(s[3]).split('-')
                if len(partes) == 3:
                    s[3] = f"{partes[2]}/{partes[1]}/{partes[0]}"
        except Exception as e:
            logger.error(f"ver_solicitacao #{id}: falha ao formatar data de nascimento '{s[3]}': {e}")

    # Histórico de edições
    cursor.execute("""
        SELECT campo, valor_antes, valor_depois, usuario, data_hora
        FROM historico_edicoes
        WHERE solicitacao_id = %s
        ORDER BY data_hora DESC
        LIMIT 50
    """, (id,))
    historico = cursor.fetchall()
    conexao.close()

    return render_template(
        "ver_solicitacao.html",
        solicitacao=s,
        historico=historico,
        json=json,
        datetime=datetime,
        current_user=current_user
    )

# =====================================================
# EDITAR SOLICITAÇÃO
# =====================================================

@app.route("/editar_solicitacao/<int:id>", methods=["GET", "POST"])
@login_required
def editar_solicitacao(id):
    conexao = get_db()
    cursor = conexao.cursor()

    # Buscar solicitação
    cursor.execute("SELECT * FROM solicitacoes WHERE id = %s", (id,))
    row = cursor.fetchone()
    if not row:
        conexao.close()
        return "Solicitação não encontrada", 404

    cols = [d[0] for d in cursor.description]
    s = dict(zip(cols, row))

    # Descriptografar CPF para exibição
    s['cpf'] = formatar_cpf(descriptografar_cpf(s['cpf'])) if s['cpf'] else ''

    # Verificar permissão: só o autor (status Cadastrada) ou admin
    eh_admin   = current_user.perfil == 'admin'
    eh_autor   = current_user.id == s['tecnico']
    ja_entregue = s.get('status') in ('Entregue', 'Ausente')

    pode_editar = eh_admin or (eh_autor and not ja_entregue)

    if not pode_editar:
        conexao.close()
        if ja_entregue:
            flash('❌ Esta solicitação já teve a entrega registrada e não pode mais ser editada.', 'danger')
        else:
            flash('❌ Você não tem permissão para editar esta solicitação.', 'danger')
        return redirect(url_for('ver_solicitacao', id=id))

    if request.method == "POST":
        nome         = request.form.get("nome", "")
        data_nasc    = request.form.get("data_nascimento", "")
        telefone     = request.form.get("telefone", "")
        email        = request.form.get("email", "")
        endereco     = request.form.get("endereco", "")
        numero       = request.form.get("numero", "")
        complemento  = request.form.get("complemento", "")
        bairro       = request.form.get("bairro", "")
        cep          = request.form.get("cep", "")
        referencia   = request.form.get("referencia", "")
        cras         = request.form.get("cras", "")
        data_escuta  = request.form.get("data_escuta", "")
        renda_bruta  = float(request.form.get("renda_bruta", 0) or 0)
        renda_rpc    = float(request.form.get("renda_per_capita", 0) or 0)
        beneficios   = request.form.get("beneficios", "")
        vuls         = request.form.getlist("vulnerabilidade")
        servicos     = request.form.getlist("servicos_suas")
        parecer      = sanitizar_html(request.form.get("parecer", ""))
        excecao      = request.form.get("excecao_art64") == "1"
        visita_dom   = request.form.get("visita_domiciliar") == "1"
        valor_bf     = 0.0
        if beneficios == "Bolsa Família":
            try:
                vbf = request.form.get("valor_bolsa_familia", "0").replace('.','').replace(',','.')
                valor_bf = float(vbf or 0)
            except (ValueError, AttributeError) as e:
                logger.error(f"Edição #{id}: valor do Bolsa Família inválido ('{request.form.get('valor_bolsa_familia')}') por {current_user.id}: {e}")
                flash('❌ Valor do Bolsa Família inválido. Digite apenas números (ex: 600,00). Suas demais alterações não foram salvas — refaça a edição.', 'danger')
                conexao.close()
                return redirect(url_for('editar_solicitacao', id=id))

        membros_nomes    = request.form.getlist("membro_nome[]")
        membros_idades   = request.form.getlist("membro_idade[]")
        membros_vinculos = request.form.getlist("membro_vinculo[]")
        composicao = []
        for i in range(len(membros_nomes)):
            if membros_nomes[i].strip():
                composicao.append({
                    'nome': membros_nomes[i],
                    'idade': membros_idades[i] if i < len(membros_idades) else '',
                    'vinculo': membros_vinculos[i] if i < len(membros_vinculos) else ''
                })

        novos = {
            'nome': nome, 'data_nascimento': data_nasc,
            'telefone': telefone, 'email': email,
            'endereco': endereco, 'numero': numero,
            'complemento': complemento, 'bairro': bairro,
            'cep': cep, 'referencia': referencia, 'cras': cras,
            'data_escuta': data_escuta,
            'renda_bruta': str(renda_bruta), 'renda_per_capita': str(renda_rpc),
            'beneficios': beneficios,
            'vulnerabilidade': ", ".join(vuls),
            'servicos_suas': ", ".join(servicos),
            'parecer': parecer,
            'excecao_art64': str(excecao),
        }

        cursor.execute("""
            UPDATE solicitacoes SET
                nome=%s, data_nascimento=%s, telefone=%s, email=%s,
                endereco=%s, numero=%s, complemento=%s, bairro=%s,
                cep=%s, referencia=%s, cras=%s, data_escuta=%s,
                total_pessoas=%s, composicao_familiar=%s,
                renda_bruta=%s, renda_per_capita=%s, beneficios=%s,
                vulnerabilidade=%s, servicos_suas=%s, parecer=%s,
                excecao_art64=%s, visita_domiciliar=%s, valor_bolsa_familia=%s
            WHERE id=%s
        """, (
            nome, data_nasc, telefone, email,
            endereco, numero, complemento, bairro,
            cep, referencia, cras, data_escuta,
            len(composicao), json.dumps(composicao, ensure_ascii=False),
            renda_bruta, renda_rpc, beneficios,
            ", ".join(vuls), ", ".join(servicos), parecer,
            excecao, visita_dom, valor_bf, id
        ))

        # ── Registrar histórico de alterações ──
        agora = datetime.now(FUSO_RONDONIA).strftime('%d/%m/%Y %H:%M:%S')
        labels = {
            'nome': 'Nome', 'data_nascimento': 'Data de nascimento',
            'telefone': 'Telefone', 'email': 'E-mail',
            'endereco': 'Endereço', 'numero': 'Número',
            'complemento': 'Complemento', 'bairro': 'Bairro',
            'cep': 'CEP', 'referencia': 'Referência', 'cras': 'CRAS',
            'data_escuta': 'Data da escuta', 'renda_bruta': 'Renda bruta',
            'renda_per_capita': 'Renda per capita', 'beneficios': 'Benefícios',
            'vulnerabilidade': 'Vulnerabilidade', 'servicos_suas': 'Serviços SUAS',
            'parecer': 'Parecer técnico', 'excecao_art64': 'Exceção Art.64',
        }
        for campo, novo_val in novos.items():
            antigo = str(s.get(campo, '') or '')
            if antigo != str(novo_val or ''):
                cursor.execute("""
                    INSERT INTO historico_edicoes
                        (solicitacao_id, usuario, campo, valor_antes, valor_depois, data_hora)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, (id, current_user.id, labels.get(campo, campo), antigo, str(novo_val), agora))

        conexao.commit()
        conexao.close()
        logger.info(f"Solicitação {id} editada por {current_user.id}")
        flash('✅ Solicitação atualizada com sucesso!', 'success')
        return redirect(url_for('ver_solicitacao', id=id))

    conexao.close()
    bairros_por_cras = get_bairros_por_cras()
    return render_template("editar_solicitacao.html", s=s, current_user=current_user,
                           bairros_por_cras=bairros_por_cras)



@app.route("/gerar_pdf/<int:id>")
@login_required
def gerar_pdf_assinatura(id):
    conexao = get_db()
    cursor = conexao.cursor()
    cursor.execute("""
        SELECT 
            s.id,                       
            s.tecnico,                  
            s.nome,                     
            s.cpf,                      
            s.data_nascimento,          
            s.telefone,                 
            s.endereco,                 
            s.numero,                   
            s.bairro,                   
            s.cras,                     
            s.renda_bruta,              
            s.renda_per_capita,         
            s.parecer,                  
            s.status,                   
            s.data_escuta,              
            s.data_solicitacao,         
            u_escuta.nome as tecnico_nome,          
            s.data_entrega,             
            s.tecnico_entrega,          
            u_entrega.nome as tecnico_entrega_nome,
            s.composicao_familiar,      
            s.beneficios,               
            s.vulnerabilidade,          
            s.servicos_suas,            
            s.total_pessoas,            
            s.cep,                      -- ÍNDICE 25
            s.referencia,               -- ÍNDICE 26
            s.complemento,              -- ÍNDICE 27
            s.email,                    -- ÍNDICE 28
            s.excecao_art64,            -- ÍNDICE 29
            s.valor_bolsa_familia,      -- ÍNDICE 30
            s.visita_domiciliar         -- ÍNDICE 31
        FROM solicitacoes s
        LEFT JOIN usuarios u_escuta ON s.tecnico = u_escuta.usuario
        LEFT JOIN usuarios u_entrega ON s.tecnico_entrega = u_entrega.usuario
        WHERE s.id = %s
    """, (id,))
   
    s = cursor.fetchone()
    conexao.close()

    if not s:
        return "Solicitação não encontrada", 404

    # s[9] = cras da solicitação
    if not pode_acessar_solicitacao(s[9]):
        flash('❌ Você não tem permissão para acessar solicitações de outra unidade.', 'danger')
        return redirect(url_for('solicitacoes'))

    # Mapeamento claro dos índices (DOCUMENTAÇÃO)
    ID = 0
    TECNICO = 1
    NOME = 2
    CPF = 3
    DATA_NASC = 4
    TELEFONE = 5
    ENDERECO = 6
    NUMERO = 7
    BAIRRO = 8
    CRAS = 9
    RENDA_BRUTA = 10
    RENDA_PER_CAPITA = 11
    PARECER = 12
    STATUS = 13
    DATA_ESCUTA = 14
    DATA_SOLICITACAO = 15
    TECNICO_NOME = 16
    DATA_ENTREGA = 17
    TECNICO_ENTREGA = 18
    TECNICO_ENTREGA_NOME = 19
    COMPOSICAO_FAMILIAR = 20
    BENEFICIOS = 21
    VULNERABILIDADE = 22
    SERVICOS_SUAS = 23
    TOTAL_PESSOAS = 24
    CEP = 25
    REFERENCIA = 26
    COMPLEMENTO = 27
    EMAIL = 28
    EXCECAO_ART64 = 29
    VALOR_BOLSA_FAMILIA = 30
    VISITA_DOMICILIAR = 31
    
    # Descriptografar CPF
    cpf_pdf = formatar_cpf(descriptografar_cpf(s[CPF])) if s[CPF] else 'N/A'
    
    # Formatar data de nascimento
    data_nasc = s[DATA_NASC] if s[DATA_NASC] else 'N/A'
    if data_nasc != 'N/A' and '-' in str(data_nasc):
        try:
            partes = str(data_nasc).split('-')
            if len(partes) == 3:
                data_nasc = f"{partes[2]}/{partes[1]}/{partes[0]}"
        except Exception as e:
            logger.error(f"gerar_pdf #{id}: falha ao formatar data de nascimento '{data_nasc}': {e}")
    
    # Formatar data da escuta
    data_escuta = s[DATA_ESCUTA] if s[DATA_ESCUTA] else 'N/A'
    if data_escuta != 'N/A' and '-' in str(data_escuta):
        try:
            partes = str(data_escuta).split('-')
            if len(partes) == 3:
                data_escuta = f"{partes[2]}/{partes[1]}/{partes[0]}"
        except Exception as e:
            logger.error(f"gerar_pdf #{id}: falha ao formatar data da escuta '{data_escuta}': {e}")
    
    # Formatar data da entrega
    data_entrega_pdf = '___/___/_______'
    if s[DATA_ENTREGA]:
        data_entrega_pdf = str(s[DATA_ENTREGA])
        if '-' in data_entrega_pdf:
            try:
                partes = data_entrega_pdf.split('-')
                if len(partes) == 3:
                    data_entrega_pdf = f"{partes[2]}/{partes[1]}/{partes[0]}"
            except Exception as e:
                logger.error(f"gerar_pdf #{id}: falha ao formatar data de entrega '{data_entrega_pdf}': {e}")
    
    # Formatar CEP
    cep_valor = ''.join(filter(str.isdigit, str(s[CEP]))) if s[CEP] else ''
    if len(cep_valor) == 8:
        cep_formatado = f"{cep_valor[:5]}-{cep_valor[5:]}"
    elif cep_valor:
        cep_formatado = cep_valor
    else:
        cep_formatado = 'N/A'
    
    numero_controle = f"CB-{datetime.now(FUSO_RONDONIA).strftime('%Y%m%d')}-{s[ID]:04d}"
    data_geracao = datetime.now(FUSO_RONDONIA).strftime('%d/%m/%Y às %H:%M:%S')
    tecnico_escuta = s[TECNICO_NOME] if s[TECNICO_NOME] else s[TECNICO]

    # ─── Gerar QR Code ────────────────────────────────────────────
    base_url = os.environ.get('BASE_URL', request.host_url.rstrip('/'))
    qr_conteudo = (
        f"Sistema SEMASF - Ji-Paraná\n"
        f"Controle: {numero_controle}\n"
        f"Técnico: {tecnico_escuta}\n"
        f"Gerado em: {data_geracao}\n"
        f"URL: {base_url}/ver_solicitacao/{s[ID]}"
    )
    qr_img = qrcode.make(qr_conteudo)
    qr_buffer = io.BytesIO()
    qr_img.save(qr_buffer, format='PNG')
    qr_buffer.seek(0)
    qr_reader = ImageReader(qr_buffer)

    # ─── Estrutura de duas passagens para "Página X de Y" ─────────
    # Passagem 1: contar páginas
    # Passagem 2: desenhar com total de páginas conhecido
    # Usamos uma classe auxiliar que desenha o rodapé no onPage

    MARGEM_RODAPE = 3.8*cm   # altura reservada para rodapé (QR + textos)
    MARGEM_TOP    = 2*cm
    LIMITE_Y      = MARGEM_RODAPE  # conteúdo não desce abaixo disso

    def desenhar_rodape(canvas_obj, pagina_atual, total_paginas):
        """Desenha rodapé padronizado em qualquer página."""
        canvas_obj.saveState()
        w, h = A4

        # Linha separadora
        canvas_obj.setStrokeColorRGB(0.5, 0.5, 0.5)
        canvas_obj.setLineWidth(0.5)
        canvas_obj.line(2*cm, MARGEM_RODAPE - 0.1*cm, w - 2*cm, MARGEM_RODAPE - 0.1*cm)

        # QR Code (canto direito do rodapé)
        qr_size = 3.0*cm
        qr_x = w - 2*cm - qr_size
        qr_y = 0.5*cm
        canvas_obj.drawImage(qr_reader, qr_x, qr_y, width=qr_size, height=qr_size)
        canvas_obj.setFont("Helvetica", 5.5)
        canvas_obj.setFillColorRGB(0.4, 0.4, 0.4)
        canvas_obj.drawCentredString(qr_x + qr_size/2, qr_y - 0.3*cm, "Assinatura Digital")

        # Textos do rodapé (à esquerda do QR)
        tx = 2*cm
        canvas_obj.setFont("Helvetica-Bold", 7.5)
        canvas_obj.setFillColorRGB(0.15, 0.15, 0.15)
        canvas_obj.drawString(tx, MARGEM_RODAPE - 0.55*cm,
            "⚠  Este documento deve ser assinado pelo beneficiário e pelo técnico responsável pela entrega.")

        canvas_obj.setFont("Helvetica", 7.5)
        canvas_obj.setFillColorRGB(0.2, 0.2, 0.2)
        canvas_obj.drawString(tx, MARGEM_RODAPE - 1.0*cm,
            "A escuta, entrega e critérios de concessão obedeceram à Lei Municipal do SUAS nº 3.603 de 01/12/2022.")
        canvas_obj.drawString(tx, MARGEM_RODAPE - 1.4*cm,
            f"Documento gerado em {data_geracao}  |  Nº de Controle: {numero_controle}")

        # Numeração de páginas
        canvas_obj.setFont("Helvetica-Bold", 8)
        canvas_obj.setFillColorRGB(0.2, 0.2, 0.2)
        canvas_obj.drawString(tx, MARGEM_RODAPE - 1.85*cm,
            f"Página {pagina_atual} de {total_paginas}")

        canvas_obj.restoreState()

    def gerar_conteudo(canvas_obj, total_paginas_conhecido=None):
        """
        Desenha todo o conteúdo do PDF.
        Se total_paginas_conhecido for None, apenas conta as páginas (passagem 1).
        Caso contrário, desenha de verdade (passagem 2).
        """
        w, h = A4
        pagina_atual = 1
        desenhar = total_paginas_conhecido is not None

        def nova_pagina():
            nonlocal pagina_atual
            if desenhar:
                desenhar_rodape(canvas_obj, pagina_atual, total_paginas_conhecido)
            canvas_obj.showPage()
            pagina_atual += 1
            return h - MARGEM_TOP

        def check_space(needed, current_y):
            if current_y - needed < LIMITE_Y:
                return nova_pagina()
            return current_y

        y = h - MARGEM_TOP

        if desenhar:
            # ── CABEÇALHO ──────────────────────────────────────────
            logo_path = os.path.join(os.path.dirname(__file__), 'static', 'img', 'logo_prefeitura.png')
            # Logo no canto DIREITO — largura fixa 5.5cm para não sobrepor o texto
            logo_largura = 5.5*cm
            logo_altura  = 1.8*cm   # bounding box; aspect real ~4.17:1 → exibe ~5.5×1.32cm
            if os.path.exists(logo_path):
                try:
                    logo_reader = ImageReader(logo_path)
                    logo_x = w - 2*cm - logo_largura
                    canvas_obj.drawImage(logo_reader, logo_x, y - logo_altura + 0.4*cm,
                                         height=logo_altura, width=logo_largura,
                                         preserveAspectRatio=True, mask='auto')
                except Exception:
                    pass
            canvas_obj.setFont("Helvetica-Bold", 14)
            canvas_obj.drawString(2*cm, y, "PREFEITURA MUNICIPAL DE JI-PARANÁ")
            y -= 0.6*cm
            canvas_obj.setFont("Helvetica-Bold", 10)
            canvas_obj.drawString(2*cm, y, "Secretaria Municipal de Assistência Social e Família - SEMASF")
            y -= 0.6*cm
            canvas_obj.setFont("Helvetica", 9)
            canvas_obj.drawString(2*cm, y, "TERMO DE CONCESSÃO DE CESTA BÁSICA")
            y -= 0.4*cm
            canvas_obj.line(2*cm, y, w - 2*cm, y)
            y -= 0.6*cm
            canvas_obj.setFont("Helvetica-Bold", 10)
            canvas_obj.drawString(2*cm, y, f"Nº de Controle: {numero_controle}")
            canvas_obj.drawString(10*cm, y, f"Data de Emissão: {datetime.now(FUSO_RONDONIA).strftime('%d/%m/%Y %H:%M')}")
            y -= 0.8*cm

            # ── 1. DADOS DO BENEFICIÁRIO ────────────────────────────
            y = check_space(5*cm, y)
            canvas_obj.setFont("Helvetica-Bold", 11)
            canvas_obj.setFillColorRGB(0, 0.4, 0)
            canvas_obj.drawString(2*cm, y, "1. DADOS DO BENEFICIÁRIO")
            canvas_obj.setFillColorRGB(0, 0, 0)
            y -= 0.6*cm
            canvas_obj.setFont("Helvetica", 10)
            for label, valor in [
                ("Nome:", s[NOME] if s[NOME] else 'N/A'),
                ("CPF:", cpf_pdf),
                ("Data de Nascimento:", data_nasc),
                ("Telefone:", s[TELEFONE] if s[TELEFONE] else 'N/A'),
                ("Email:", s[EMAIL] if s[EMAIL] else 'N/A'),
            ]:
                canvas_obj.drawString(2.5*cm, y, f"{label} {valor}")
                y -= 0.45*cm
            y -= 0.2*cm

            # ── 2. ENDEREÇO ─────────────────────────────────────────
            y = check_space(7*cm, y)
            canvas_obj.setFont("Helvetica-Bold", 11)
            canvas_obj.setFillColorRGB(0, 0.4, 0)
            canvas_obj.drawString(2*cm, y, "2. ENDEREÇO")
            canvas_obj.setFillColorRGB(0, 0, 0)
            y -= 0.6*cm
            canvas_obj.setFont("Helvetica", 10)
            end_val = s[ENDERECO] if s[ENDERECO] else ''
            num_val = s[NUMERO] if s[NUMERO] else 'S/N'
            comp_val = s[COMPLEMENTO] if s[COMPLEMENTO] else ''
            end_completo = f"{end_val}, {num_val}" + (f" - {comp_val}" if comp_val else "")
            for label, valor in [
                ("Endereço:", end_completo[:60]),
                ("Bairro:", s[BAIRRO] if s[BAIRRO] else 'N/A'),
                ("CEP:", cep_formatado),
                ("Referência:", s[REFERENCIA][:50] if s[REFERENCIA] else 'N/A'),
                ("CRAS:", s[CRAS] if s[CRAS] else 'N/A'),
            ]:
                canvas_obj.drawString(2.5*cm, y, f"{label} {valor}")
                y -= 0.45*cm
            y -= 0.2*cm

            # ── 3. INFORMAÇÕES SOCIOECONÔMICAS ──────────────────────
            y = check_space(6*cm, y)
            canvas_obj.setFont("Helvetica-Bold", 11)
            canvas_obj.setFillColorRGB(0, 0.4, 0)
            canvas_obj.drawString(2*cm, y, "3. INFORMAÇÕES SOCIOECONÔMICAS")
            canvas_obj.setFillColorRGB(0, 0, 0)
            y -= 0.6*cm
            canvas_obj.setFont("Helvetica", 10)
            rb_val = s[RENDA_BRUTA] if s[RENDA_BRUTA] and float(s[RENDA_BRUTA]) > 0 else 0
            rpc_val = s[RENDA_PER_CAPITA] if s[RENDA_PER_CAPITA] and float(s[RENDA_PER_CAPITA]) > 0 else 0
            for label, valor in [
                ("Renda Bruta Familiar:", f"R$ {float(rb_val):.2f}" if rb_val > 0 else 'N/A'),
                ("Renda Per Capita:", f"R$ {float(rpc_val):.2f}" if rpc_val > 0 else 'N/A'),
                ("Benefícios:", ((s[BENEFICIOS] if s[BENEFICIOS] else 'Nenhum') + (f" — R$ {float(s[VALOR_BOLSA_FAMILIA]):.2f}" if s[BENEFICIOS] == 'Bolsa Família' and s[VALOR_BOLSA_FAMILIA] else ''))[:60]),
                ("Vulnerabilidades:", (s[VULNERABILIDADE] if s[VULNERABILIDADE] else 'Não informado')[:50]),
                ("Serviços SUAS:", (s[SERVICOS_SUAS] if s[SERVICOS_SUAS] else 'Não informado')[:50]),
            ]:
                canvas_obj.drawString(2.5*cm, y, f"{label} {valor}")
                y -= 0.45*cm
            y -= 0.2*cm

            # ── 4. COMPOSIÇÃO FAMILIAR ──────────────────────────────
            y = check_space(8*cm, y)
            canvas_obj.setFont("Helvetica-Bold", 11)
            canvas_obj.setFillColorRGB(0, 0.4, 0)
            canvas_obj.drawString(2*cm, y, "4. COMPOSIÇÃO FAMILIAR")
            canvas_obj.setFillColorRGB(0, 0, 0)
            y -= 0.6*cm
            canvas_obj.setFont("Helvetica", 10)
            total_pess = s[TOTAL_PESSOAS] if s[TOTAL_PESSOAS] and s[TOTAL_PESSOAS] > 0 else 1
            canvas_obj.drawString(2.5*cm, y, f"Total de pessoas: {total_pess}")
            y -= 0.5*cm
            comp_familiar = s[COMPOSICAO_FAMILIAR] if s[COMPOSICAO_FAMILIAR] else '[]'
            try:
                membros = json.loads(comp_familiar)
                if membros:
                    canvas_obj.setFont("Helvetica-Bold", 9)
                    canvas_obj.drawString(3*cm, y, "Nome")
                    canvas_obj.drawString(9*cm, y, "Idade")
                    canvas_obj.drawString(13*cm, y, "Vínculo")
                    y -= 0.4*cm
                    canvas_obj.setFont("Helvetica", 9)
                    for membro in membros[:8]:
                        y = check_space(1*cm, y)
                        canvas_obj.drawString(3*cm, y, membro.get('nome', '')[:25])
                        canvas_obj.drawString(9*cm, y, str(membro.get('idade', '')))
                        canvas_obj.drawString(13*cm, y, membro.get('vinculo', '')[:15])
                        y -= 0.35*cm
                else:
                    canvas_obj.drawString(2.5*cm, y, "Família unipessoal")
                    y -= 0.35*cm
            except Exception as e:
                logger.error(f"gerar_pdf #{id}: composição familiar corrompida ('{comp_familiar}'): {e}")
                canvas_obj.drawString(2.5*cm, y, "Não informado")
                y -= 0.35*cm
            y -= 0.2*cm

            # ── 5. REGISTRO DE ATENDIMENTO ──────────────────────────
            y = check_space(4*cm, y)
            canvas_obj.setFont("Helvetica-Bold", 11)
            canvas_obj.setFillColorRGB(0, 0.4, 0)
            canvas_obj.drawString(2*cm, y, "5. REGISTRO DE ATENDIMENTO")
            canvas_obj.setFillColorRGB(0, 0, 0)
            y -= 0.6*cm
            canvas_obj.setFont("Helvetica", 10)
            for label, valor in [
                ("Técnico que realizou a Escuta:", tecnico_escuta),
                ("Data da Escuta Técnica:", data_escuta),
                ("Status:", s[STATUS] if s[STATUS] else 'N/A'),
            ]:
                canvas_obj.drawString(2.5*cm, y, f"{label} {valor}")
                y -= 0.45*cm
            y -= 0.2*cm

            # ── 6. PARECER TÉCNICO ──────────────────────────────────
            y = check_space(5*cm, y)
            canvas_obj.setFont("Helvetica-Bold", 11)
            canvas_obj.setFillColorRGB(0, 0.4, 0)
            canvas_obj.drawString(2*cm, y, "6. PARECER TÉCNICO")
            canvas_obj.setFillColorRGB(0, 0, 0)
            y -= 0.6*cm

            # Aviso de exceção Art. 64
            if s[EXCECAO_ART64]:
                box_x = 2*cm
                box_largura = w - 4*cm  # mesma largura útil da página (margens de 2cm)
                box_altura = 1.15*cm
                box_topo = y
                # Moldura
                canvas_obj.setStrokeColorRGB(0.7, 0.1, 0.1)
                canvas_obj.setLineWidth(1)
                canvas_obj.rect(box_x, box_topo - box_altura, box_largura, box_altura, fill=0)
                # Texto dentro da moldura (com recuo interno)
                canvas_obj.setFillColorRGB(0.7, 0.1, 0.1)
                canvas_obj.setFont("Helvetica-Bold", 8)
                canvas_obj.drawString(box_x + 0.3*cm, box_topo - 0.4*cm,
                    "CONCESSÃO EXCEPCIONAL — Art. 64 da Lei Municipal nº 3.603/2022")
                canvas_obj.setFont("Helvetica", 7.5)
                canvas_obj.drawString(box_x + 0.3*cm, box_topo - 0.72*cm,
                    "Situação não contemplada nos critérios ordinários, autorizada mediante parecer técnico social")
                canvas_obj.drawString(box_x + 0.3*cm, box_topo - 1.0*cm,
                    "e autorização do gestor da SEMASF.")
                canvas_obj.setFillColorRGB(0, 0, 0)
                canvas_obj.setStrokeColorRGB(0, 0, 0)
                y -= (box_altura + 0.4*cm)

            # Aviso de visita domiciliar
            if s[VISITA_DOMICILIAR]:
                vd_x = 2*cm
                vd_largura = w - 4*cm
                vd_altura = 0.9*cm
                canvas_obj.setStrokeColorRGB(0.0, 0.45, 0.7)
                canvas_obj.setLineWidth(1)
                canvas_obj.rect(vd_x, y - vd_altura, vd_largura, vd_altura, fill=0)
                canvas_obj.setFillColorRGB(0.0, 0.45, 0.7)
                canvas_obj.setFont("Helvetica-Bold", 8)
                canvas_obj.drawString(vd_x + 0.3*cm, y - 0.38*cm,
                    "VISITA DOMICILIAR SOLICITADA — Averiguação das condições relatadas pelo beneficiário.")
                canvas_obj.setFont("Helvetica", 7.5)
                canvas_obj.drawString(vd_x + 0.3*cm, y - 0.68*cm,
                    "O técnico responsável identificou necessidade de verificação in loco antes da concessão.")
                canvas_obj.setFillColorRGB(0, 0, 0)
                canvas_obj.setStrokeColorRGB(0, 0, 0)
                y -= (vd_altura + 0.4*cm)

            parecer_txt = _strip_html(s[PARECER]) if s[PARECER] else 'Sem parecer técnico registrado.'
            text_object = canvas_obj.beginText(2.5*cm, y)
            text_object.setFont("Helvetica", 9)
            max_w = w - 5*cm

            for paragrafo in parecer_txt.split('\n'):
                paragrafo = paragrafo.strip()
                # Linha em branco entre parágrafos
                if not paragrafo:
                    text_object.textLine('')
                    continue
                linha = ""
                for palavra in paragrafo.split():
                    teste = linha + " " + palavra if linha else palavra
                    if canvas_obj.stringWidth(teste, "Helvetica", 9) <= max_w:
                        linha = teste
                    else:
                        text_object.textLine(linha)
                        linha = palavra
                        if text_object.getY() < LIMITE_Y:
                            canvas_obj.drawText(text_object)
                            y = nova_pagina()
                            text_object = canvas_obj.beginText(2.5*cm, y)
                            text_object.setFont("Helvetica", 9)
                if linha:
                    text_object.textLine(linha)
                # Espaço entre parágrafos
                text_object.textLine('')

            canvas_obj.drawText(text_object)
            y = text_object.getY() - 1*cm

            # ── 7. ASSINATURAS ──────────────────────────────────────
            if y < 8*cm + LIMITE_Y:
                y = nova_pagina()
            else:
                y -= 0.5*cm
            canvas_obj.setFont("Helvetica-Bold", 11)
            canvas_obj.setFillColorRGB(0, 0.4, 0)
            canvas_obj.drawString(2*cm, y, "7. ASSINATURAS")
            canvas_obj.setFillColorRGB(0, 0, 0)
            y -= 0.8*cm
            canvas_obj.setFont("Helvetica-Bold", 10)
            canvas_obj.drawString(2*cm, y, f"Data da Entrega: {data_entrega_pdf}")
            y -= 1.2*cm
            canvas_obj.setFont("Helvetica", 10)
            canvas_obj.line(2*cm, y, 9.5*cm, y)
            canvas_obj.drawString(2*cm, y - 0.4*cm, "Assinatura do Beneficiário")
            canvas_obj.line(10.5*cm, y, 19*cm, y)
            canvas_obj.drawString(10.5*cm, y - 0.4*cm, "Técnico Responsável pela Entrega")
            y -= 2.5*cm
            if y > LIMITE_Y + 3*cm:
                canvas_obj.rect(2*cm, y - 1.5*cm, 6*cm, 2*cm)
                canvas_obj.setFont("Helvetica", 8)
                canvas_obj.drawString(2.3*cm, y - 0.5*cm, "CARIMBO DO CRAS")
                canvas_obj.drawString(2.3*cm, y - 0.9*cm, s[CRAS] if s[CRAS] else '')

            # Rodapé da última página
            desenhar_rodape(canvas_obj, pagina_atual, total_paginas_conhecido)

        else:
            # ── Passagem 1: simulação apenas para contar páginas ────
            # Replica a mesma lógica de quebra de página sem desenhar
            y -= 3.5*cm  # cabeçalho
            y = check_space(5*cm, y)
            y -= 5 * 0.45*cm + 0.2*cm + 0.6*cm  # dados beneficiário

            y = check_space(7*cm, y)
            y -= 5 * 0.45*cm + 0.2*cm + 0.6*cm  # endereço

            y = check_space(6*cm, y)
            y -= 5 * 0.45*cm + 0.2*cm + 0.6*cm  # socioeconômico

            y = check_space(8*cm, y)
            y -= 0.6*cm + 0.5*cm  # composição - cabeçalho
            try:
                membros = json.loads(s[COMPOSICAO_FAMILIAR] if s[COMPOSICAO_FAMILIAR] else '[]')
                for membro in membros[:8]:
                    y = check_space(1*cm, y)
                    y -= 0.35*cm
            except Exception as e:
                logger.error(f"gerar_pdf #{id}: composição familiar corrompida na contagem de páginas: {e}")
                y -= 0.35*cm
            y -= 0.2*cm

            y = check_space(4*cm, y)
            y -= 3 * 0.45*cm + 0.2*cm + 0.6*cm  # atendimento

            y = check_space(5*cm, y)
            y -= 0.6*cm
            # simular quebra de página do parecer
            parecer_txt = _strip_html(s[PARECER]) if s[PARECER] else 'Sem parecer técnico registrado.'
            linhas_estimadas = max(1, len(parecer_txt) // 80)
            for _ in range(linhas_estimadas):
                if y - 0.4*cm < LIMITE_Y:
                    y = nova_pagina()
                y -= 0.4*cm
            y -= 1*cm

            if y < 8*cm + LIMITE_Y:
                y = nova_pagina()

        return pagina_atual

    # ── Passagem 1: contar páginas ─────────────────────────────────
    counter_buf = io.BytesIO()
    c_count = canvas.Canvas(counter_buf, pagesize=A4)
    total_paginas = gerar_conteudo(c_count, total_paginas_conhecido=None)

    # ── Passagem 2: gerar PDF real ─────────────────────────────────
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    gerar_conteudo(c, total_paginas_conhecido=total_paginas)
    c.save()
    buffer.seek(0)

    return send_file(buffer, as_attachment=True, download_name=f"termo_entrega_{numero_controle}.pdf", mimetype='application/pdf')

# =====================================================
# REGISTRAR ENTREGA
# =====================================================

@app.route("/registrar_entrega/<int:id>", methods=["POST"])
@login_required
def registrar_entrega(id):
    status = request.form.get("status_entrega")
    data = request.form.get("data_entrega")
    observacoes = request.form.get("observacoes", "").strip()
    if status not in ('Entregue', 'Ausente') or not data:
        flash("Preencha todos os campos!", "danger")
        return redirect(url_for("solicitacoes"))
    conexao = get_db()
    cursor = conexao.cursor()

    cursor.execute("SELECT status, cras FROM solicitacoes WHERE id=%s", (id,))
    row = cursor.fetchone()
    if not row:
        conexao.close()
        flash("Solicitação não encontrada.", "danger")
        return redirect(url_for("solicitacoes"))
    status_atual, cras_solic = row

    if not pode_acessar_solicitacao(cras_solic):
        conexao.close()
        flash("❌ Você não tem permissão para registrar entrega de solicitações de outra unidade.", "danger")
        return redirect(url_for("solicitacoes"))

    # Evita registrar entrega em solicitação cancelada ou sobrescrever
    # silenciosamente uma entrega já registrada
    if status_atual != 'Cadastrada':
        conexao.close()
        flash(f"❌ Só é possível registrar entrega de solicitações com status 'Cadastrada'. Status atual: {status_atual}.", "danger")
        return redirect(url_for("solicitacoes"))

    cursor.execute(
        "UPDATE solicitacoes SET status=%s, data_entrega=%s, tecnico_entrega=%s, observacoes_entrega=%s WHERE id=%s",
        (status, data, current_user.id, observacoes or None, id)
    )
    cursor.execute("""
        INSERT INTO historico_edicoes (solicitacao_id, usuario, campo, valor_antes, valor_depois, data_hora)
        VALUES (%s, %s, 'status', 'Cadastrada', %s, %s)
    """, (id, current_user.id, status, datetime.now(FUSO_RONDONIA).strftime("%d/%m/%Y %H:%M:%S")))
    conexao.commit()
    conexao.close()
    logger.info(f"Entrega: ID={id}, Status={status}")
    flash('✅ Registrado!' if status == 'Entregue' else '❌ Ausência registrada.', 'success')
    return redirect(url_for("solicitacoes"))

# =====================================================
# NOVA TENTATIVA DE ENTREGA
# =====================================================

@app.route("/nova_tentativa/<int:id>", methods=["POST"])
@login_required
def nova_tentativa(id):
    """Reseta status para Cadastrada e incrementa contador de tentativas."""
    conexao = get_db()
    cursor  = conexao.cursor()
    cursor.execute("SELECT status, num_tentativas, cras FROM solicitacoes WHERE id = %s", (id,))
    row = cursor.fetchone()
    if not row or row[0] != 'Ausente':
        conexao.close()
        flash("Solicitação não encontrada ou não está com status Ausente.", "danger")
        return redirect(url_for("solicitacoes"))

    if not pode_acessar_solicitacao(row[2]):
        conexao.close()
        flash("❌ Você não tem permissão para reabrir solicitações de outra unidade.", "danger")
        return redirect(url_for("solicitacoes"))

    tentativa_atual = (row[1] or 1) + 1
    data_hora = datetime.now(FUSO_RONDONIA).strftime("%d/%m/%Y %H:%M:%S")

    cursor.execute("""
        UPDATE solicitacoes
        SET status = 'Cadastrada',
            data_entrega = NULL,
            tecnico_entrega = NULL,
            observacoes_entrega = NULL,
            num_tentativas = %s
        WHERE id = %s
    """, (tentativa_atual, id))

    cursor.execute("""
        INSERT INTO historico_edicoes (solicitacao_id, usuario, campo, valor_antes, valor_depois, data_hora)
        VALUES (%s, %s, 'status', 'Ausente', %s, %s)
    """, (id, current_user.id, f'Cadastrada (tentativa {tentativa_atual})', data_hora))

    conexao.commit()
    conexao.close()

    logger.info(f"Nova tentativa: ID={id}, tentativa={tentativa_atual}, tecnico={current_user.id}")
    flash(f"✅ Solicitação #{id} reaberta para nova tentativa de entrega (tentativa {tentativa_atual}).", "success")
    return redirect(url_for("ver_solicitacao", id=id))


# =====================================================
# LISTA DE ENTREGA — o técnico define quantas cestas tem
# no dia, o sistema sugere as solicitações mais antigas,
# ele ajusta a seleção e gera o PDF para carimbo/assinatura
# =====================================================

def _filtro_unidade_entrega(unidade):
    """Retorna (join, condicao, params) para filtrar as solicitações pela
    unidade RESPONSÁVEL PELA ENTREGA.

    - CRAS X ........: escutas feitas pela equipe do CRAS X, MAIS as
                       escutas feitas pela Equipe Volante para famílias de
                       território URBANO do CRAS X (a Volante raramente
                       escuta, mas quando o faz para um caso urbano, a
                       entrega é do CRAS de referência, não da Volante).
                       Exclui sempre bairros da Equipe Volante (área rural).
                       Se a escuta foi registrada por admin/gestor ou
                       usuário excluído, usa o território como fallback.
    - CREAS .........: escutas feitas por técnicos do CREAS
                       (CRAS e CREAS atendem o mesmo território)
    - EQUIPE VOLANTE : bairros marcados como entrega da volante (área
                       rural) + outros municípios, independente de quem
                       registrou a escuta — inclusive quando é o próprio
                       CRAS que escuta e encaminha a família rural
    """
    if not unidade:
        return "", "", []
    if unidade == 'CREAS':
        return ("JOIN usuarios u_fil ON s.tecnico = u_fil.usuario",
                "u_fil.perfil = 'creas'", [])
    if unidade == 'EQUIPE VOLANTE':
        return ("",
                "(s.cras = 'EQUIPE VOLANTE' OR COALESCE(s.bairro, '') IN "
                "(SELECT bairro FROM cras_bairros WHERE entrega_volante = TRUE))",
                [])
    # COALESCE no bairro é necessário porque "X NOT IN (subconsulta)" com X
    # NULL resulta em NULL (não TRUE) em SQL — sem isso, uma solicitação com
    # bairro vazio/nulo (dado legado) sumiria silenciosamente de toda lista
    # de entrega, mesmo tendo sido escutada normalmente pela equipe do CRAS.
    return ("LEFT JOIN usuarios u_fil ON s.tecnico = u_fil.usuario",
            "COALESCE(s.bairro, '') NOT IN (SELECT bairro FROM cras_bairros WHERE entrega_volante = TRUE)"
            " AND ((u_fil.perfil = 'cras' AND u_fil.cras = %s)"
            "      OR (s.cras = %s AND COALESCE(u_fil.perfil, '') NOT IN ('cras', 'creas')))",
            [unidade, unidade])


@app.route("/lista_entrega")
@login_required
def lista_entrega():
    qtd = request.args.get('qtd', type=int)
    unidade = request.args.get('cras', '').strip()
    # Cada perfil tem sua unidade de entrega natural
    if current_user.perfil == 'cras':
        unidade = current_user.cras or ''
    elif current_user.perfil == 'creas' and not unidade:
        unidade = 'CREAS'
    elif current_user.perfil == 'cras_volante' and not unidade:
        unidade = 'EQUIPE VOLANTE'
    if unidade == 'ADMINISTRAÇÃO':
        unidade = ''

    sugestao = None
    if qtd and qtd > 0:
        qtd = min(qtd, 100)
        join, cond, params = _filtro_unidade_entrega(unidade)
        where = "s.status = 'Cadastrada'" + (f" AND {cond}" if cond else "")
        conexao = get_db()
        cursor  = conexao.cursor()
        # Mais antigas primeiro (fila por ordem de chegada)
        cursor.execute(f"""
            SELECT s.id, s.nome, s.cpf, s.bairro, s.data_solicitacao
            FROM solicitacoes s {join}
            WHERE {where}
            ORDER BY s.id ASC
            LIMIT %s
        """, params + [qtd])
        rows = cursor.fetchall()
        conexao.close()
        sugestao = [{
            'id': r[0],
            'nome': r[1] or '—',
            'cpf': formatar_cpf(descriptografar_cpf(r[2])) if r[2] else '—',
            'bairro': r[3] or '—',
            'data': str(r[4])[:10] if r[4] else '—',
        } for r in rows]

    pode_escolher_unidade = current_user.perfil in ['admin', 'gestor', 'creas', 'cras_volante']
    lista_unidades = (get_lista_cras() + ['CREAS', 'EQUIPE VOLANTE']) if pode_escolher_unidade else []
    return render_template("lista_entrega.html",
                           qtd=qtd,
                           cras_filtro=unidade,
                           sugestao=sugestao,
                           pode_escolher_unidade=pode_escolher_unidade,
                           lista_unidades=lista_unidades,
                           current_user=current_user)


@app.route("/api/lista_entrega/buscar_cpf/<cpf>")
@login_required
def api_lista_entrega_buscar_cpf(cpf):
    """Busca solicitações Cadastradas ou já Entregues de um CPF para incluir na lista de entrega."""
    cpf_limpo = ''.join(filter(str.isdigit, str(cpf)))
    if len(cpf_limpo) != 11 or not validar_cpf(cpf_limpo):
        return jsonify({'erro': 'CPF inválido', 'encontrados': []})
    filtros = ["s.cpf_hash = %s", "s.status IN ('Cadastrada', 'Entregue')"]
    params  = [hash_cpf(cpf_limpo)]
    join    = ""
    # Técnico de CRAS: escutas da própria equipe ou do próprio território
    if current_user.perfil == 'cras':
        join = "LEFT JOIN usuarios u_fil ON s.tecnico = u_fil.usuario"
        filtros.append("((u_fil.perfil = 'cras' AND u_fil.cras = %s) OR s.cras = %s)")
        params.extend([current_user.cras or '', current_user.cras or ''])
    conexao = get_db()
    cursor  = conexao.cursor()
    cursor.execute(f"""
        SELECT s.id, s.nome, s.cpf, s.bairro, s.data_solicitacao, s.status
        FROM solicitacoes s {join}
        WHERE {' AND '.join(filtros)}
        ORDER BY s.id ASC
    """, params)
    rows = cursor.fetchall()
    conexao.close()
    encontrados = [{
        'id': r[0],
        'nome': r[1] or '—',
        'cpf': formatar_cpf(descriptografar_cpf(r[2])) if r[2] else '—',
        'bairro': r[3] or '—',
        'data': str(r[4])[:10] if r[4] else '—',
        'entregue': r[5] == 'Entregue',
    } for r in rows]
    return jsonify({'encontrados': encontrados})


class _CanvasNumerado(canvas.Canvas):
    """Canvas que numera "Página X de Y" em duas passagens — reaproveitado
    pelos PDFs de Lista de Entrega e de Solicitações."""
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._estados = []

    def showPage(self):
        self._estados.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        total = len(self._estados)
        for estado in self._estados:
            self.__dict__.update(estado)
            largura = self._pagesize[0]
            self.setFont("Helvetica", 7.5)
            self.setFillColorRGB(0.2, 0.2, 0.2)
            self.drawRightString(largura - 1*cm, 0.8*cm, f"Página {self._pageNumber} de {total}")
            super().showPage()
        super().save()


@app.route("/lista_entrega/pdf")
@login_required
def lista_entrega_pdf():
    # Somente as solicitações selecionadas na tela de montagem
    ids_param = request.args.get('ids', '').strip()
    ids = [int(x) for x in ids_param.split(',') if x.strip().isdigit()][:100]
    if not ids:
        flash("Monte a lista de entrega antes de gerar o PDF.", "warning")
        return redirect(url_for("lista_entrega"))

    filtros = ["s.status IN ('Cadastrada', 'Entregue')", "s.id = ANY(%s)"]
    params  = [ids]
    join    = ""
    # Técnico de CRAS: escutas da própria equipe ou do próprio território
    if current_user.perfil == 'cras':
        join = "LEFT JOIN usuarios u_fil ON s.tecnico = u_fil.usuario"
        filtros.append("((u_fil.perfil = 'cras' AND u_fil.cras = %s) OR s.cras = %s)")
        params.extend([current_user.cras or '', current_user.cras or ''])

    conexao = get_db()
    cursor  = conexao.cursor()
    cursor.execute(f"""
        SELECT s.nome, s.cpf, s.endereco, s.numero, s.bairro, s.telefone,
               s.renda_per_capita, s.visita_domiciliar, s.cras, s.status
        FROM solicitacoes s {join}
        WHERE {' AND '.join(filtros)}
        ORDER BY s.bairro, s.nome
    """, params)
    linhas = cursor.fetchall()
    conexao.close()

    if not linhas:
        flash("Nenhuma solicitação válida (status Cadastrada ou Entregue) entre as selecionadas.", "warning")
        return redirect(url_for("lista_entrega"))

    hoje = datetime.now(FUSO_RONDONIA)
    # Rótulo derivado sempre dos dados reais retornados pela consulta —
    # nunca de um parâmetro de URL, já que este é um documento oficial
    # (carimbado e assinado) e um rótulo vindo da querystring poderia ser
    # manipulado para exibir uma unidade diferente da real.
    unidades = sorted({r[8] for r in linhas if r[8]})
    unidade_label = unidades[0] if len(unidades) == 1 else "DIVERSAS UNIDADES"

    # ── Cabeçalho e rodapé desenhados em toda página ──────────────
    PAGE_W, PAGE_H = landscape(A4)

    def _cabecalho_lista(canvas_obj, doc):
        canvas_obj.saveState()
        logo_path = os.path.join(os.path.dirname(__file__), 'static', 'img', 'logo_prefeitura.png')
        if os.path.exists(logo_path):
            try:
                canvas_obj.drawImage(ImageReader(logo_path), 1*cm, PAGE_H - 2.2*cm,
                                     width=4.5*cm, height=1.5*cm,
                                     preserveAspectRatio=True, mask='auto')
            except Exception:
                pass
        canvas_obj.setFont("Helvetica-Bold", 12)
        canvas_obj.drawCentredString(PAGE_W/2, PAGE_H - 1.1*cm, "PREFEITURA MUNICIPAL DE JI-PARANÁ")
        canvas_obj.setFont("Helvetica", 9)
        canvas_obj.drawCentredString(PAGE_W/2, PAGE_H - 1.55*cm,
                                     "Secretaria Municipal de Assistência Social e Família - SEMASF")
        canvas_obj.setFont("Helvetica-Bold", 11)
        canvas_obj.drawCentredString(PAGE_W/2, PAGE_H - 2.15*cm,
                                     "LISTA DE ENTREGA DO BENEFÍCIO EVENTUAL EMERGENCIAL (CESTA BÁSICA)")
        canvas_obj.setFont("Helvetica-Bold", 9)
        canvas_obj.drawCentredString(PAGE_W/2, PAGE_H - 2.6*cm,
                                     f"DATA: {hoje.strftime('%d/%m/%Y')}   |   UNIDADE: {unidade_label}   |   CESTAS: {len(linhas)}")
        canvas_obj.setLineWidth(0.8)
        canvas_obj.line(1*cm, PAGE_H - 2.8*cm, PAGE_W - 1*cm, PAGE_H - 2.8*cm)

        # Rodapé institucional
        canvas_obj.setFont("Helvetica", 6.5)
        canvas_obj.setFillColorRGB(0.35, 0.35, 0.35)
        canvas_obj.drawString(1*cm, 0.8*cm,
            f"Documento gerado pelo Sistema de Cestas Básicas SEMASF em {hoje.strftime('%d/%m/%Y às %H:%M')} (horário de Rondônia)")
        canvas_obj.restoreState()

    # ── Montagem da tabela ─────────────────────────────────────────
    estilo_celula = ParagraphStyle('celula', fontName='Helvetica', fontSize=7.5, leading=9)
    estilo_check  = ParagraphStyle('check',  fontName='Helvetica', fontSize=7.5, leading=10, alignment=1)

    cabecalho = ['ORD', 'Nome do(a) RF', 'CPF', 'Endereço', 'Bairro',
                 'Telefone', 'Renda Per Capita', 'Entrega', 'Sol. Visita']
    dados_tabela = [cabecalho]

    for i, (nome, cpf, endereco, numero, bairro, telefone, rpc, visita, _cras, status) in enumerate(linhas, 1):
        cpf_fmt = formatar_cpf(descriptografar_cpf(cpf)) if cpf else '—'
        end_fmt = f"{endereco or ''}, nº {numero}" if numero else (endereco or '—')
        tel_fmt = telefone or 'Não possui'
        rpc_fmt = f"R$ {float(rpc):.2f}".replace('.', ',') if rpc is not None else 'R$ 0,00'
        # Entrega: em branco, para marcar à caneta no ato da entrega — exceto
        # quando a solicitação já está registrada como Entregue no sistema,
        # caso em que o campo vem pré-marcado (não é para marcar de novo)
        if status == 'Entregue':
            chk_entrega = Paragraph('(X) Sim<br/>(&nbsp;&nbsp;) Não', estilo_check)
        else:
            chk_entrega = Paragraph('(&nbsp;&nbsp;) Sim<br/>(&nbsp;&nbsp;) Não', estilo_check)
        # Sol. Visita: pré-marcada conforme registrado no sistema
        if visita:
            chk_visita = Paragraph('(X) Sim<br/>(&nbsp;&nbsp;) Não', estilo_check)
        else:
            chk_visita = Paragraph('(&nbsp;&nbsp;) Sim<br/>(X) Não', estilo_check)
        dados_tabela.append([
            f"{i:02d}",
            Paragraph((nome or '—'), estilo_celula),
            cpf_fmt,
            Paragraph(end_fmt, estilo_celula),
            Paragraph(bairro or '—', estilo_celula),
            tel_fmt,
            rpc_fmt,
            chk_entrega,
            chk_visita,
        ])

    larguras = [1.1*cm, 5.6*cm, 3.0*cm, 5.8*cm, 3.2*cm, 3.0*cm, 2.4*cm, 1.9*cm, 1.9*cm]
    tabela = Table(dados_tabela, colWidths=larguras, repeatRows=1)
    tabela.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0), colors.HexColor('#1B2F5E')),
        ('TEXTCOLOR',     (0, 0), (-1, 0), colors.white),
        ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, 0), 7.5),
        ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',      (0, 1), (-1, -1), 7.5),
        ('ALIGN',         (0, 0), (0, -1), 'CENTER'),
        ('ALIGN',         (2, 1), (2, -1), 'CENTER'),
        ('ALIGN',         (5, 1), (6, -1), 'CENTER'),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID',          (0, 0), (-1, -1), 0.5, colors.black),
        ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F4F8')]),
        ('TOPPADDING',    (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))

    estilo_assinatura = ParagraphStyle('assinatura', fontName='Helvetica', fontSize=9, leading=14)
    assinaturas = [
        Spacer(1, 1.4*cm),
        Paragraph("Técnico de Referência Responsável pela entrega: ________________________________________________",
                  estilo_assinatura),
        Spacer(1, 0.9*cm),
        Paragraph("Motorista Responsável pelo recebimento das escutas: _____________________________________________",
                  estilo_assinatura),
    ]

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=1*cm, rightMargin=1*cm,
        topMargin=3.1*cm, bottomMargin=1.4*cm,
    )
    doc.build([tabela] + assinaturas,
              onFirstPage=_cabecalho_lista, onLaterPages=_cabecalho_lista,
              canvasmaker=_CanvasNumerado)
    buffer.seek(0)

    slug_unidade = unidade_label.lower().replace(' ', '_')
    nome_arquivo = f"lista_entrega_{slug_unidade}_{hoje.strftime('%Y%m%d')}.pdf"
    logger.info(f"Lista de entrega gerada por {current_user.id} ({unidade_label}, {len(linhas)} beneficiários)")
    return send_file(buffer, as_attachment=True, download_name=nome_arquivo, mimetype='application/pdf')


# =====================================================
# FOTOS DAS ATIVIDADES (evidências para o Quadrimestral)
# =====================================================

def _url_foto(path):
    if not path:
        return None
    try:
        return storage_fotos.url_publica(path)
    except storage_fotos.StorageNaoConfigurado:
        return None

def _pode_gerenciar_atividade(criado_por):
    return current_user.perfil in ('admin', 'gestor') or criado_por == current_user.id

def _quadrimestre_da_data(data_str):
    """(ano, quadrimestre 1-3) a partir de uma data 'YYYY-MM-DD'. Convenção da
    audiência pública: 1º=jan-abr, 2º=mai-ago, 3º=set-dez."""
    try:
        ano, mes, _ = data_str.split('-')
        mes = int(mes)
    except (ValueError, AttributeError):
        return None, None
    quadrimestre = 1 if mes <= 4 else (2 if mes <= 8 else 3)
    return int(ano), quadrimestre

def _periodo_do_quadrimestre(ano, quadrimestre):
    faixas = {1: ('01-01', '04-30'), 2: ('05-01', '08-31'), 3: ('09-01', '12-31')}
    inicio, fim = faixas.get(quadrimestre, faixas[1])
    return f"{ano}-{inicio}", f"{ano}-{fim}"

def _nome_arquivo_foto(servico, data_atividade, titulo, foto_id):
    base = f"{servico or 'sem-servico'}_{data_atividade or ''}_{titulo or ''}"
    base = _re.sub(r'[^A-Za-z0-9._-]+', '-', base).strip('-')
    return f"{base}_{foto_id}.jpg"

def _pode_acessar_atividades():
    """Admin/gestor sempre têm acesso; os demais perfis (cras/creas/
    cras_volante) só se tiverem a permissão extra 'acesso_atividades'
    marcada — mantendo intactos os privilégios normais de solicitações."""
    return current_user.perfil in ('admin', 'gestor') or current_user.acesso_atividades

@app.route("/atividades")
@login_required
def atividades():
    if not _pode_acessar_atividades():
        flash('❌ Você não tem acesso a essa área.', 'danger')
        return redirect(url_for(pagina_inicial(current_user.perfil)))

    periodo_inicio = request.args.get('periodo_inicio', '').strip()
    periodo_fim    = request.args.get('periodo_fim', '').strip()
    busca_servico  = request.args.get('busca_servico', '').strip()
    quadrimestre   = request.args.get('quadrimestre', '').strip()
    ano            = request.args.get('ano', '').strip()
    ver_tudo       = request.args.get('tudo', '').strip() == '1'

    hoje = datetime.now(FUSO_RONDONIA)
    ano_atual, quad_atual = _quadrimestre_da_data(hoje.strftime('%Y-%m-%d'))

    if quadrimestre and ano:
        # Atalho de quadrimestre tem prioridade sobre datas manuais
        try:
            periodo_inicio, periodo_fim = _periodo_do_quadrimestre(int(ano), int(quadrimestre))
        except ValueError:
            quadrimestre = ''
    elif not periodo_inicio and not periodo_fim and not ver_tudo:
        # Nenhum filtro informado (e não pediram "ver tudo"): cai no
        # quadrimestre corrente por padrão
        quadrimestre, ano = str(quad_atual), str(ano_atual)
        periodo_inicio, periodo_fim = _periodo_do_quadrimestre(ano_atual, quad_atual)

    filtros_periodo = []
    params_periodo  = []
    if periodo_inicio:
        filtros_periodo.append("a.data_atividade >= %s")
        params_periodo.append(periodo_inicio)
    if periodo_fim:
        filtros_periodo.append("a.data_atividade <= %s")
        params_periodo.append(periodo_fim)
    where_periodo = ("WHERE " + " AND ".join(filtros_periodo)) if filtros_periodo else ""

    filtros_lista = list(filtros_periodo)
    params_lista  = list(params_periodo)
    if busca_servico:
        filtros_lista.append("a.servico = %s")
        params_lista.append(busca_servico)
    where_lista = ("WHERE " + " AND ".join(filtros_lista)) if filtros_lista else ""

    conexao = get_db()
    cursor = conexao.cursor()
    cursor.execute(f"""
        SELECT a.id, a.titulo, a.data_atividade, a.servico, a.criado_por,
               COALESCE(u.nome, a.criado_por),
               (SELECT COUNT(*) FROM fotos_atividade f WHERE f.atividade_id = a.id),
               (SELECT f.storage_path FROM fotos_atividade f
                WHERE f.atividade_id = a.id ORDER BY f.ordem, f.id LIMIT 1)
        FROM atividades_fotos a
        LEFT JOIN usuarios u ON a.criado_por = u.usuario
        {where_lista}
        ORDER BY a.data_atividade DESC, a.id DESC
    """, params_lista)
    linhas = cursor.fetchall()

    # Contagem de fotos por serviço dentro do PERÍODO (sem aplicar o filtro
    # de serviço) — é o que alimenta o painel de resumo abaixo.
    cursor.execute(f"""
        SELECT a.servico, COUNT(f.id)
        FROM atividades_fotos a
        LEFT JOIN fotos_atividade f ON f.atividade_id = a.id
        {where_periodo}
        GROUP BY a.servico
    """, params_periodo)
    contagem_servico = {servico: total for servico, total in cursor.fetchall()}
    conexao.close()

    lista_atividades = []
    for (aid, titulo, data_ativ, servico, criado_por, criado_por_nome, num_fotos, capa) in linhas:
        lista_atividades.append({
            'id': aid, 'titulo': titulo, 'data_atividade': data_ativ,
            'servico': servico, 'criado_por': criado_por, 'criado_por_nome': criado_por_nome,
            'num_fotos': num_fotos, 'capa_url': _url_foto(capa),
            'pode_gerenciar': _pode_gerenciar_atividade(criado_por),
        })

    lista_unidades = get_lista_cras() + ['CREAS', 'EQUIPE VOLANTE'] + get_lista_servicos()
    resumo_servicos = [
        {'servico': u, 'num_fotos': contagem_servico.get(u, 0)}
        for u in lista_unidades
    ]
    qtd_servicos_enviaram = sum(1 for r in resumo_servicos if r['num_fotos'] > 0)

    return render_template(
        "atividades.html",
        atividades=lista_atividades,
        lista_unidades=lista_unidades,
        periodo_inicio=periodo_inicio,
        periodo_fim=periodo_fim,
        busca_servico=busca_servico,
        storage_configurado=storage_fotos.esta_configurado(),
        quadrimestre=quadrimestre,
        ano=ano,
        anos_disponiveis=list(range(ano_atual - 1, ano_atual + 2)),
        resumo_servicos=resumo_servicos,
        qtd_servicos_enviaram=qtd_servicos_enviaram,
        qtd_servicos_total=len(lista_unidades),
        total_fotos_periodo=sum(contagem_servico.values()),
    )

@app.route("/atividades/nova", methods=["GET", "POST"])
@login_required
def nova_atividade():
    if not _pode_acessar_atividades():
        flash('❌ Você não tem acesso a essa área.', 'danger')
        return redirect(url_for(pagina_inicial(current_user.perfil)))

    lista_unidades = get_lista_cras() + ['CREAS', 'EQUIPE VOLANTE'] + get_lista_servicos()

    if request.method == "POST":
        titulo = request.form.get("titulo", "").strip()
        data_atividade = request.form.get("data_atividade", "").strip()
        servico = request.form.get("servico", "").strip()
        descricao = request.form.get("descricao", "").strip()
        # Pareia arquivo+legenda pelo índice da LINHA antes de filtrar — se
        # filtrássemos os arquivos vazios primeiro, uma linha sem foto (mas
        # com legenda preenchida) desalinharia as legendas das linhas seguintes.
        arquivos_brutos = request.files.getlist("fotos[]")
        legendas_brutas = request.form.getlist("legenda[]")
        pares = [(f, legendas_brutas[i] if i < len(legendas_brutas) else '')
                 for i, f in enumerate(arquivos_brutos) if f and f.filename]

        if not titulo or not data_atividade:
            flash('❌ Preencha o título e a data da atividade.', 'danger')
            return render_template("nova_atividade.html", lista_unidades=lista_unidades)
        if not pares:
            flash('❌ Envie pelo menos uma foto.', 'danger')
            return render_template("nova_atividade.html", lista_unidades=lista_unidades)
        if not storage_fotos.esta_configurado():
            flash('❌ Envio de fotos ainda não está configurado. Contate o administrador.', 'danger')
            return render_template("nova_atividade.html", lista_unidades=lista_unidades)

        # Comprime tudo antes de gravar qualquer coisa no banco, para não
        # deixar uma atividade "pela metade" se um dos arquivos for inválido.
        try:
            comprimidas = [(storage_fotos.comprimir_imagem(f), legenda) for f, legenda in pares]
        except storage_fotos.ImagemInvalida as e:
            flash(f'❌ {e}', 'danger')
            return render_template("nova_atividade.html", lista_unidades=lista_unidades)

        agora = datetime.now(FUSO_RONDONIA).strftime("%d/%m/%Y %H:%M:%S")
        conexao = get_db()
        cursor = conexao.cursor()
        cursor.execute("""
            INSERT INTO atividades_fotos (titulo, data_atividade, servico, descricao, criado_por, criado_em)
            VALUES (%s,%s,%s,%s,%s,%s) RETURNING id
        """, (titulo, data_atividade, servico, descricao, current_user.id, agora))
        atividade_id = cursor.fetchone()[0]

        falhas = 0
        for i, (dados_bytes, legenda) in enumerate(comprimidas):
            path = storage_fotos.gerar_caminho(atividade_id)
            try:
                storage_fotos.upload_foto(dados_bytes, path)
                cursor.execute("""
                    INSERT INTO fotos_atividade (atividade_id, storage_path, legenda, ordem, criado_em)
                    VALUES (%s,%s,%s,%s,%s)
                """, (atividade_id, path, legenda, i, agora))
            except Exception as e:
                falhas += 1
                msg = f"nova_atividade: falha ao enviar foto {i} da atividade {atividade_id}: {type(e).__name__}: {e}"
                logger.error(msg)
                print(msg)  # também visível no console/log do Render (logger só grava em arquivo local)

        conexao.commit()
        conexao.close()
        logger.info(f"Atividade cadastrada: '{titulo}' por {current_user.id} ({len(comprimidas) - falhas} foto(s))")

        if falhas:
            flash(f'⚠️ Atividade criada, mas {falhas} foto(s) falharam ao enviar. Tente adicioná-las novamente na edição.', 'warning')
        else:
            flash('✅ Atividade registrada!', 'success')
        return redirect(url_for("ver_atividade", id=atividade_id))

    return render_template("nova_atividade.html", lista_unidades=lista_unidades)

@app.route("/atividades/<int:id>")
@login_required
def ver_atividade(id):
    if not _pode_acessar_atividades():
        flash('❌ Você não tem acesso a essa área.', 'danger')
        return redirect(url_for(pagina_inicial(current_user.perfil)))

    conexao = get_db()
    cursor = conexao.cursor()
    cursor.execute("""
        SELECT a.id, a.titulo, a.data_atividade, a.servico, a.descricao,
               a.criado_por, COALESCE(u.nome, a.criado_por), a.criado_em
        FROM atividades_fotos a
        LEFT JOIN usuarios u ON a.criado_por = u.usuario
        WHERE a.id = %s
    """, (id,))
    row = cursor.fetchone()
    if not row:
        conexao.close()
        return "Atividade não encontrada", 404

    cursor.execute("""
        SELECT id, storage_path, legenda FROM fotos_atividade
        WHERE atividade_id = %s ORDER BY ordem, id
    """, (id,))
    fotos = [{'id': fid, 'url': _url_foto(path), 'legenda': legenda}
             for (fid, path, legenda) in cursor.fetchall()]
    conexao.close()

    atividade = {
        'id': row[0], 'titulo': row[1], 'data_atividade': row[2], 'servico': row[3],
        'descricao': row[4], 'criado_por': row[5], 'criado_por_nome': row[6], 'criado_em': row[7],
    }

    return render_template(
        "ver_atividade.html",
        atividade=atividade,
        fotos=fotos,
        pode_gerenciar=_pode_gerenciar_atividade(atividade['criado_por']),
    )

@app.route("/atividades/<int:id>/editar", methods=["GET", "POST"])
@login_required
def editar_atividade(id):
    if not _pode_acessar_atividades():
        flash('❌ Você não tem acesso a essa área.', 'danger')
        return redirect(url_for(pagina_inicial(current_user.perfil)))

    conexao = get_db()
    cursor = conexao.cursor()
    cursor.execute("SELECT titulo, data_atividade, servico, descricao, criado_por FROM atividades_fotos WHERE id = %s", (id,))
    row = cursor.fetchone()
    if not row:
        conexao.close()
        return "Atividade não encontrada", 404

    if not _pode_gerenciar_atividade(row[4]):
        conexao.close()
        flash('❌ Você não tem permissão para editar esta atividade.', 'danger')
        return redirect(url_for('ver_atividade', id=id))

    lista_unidades = get_lista_cras() + ['CREAS', 'EQUIPE VOLANTE'] + get_lista_servicos()

    if request.method == "POST":
        titulo = request.form.get("titulo", "").strip()
        data_atividade = request.form.get("data_atividade", "").strip()
        servico = request.form.get("servico", "").strip()
        descricao = request.form.get("descricao", "").strip()

        if not titulo or not data_atividade:
            flash('❌ Preencha o título e a data da atividade.', 'danger')
        else:
            cursor.execute("""
                UPDATE atividades_fotos SET titulo=%s, data_atividade=%s, servico=%s, descricao=%s
                WHERE id=%s
            """, (titulo, data_atividade, servico, descricao, id))

            # Legendas das fotos já existentes (campo legenda_<id> por foto)
            cursor.execute("SELECT id FROM fotos_atividade WHERE atividade_id = %s", (id,))
            for (foto_id,) in cursor.fetchall():
                campo = f"legenda_{foto_id}"
                if campo in request.form:
                    cursor.execute("UPDATE fotos_atividade SET legenda=%s WHERE id=%s",
                                   (request.form.get(campo, "").strip(), foto_id))

            # Novas fotos adicionadas na edição — pareia arquivo+legenda pelo
            # índice da linha antes de filtrar (mesmo motivo do nova_atividade:
            # uma linha sem foto mas com legenda desalinharia as seguintes).
            arquivos_brutos = request.files.getlist("fotos[]")
            legendas_brutas = request.form.getlist("legenda[]")
            pares = [(f, legendas_brutas[i] if i < len(legendas_brutas) else '')
                     for i, f in enumerate(arquivos_brutos) if f and f.filename]
            if pares:
                if not storage_fotos.esta_configurado():
                    flash('⚠️ Metadados salvos, mas o envio de fotos não está configurado.', 'warning')
                else:
                    agora = datetime.now(FUSO_RONDONIA).strftime("%d/%m/%Y %H:%M:%S")
                    cursor.execute("SELECT COALESCE(MAX(ordem), -1) FROM fotos_atividade WHERE atividade_id = %s", (id,))
                    proxima_ordem = cursor.fetchone()[0] + 1
                    falhas = 0
                    for i, (arquivo, legenda) in enumerate(pares):
                        try:
                            dados_bytes = storage_fotos.comprimir_imagem(arquivo)
                            path = storage_fotos.gerar_caminho(id)
                            storage_fotos.upload_foto(dados_bytes, path)
                            cursor.execute("""
                                INSERT INTO fotos_atividade (atividade_id, storage_path, legenda, ordem, criado_em)
                                VALUES (%s,%s,%s,%s,%s)
                            """, (id, path, legenda, proxima_ordem + i, agora))
                        except Exception as e:
                            falhas += 1
                            msg = f"editar_atividade: falha ao adicionar foto à atividade {id}: {type(e).__name__}: {e}"
                            logger.error(msg)
                            print(msg)  # também visível no console/log do Render (logger só grava em arquivo local)
                    if falhas:
                        flash(f'⚠️ {falhas} foto(s) não puderam ser adicionadas.', 'warning')

            conexao.commit()
            conexao.close()
            logger.info(f"Atividade #{id} editada por {current_user.id}")
            flash('✅ Atividade atualizada!', 'success')
            return redirect(url_for('ver_atividade', id=id))

    cursor.execute("SELECT id, storage_path, legenda FROM fotos_atividade WHERE atividade_id = %s ORDER BY ordem, id", (id,))
    fotos = [{'id': fid, 'url': _url_foto(path), 'legenda': legenda}
             for (fid, path, legenda) in cursor.fetchall()]
    conexao.close()

    atividade = {'id': id, 'titulo': row[0], 'data_atividade': row[1], 'servico': row[2], 'descricao': row[3]}
    return render_template("editar_atividade.html", atividade=atividade, fotos=fotos, lista_unidades=lista_unidades)

@app.route("/atividades/<int:id>/excluir", methods=["POST"])
@login_required
def excluir_atividade(id):
    conexao = get_db()
    cursor = conexao.cursor()
    cursor.execute("SELECT criado_por, titulo FROM atividades_fotos WHERE id = %s", (id,))
    row = cursor.fetchone()
    if not row:
        conexao.close()
        return "Atividade não encontrada", 404
    if not _pode_gerenciar_atividade(row[0]):
        conexao.close()
        flash('❌ Você não tem permissão para excluir esta atividade.', 'danger')
        return redirect(url_for('ver_atividade', id=id))
    titulo = row[1]

    cursor.execute("SELECT storage_path FROM fotos_atividade WHERE atividade_id = %s", (id,))
    caminhos = [r[0] for r in cursor.fetchall()]
    cursor.execute("DELETE FROM atividades_fotos WHERE id = %s", (id,))
    conexao.commit()
    conexao.close()

    for path in caminhos:
        try:
            storage_fotos.excluir_foto(path)
        except Exception as e:
            logger.error(f"excluir_atividade: falha ao excluir foto '{path}' do Storage: {e}")

    logger.info(f"Atividade #{id} '{titulo}' excluída ({len(caminhos)} foto(s)) por {current_user.id}")
    flash('✅ Atividade excluída.', 'success')
    return redirect(url_for('atividades'))

@app.route("/atividades/foto/<int:foto_id>/excluir", methods=["POST"])
@login_required
def excluir_foto_atividade(foto_id):
    conexao = get_db()
    cursor = conexao.cursor()
    cursor.execute("""
        SELECT f.atividade_id, f.storage_path, a.criado_por, a.titulo
        FROM fotos_atividade f JOIN atividades_fotos a ON f.atividade_id = a.id
        WHERE f.id = %s
    """, (foto_id,))
    row = cursor.fetchone()
    if not row:
        conexao.close()
        return "Foto não encontrada", 404
    atividade_id, path, criado_por, titulo = row

    if not _pode_gerenciar_atividade(criado_por):
        conexao.close()
        flash('❌ Você não tem permissão para excluir esta foto.', 'danger')
        return redirect(url_for('editar_atividade', id=atividade_id))

    cursor.execute("DELETE FROM fotos_atividade WHERE id = %s", (foto_id,))
    conexao.commit()
    conexao.close()

    try:
        storage_fotos.excluir_foto(path)
    except Exception as e:
        logger.error(f"excluir_foto_atividade: falha ao excluir '{path}' do Storage: {e}")

    logger.info(f"Foto #{foto_id} ('{path}') da atividade #{atividade_id} '{titulo}' excluída por {current_user.id}")
    flash('✅ Foto removida.', 'success')
    return redirect(url_for('editar_atividade', id=atividade_id))

@app.route("/atividades/foto/<int:foto_id>/download")
@login_required
def download_foto_atividade(foto_id):
    if not _pode_acessar_atividades():
        flash('❌ Você não tem acesso a essa área.', 'danger')
        return redirect(url_for(pagina_inicial(current_user.perfil)))

    conexao = get_db()
    cursor = conexao.cursor()
    cursor.execute("""
        SELECT f.storage_path, a.servico, a.data_atividade, a.titulo
        FROM fotos_atividade f JOIN atividades_fotos a ON f.atividade_id = a.id
        WHERE f.id = %s
    """, (foto_id,))
    row = cursor.fetchone()
    conexao.close()
    if not row:
        return "Foto não encontrada", 404
    path, servico, data_ativ, titulo = row

    try:
        dados = storage_fotos.baixar_foto(path)
    except Exception as e:
        logger.error(f"download_foto_atividade: falha ao baixar '{path}': {e}")
        flash('❌ Não foi possível baixar esta foto agora. Tente novamente.', 'danger')
        return redirect(request.referrer or url_for('atividades'))

    nome = _nome_arquivo_foto(servico, data_ativ, titulo, foto_id)
    return send_file(io.BytesIO(dados), mimetype='image/jpeg', as_attachment=True, download_name=nome)

@app.route("/atividades/download_zip")
@login_required
def download_zip_atividades():
    """Baixa em um .zip as fotos de uma atividade específica (atividade_id)
    ou de um serviço num período/quadrimestre — é o fluxo usado pela
    vigilância a cada 4 meses para reunir as fotos do Quadrimestral."""
    if not _pode_acessar_atividades():
        flash('❌ Você não tem acesso a essa área.', 'danger')
        return redirect(url_for(pagina_inicial(current_user.perfil)))

    atividade_id  = request.args.get('atividade_id', '').strip()
    servico       = request.args.get('servico', '').strip()
    periodo_inicio = request.args.get('periodo_inicio', '').strip()
    periodo_fim    = request.args.get('periodo_fim', '').strip()
    quadrimestre  = request.args.get('quadrimestre', '').strip()
    ano           = request.args.get('ano', '').strip()
    if quadrimestre and ano:
        try:
            periodo_inicio, periodo_fim = _periodo_do_quadrimestre(int(ano), int(quadrimestre))
        except ValueError:
            pass

    conexao = get_db()
    cursor = conexao.cursor()

    if atividade_id:
        cursor.execute("""
            SELECT a.id, a.titulo, a.servico, a.data_atividade
            FROM atividades_fotos a WHERE a.id = %s
        """, (atividade_id,))
        atividades_rows = cursor.fetchall()
        nome_zip = f"atividade_{atividade_id}.zip"
    else:
        filtros = []
        params = []
        if periodo_inicio:
            filtros.append("a.data_atividade >= %s"); params.append(periodo_inicio)
        if periodo_fim:
            filtros.append("a.data_atividade <= %s"); params.append(periodo_fim)
        if servico:
            filtros.append("a.servico = %s"); params.append(servico)
        where = ("WHERE " + " AND ".join(filtros)) if filtros else ""
        cursor.execute(f"""
            SELECT a.id, a.titulo, a.servico, a.data_atividade
            FROM atividades_fotos a
            {where}
            ORDER BY a.data_atividade, a.id
        """, params)
        atividades_rows = cursor.fetchall()
        partes_nome = [servico or 'todos-servicos']
        if periodo_inicio or periodo_fim:
            partes_nome.append(f"{periodo_inicio or 'inicio'}_a_{periodo_fim or 'fim'}")
        nome_zip = _re.sub(r'[^A-Za-z0-9._-]+', '-', '_'.join(partes_nome)).strip('-') + '.zip'

    if not atividades_rows:
        conexao.close()
        flash('❌ Nenhuma atividade encontrada para esse filtro.', 'warning')
        return redirect(url_for('atividades'))

    ids = [row[0] for row in atividades_rows]
    info_atividade = {row[0]: {'titulo': row[1], 'servico': row[2], 'data': row[3]} for row in atividades_rows}
    cursor.execute("""
        SELECT id, atividade_id, storage_path FROM fotos_atividade
        WHERE atividade_id = ANY(%s) ORDER BY atividade_id, ordem, id
    """, (ids,))
    fotos_rows = cursor.fetchall()
    conexao.close()

    if not fotos_rows:
        flash('❌ Nenhuma foto encontrada para esse filtro.', 'warning')
        return redirect(url_for('atividades'))

    buffer = io.BytesIO()
    falhas = 0
    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for foto_id, ativ_id, path in fotos_rows:
            info = info_atividade[ativ_id]
            try:
                dados = storage_fotos.baixar_foto(path)
            except Exception as e:
                falhas += 1
                logger.error(f"download_zip_atividades: falha ao baixar '{path}': {e}")
                continue
            pasta = _re.sub(r'[^A-Za-z0-9._-]+', '-', info['servico'] or 'sem-servico').strip('-')
            nome = _nome_arquivo_foto(info['servico'], info['data'], info['titulo'], foto_id)
            zf.writestr(f"{pasta}/{nome}", dados)

    if falhas:
        flash(f'⚠️ {falhas} foto(s) não puderam ser incluídas no zip (falha ao baixar do armazenamento).', 'warning')

    buffer.seek(0)
    return send_file(buffer, mimetype='application/zip', as_attachment=True, download_name=nome_zip)

# =====================================================
# DASHBOARD
# =====================================================

@app.route("/dashboard")
@login_required
def dashboard():
    if current_user.perfil not in ['admin', 'gestor']:
        return redirect(url_for('solicitacoes'))

    conexao = get_db()
    cursor  = conexao.cursor()

    # Totais gerais (canceladas ficam fora do total ativo, mas são contadas à parte)
    cursor.execute("SELECT COUNT(*) FROM solicitacoes WHERE status != 'Cancelada'")
    total = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM solicitacoes WHERE status='Entregue'")
    entregues = cursor.fetchone()[0]
    # Ausências acumuladas: conta toda vez que um técnico foi até a casa e não
    # encontrou ninguém, mesmo que a solicitação já tenha sido entregue depois
    # (ver rota /nova_tentativa, que reabre um "Ausente" para "Cadastrada").
    cursor.execute("""
        SELECT COALESCE(SUM((COALESCE(num_tentativas, 1) - 1)
                             + CASE WHEN status='Ausente' THEN 1 ELSE 0 END), 0)
        FROM solicitacoes
    """)
    ausentes = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM solicitacoes WHERE status='Cadastrada'")
    pendentes = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM solicitacoes WHERE status='Cancelada'")
    canceladas = cursor.fetchone()[0]

    # Tempo médio de espera da concessão: escuta → entrega, considerando
    # apenas solicitações já entregues (evita distorção de escutas ainda em aberto).
    cursor.execute("""
        SELECT ROUND(AVG(data_entrega::date - data_escuta::date)::numeric, 1)
        FROM solicitacoes
        WHERE status = 'Entregue'
          AND data_escuta ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}'
          AND data_entrega ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}'
    """)
    tempo_medio_espera = cursor.fetchone()[0]

    # Escuta mais antiga ainda aguardando concessão (status pendente)
    cursor.execute("""
        SELECT id, nome, (CURRENT_DATE - data_escuta::date) AS dias
        FROM solicitacoes
        WHERE status = 'Cadastrada' AND data_escuta ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}'
        ORDER BY data_escuta::date ASC
        LIMIT 1
    """)
    escuta_mais_antiga = cursor.fetchone()

    # Por equipe (CRAS/perfil do técnico que registrou — produção real da equipe)
    # Não existe mais um bucket "ADMINISTRAÇÃO": solicitações de contas admin/gestor
    # caem na unidade informada no cadastro (s.cras) ou em "Não informado", sem
    # perder a contagem. Todas as canceladas do sistema somam numa coluna final única.
    #
    # "Ausentes" é acumulativo: num_tentativas só sobe quando uma visita "Ausente"
    # é reaberta (rota /nova_tentativa), então (num_tentativas - 1) preserva quantas
    # vezes o técnico foi até a casa e não encontrou ninguém, mesmo que a solicitação
    # já tenha sido entregue depois. Sem isso, cada reabertura apagava a visita anterior.
    cursor.execute("""
        SELECT
            CASE
                WHEN u.perfil = 'creas'       THEN 'CREAS'
                WHEN u.perfil = 'cras_volante' THEN 'EQUIPE VOLANTE'
                ELSE COALESCE(NULLIF(u.cras, ''), s.cras, 'Não informado')
            END AS equipe,
            SUM(CASE WHEN s.status != 'Cancelada' THEN 1 ELSE 0 END) AS total,
            SUM(CASE WHEN s.status='Entregue'  THEN 1 ELSE 0 END) AS entregues,
            SUM((COALESCE(s.num_tentativas, 1) - 1)
                + CASE WHEN s.status='Ausente' THEN 1 ELSE 0 END) AS ausentes,
            SUM(CASE WHEN s.status='Entregue' AND COALESCE(s.num_tentativas, 1) >= 2
                     THEN 1 ELSE 0 END) AS entregues_apos_tentativa
        FROM solicitacoes s
        LEFT JOIN usuarios u ON s.tecnico = u.usuario
        GROUP BY equipe ORDER BY total DESC
    """)
    por_cras = [list(row) + [0] for row in cursor.fetchall()]
    por_cras.append(['Canceladas', 0, 0, 0, 0, canceladas])

    # Últimos 6 meses (para gráfico de linha) — solicitações feitas (exclui
    # canceladas) e entregues, agrupadas por mês. UNION ALL + GROUP BY externo
    # pra alinhar as duas séries no mesmo conjunto de meses (a união dos meses
    # em que houve solicitação OU entrega).
    cursor.execute("""
        SELECT mes, SUM(solicitadas) AS solicitadas, SUM(entregues) AS entregues
        FROM (
            SELECT TO_CHAR(TO_DATE(SUBSTRING(data_solicitacao, 7, 4) || '-' ||
                                    SUBSTRING(data_solicitacao, 4, 2) || '-01', 'YYYY-MM-DD'), 'YYYY-MM') AS mes,
                   COUNT(*) AS solicitadas, 0 AS entregues
            FROM solicitacoes
            WHERE data_solicitacao ~ '^[0-9]{2}/[0-9]{2}/[0-9]{4}'
              AND status != 'Cancelada'
            GROUP BY mes
            UNION ALL
            SELECT SUBSTRING(data_entrega, 1, 7) AS mes, 0 AS solicitadas, COUNT(*) AS entregues
            FROM solicitacoes
            WHERE data_entrega ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}'
              AND status = 'Entregue'
            GROUP BY mes
        ) x
        GROUP BY mes ORDER BY mes DESC LIMIT 6
    """)
    por_mes_raw = cursor.fetchall()
    por_mes = list(reversed(por_mes_raw))

    conexao.close()

    return render_template(
        "dashboard.html",
        total_solicitacoes=total,
        total_entregues=entregues,
        total_ausentes=ausentes,
        total_pendentes=pendentes,
        total_canceladas=canceladas,
        tempo_medio_espera=tempo_medio_espera,
        escuta_mais_antiga=escuta_mais_antiga,
        por_cras=por_cras,
        por_mes=por_mes,
        datetime=datetime,
        current_user=current_user
    )

@app.route("/dashboard/escutas_aguardando")
@login_required
def escutas_aguardando():
    if current_user.perfil not in ['admin', 'gestor']:
        return redirect(url_for('solicitacoes'))

    conexao = get_db()
    cursor  = conexao.cursor()
    cursor.execute("""
        SELECT s.id, s.nome, COALESCE(u.cras, s.cras, 'Não informado') AS unidade,
               COALESCE(u.nome, s.tecnico) AS tecnico_nome, s.data_escuta,
               (CURRENT_DATE - s.data_escuta::date) AS dias
        FROM solicitacoes s
        LEFT JOIN usuarios u ON s.tecnico = u.usuario
        WHERE s.status = 'Cadastrada' AND s.data_escuta ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}'
        ORDER BY s.data_escuta::date ASC
        LIMIT 10
    """)
    dados_raw = cursor.fetchall()
    conexao.close()

    dados = []
    for row in dados_raw:
        row = list(row)
        partes = str(row[4]).split('-')
        if len(partes) == 3:
            row[4] = f"{partes[2]}/{partes[1]}/{partes[0]}"
        dados.append(row)

    return render_template(
        "escutas_aguardando.html",
        dados=dados,
        current_user=current_user
    )

# =====================================================
# RELATÓRIO
# =====================================================

def _periodo_sql(periodo_inicio, periodo_fim):
    """Retorna (condicao_sql, params) para filtro de período YYYY-MM.
    Se nenhum período for informado, retorna condição vazia (mostra tudo)."""
    if not periodo_inicio and not periodo_fim:
        return "", []
    inicio = periodo_inicio or '2000-01'
    fim    = periodo_fim    or '2099-12'
    cond = """(
        (data_solicitacao ~ '^[0-9]' AND
         (SUBSTRING(data_solicitacao, 7, 4) || '-' || SUBSTRING(data_solicitacao, 4, 2)) BETWEEN %s AND %s)
        OR (data_entrega IS NOT NULL AND data_entrega != '' AND SUBSTRING(data_entrega::text, 1, 7) BETWEEN %s AND %s)
    )"""
    return cond, [inicio, fim, inicio, fim]

def _periodo_sql_dia(data_inicio, data_fim):
    """Como _periodo_sql, mas granularidade de dia (YYYY-MM-DD) em vez de
    mês — usada só pela listagem de Solicitações, que filtra por data
    específica. Mesma semântica: casa na data da escuta OU da entrega."""
    if not data_inicio and not data_fim:
        return "", []
    inicio = data_inicio or '2000-01-01'
    fim    = data_fim    or '2099-12-31'
    cond = """(
        (data_solicitacao ~ '^[0-9]' AND
         (SUBSTRING(data_solicitacao, 7, 4) || '-' || SUBSTRING(data_solicitacao, 4, 2) || '-' || SUBSTRING(data_solicitacao, 1, 2)) BETWEEN %s AND %s)
        OR (data_entrega IS NOT NULL AND data_entrega != '' AND SUBSTRING(data_entrega::text, 1, 10) BETWEEN %s AND %s)
    )"""
    return cond, [inicio, fim, inicio, fim]

def _label_periodo(periodo_inicio, periodo_fim):
    """Gera texto legível para o período selecionado."""
    nomes = ['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez']
    def fmt(p):
        try:
            a, m = p.split('-')
            return f"{nomes[int(m)-1]}/{a}"
        except Exception as e:
            logger.error(f"_label_periodo: falha ao formatar período '{p}': {e}")
            return p
    if periodo_inicio and periodo_fim:
        if periodo_inicio == periodo_fim:
            return fmt(periodo_inicio)
        return f"{fmt(periodo_inicio)} a {fmt(periodo_fim)}"
    if periodo_inicio:
        return f"A partir de {fmt(periodo_inicio)}"
    if periodo_fim:
        return f"Até {fmt(periodo_fim)}"
    return "Todo o período"

@app.route("/relatorio")
@login_required
def relatorio():
    periodo_inicio = request.args.get('periodo_inicio', '').strip()
    periodo_fim    = request.args.get('periodo_fim', '').strip()

    nomes_meses = ['Janeiro','Fevereiro','Março','Abril','Maio','Junho',
                   'Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']
    lista_meses = []
    agora = datetime.now(FUSO_RONDONIA)
    for i in range(24):
        data = agora - timedelta(days=30 * i)
        if data.year < 2026:
            break
        valor = data.strftime('%Y-%m')
        nome = f"{nomes_meses[data.month - 1]}/{data.year}"
        lista_meses.append({'valor': valor, 'nome': nome})

    cond_periodo, p_periodo = _periodo_sql(periodo_inicio, periodo_fim)
    w = f"WHERE {cond_periodo}" if cond_periodo else ""

    def _and(base):
        return (base + " AND ") if base else "WHERE "

    conexao = get_db()
    cursor = conexao.cursor()

    cursor.execute(f"SELECT COUNT(*) FROM solicitacoes {_and(w)}status != 'Cancelada'", p_periodo)
    total_solicitacoes = cursor.fetchone()[0]

    cursor.execute(f"SELECT COUNT(*) FROM solicitacoes {_and(w)}status='Entregue'", p_periodo)
    total_entregues = cursor.fetchone()[0]

    cursor.execute(f"""
        SELECT COALESCE(SUM((COALESCE(num_tentativas, 1) - 1)
                             + CASE WHEN status='Ausente' THEN 1 ELSE 0 END), 0)
        FROM solicitacoes {w}
    """, p_periodo)
    total_ausentes = cursor.fetchone()[0]

    cursor.execute(f"SELECT COUNT(*) FROM solicitacoes {_and(w)}status='Cadastrada'", p_periodo)
    total_pendentes = cursor.fetchone()[0]

    cursor.execute(f"SELECT COUNT(*) FROM solicitacoes {_and(w)}excecao_art64=TRUE", p_periodo)
    total_excecoes = cursor.fetchone()[0]

    # Por Unidade — separa escutas (tecnico) de entregas/ausentes (tecnico_entrega)
    _u_case = """CASE
                    WHEN u.perfil='creas'             THEN 'CREAS'
                    WHEN u.perfil='cras_volante'       THEN 'EQUIPE VOLANTE'
                    WHEN u.perfil IN ('admin','gestor') THEN 'ADMINISTRAÇÃO'
                    ELSE COALESCE(u.cras, s.cras, 'Não informado')
                 END"""
    cursor.execute(f"""
        SELECT unidade,
               SUM(escutas) as escutas,
               SUM(entregas) as entregas,
               SUM(ausentes) as ausentes,
               SUM(escutas + entregas + ausentes) as total
        FROM (
            SELECT {_u_case} AS unidade,
                   COUNT(*) as escutas, 0 as entregas, 0 as ausentes
            FROM solicitacoes s LEFT JOIN usuarios u ON s.tecnico = u.usuario
            {_and(w)}s.status != 'Cancelada'
            GROUP BY unidade
            UNION ALL
            SELECT {_u_case} AS unidade,
                   0 as escutas,
                   SUM(CASE WHEN s.status='Entregue' THEN 1 ELSE 0 END) as entregas,
                   SUM((COALESCE(s.num_tentativas, 1) - 1)
                       + CASE WHEN s.status='Ausente' THEN 1 ELSE 0 END) as ausentes
            FROM solicitacoes s LEFT JOIN usuarios u ON s.tecnico_entrega = u.usuario
            {_and(w)}s.tecnico_entrega IS NOT NULL AND s.status != 'Cancelada'
            GROUP BY unidade
        ) t
        GROUP BY unidade ORDER BY total DESC
    """, p_periodo + p_periodo)
    por_cras = cursor.fetchall()

    # Por Técnico — separa escutas de entregas/ausentes
    cursor.execute(f"""
        SELECT nome_tecnico,
               SUM(escutas) as escutas,
               SUM(entregas) as entregas,
               SUM(ausentes) as ausentes,
               SUM(escutas + entregas + ausentes) as total
        FROM (
            SELECT COALESCE(u.nome, s.tecnico) as nome_tecnico,
                   COUNT(*) as escutas, 0 as entregas, 0 as ausentes
            FROM solicitacoes s LEFT JOIN usuarios u ON s.tecnico = u.usuario
            {_and(w)}s.status != 'Cancelada'
            GROUP BY COALESCE(u.nome, s.tecnico)
            UNION ALL
            SELECT COALESCE(u.nome, s.tecnico_entrega) as nome_tecnico,
                   0 as escutas,
                   SUM(CASE WHEN s.status='Entregue' THEN 1 ELSE 0 END) as entregas,
                   SUM((COALESCE(s.num_tentativas, 1) - 1)
                       + CASE WHEN s.status='Ausente' THEN 1 ELSE 0 END) as ausentes
            FROM solicitacoes s LEFT JOIN usuarios u ON s.tecnico_entrega = u.usuario
            {_and(w)}s.tecnico_entrega IS NOT NULL AND s.status != 'Cancelada'
            GROUP BY COALESCE(u.nome, s.tecnico_entrega)
        ) t
        GROUP BY nome_tecnico ORDER BY total DESC
    """, p_periodo + p_periodo)
    por_tecnico = cursor.fetchall()

    # Últimas entregas do período (máx 20)
    cursor.execute(f"""
        SELECT s.nome, s.cpf, s.bairro,
               CASE
                   WHEN u_esc.perfil = 'creas'             THEN 'CREAS'
                   WHEN u_esc.perfil = 'cras_volante'       THEN 'EQUIPE VOLANTE'
                   WHEN u_esc.perfil IN ('admin','gestor')  THEN 'ADMINISTRAÇÃO'
                   ELSE COALESCE(u_esc.cras, s.cras, 'Não informado')
               END AS unidade,
               s.status, s.data_entrega,
               COALESCE(u_ent.nome, s.tecnico_entrega) as tecnico
        FROM solicitacoes s
        LEFT JOIN usuarios u_esc ON s.tecnico = u_esc.usuario
        LEFT JOIN usuarios u_ent ON s.tecnico_entrega = u_ent.usuario
        {_and(w)}s.status IN ('Entregue', 'Ausente')
        ORDER BY s.id DESC LIMIT 20
    """, p_periodo)
    raw_entregas = cursor.fetchall()

    ultimas_entregas = []
    for row in raw_entregas:
        row = list(row)
        if row[1]:
            row[1] = formatar_cpf(descriptografar_cpf(row[1]))
        ultimas_entregas.append(tuple(row))

    # Recorrência: beneficiários com mais de 1 cesta entregue (sem filtro de período)
    cursor.execute("""
        SELECT cpf_hash, MAX(nome), MAX(cpf), COUNT(*), MAX(data_entrega)
        FROM solicitacoes
        WHERE status='Entregue' AND cpf_hash IS NOT NULL
        GROUP BY cpf_hash HAVING COUNT(*) > 1
        ORDER BY COUNT(*) DESC LIMIT 30
    """)
    raw_recorrencia = cursor.fetchall()
    conexao.close()

    recorrencia = []
    for row in raw_recorrencia:
        cpf_legivel = formatar_cpf(descriptografar_cpf(row[2])) if row[2] else 'N/A'
        recorrencia.append({
            'cpf': cpf_legivel, 'nome': row[1] or 'N/A',
            'total_recebido': row[3],
            'ultima_entrega': str(row[4]) if row[4] else 'N/A',
        })

    label_periodo = _label_periodo(periodo_inicio, periodo_fim)

    return render_template(
        "relatorio.html",
        lista_meses=lista_meses,
        periodo_inicio=periodo_inicio,
        periodo_fim=periodo_fim,
        label_periodo=label_periodo,
        total_solicitacoes=total_solicitacoes,
        total_entregues=total_entregues,
        total_ausentes=total_ausentes,
        total_pendentes=total_pendentes,
        total_excecoes=total_excecoes,
        por_cras=por_cras,
        por_tecnico=por_tecnico,
        ultimas_entregas=ultimas_entregas,
        recorrencia=recorrencia,
        datetime=datetime,
        current_user=current_user
    )

# =====================================================
# EXPORTAÇÃO PDF DO RELATÓRIO
# =====================================================

def _coletar_dados_relatorio(periodo_inicio, periodo_fim):
    """Coleta os totais e tabelas do relatório para um período. Retorna um dict."""
    cond_periodo, p_periodo = _periodo_sql(periodo_inicio, periodo_fim)
    where = f"WHERE {cond_periodo}" if cond_periodo else ""

    conexao = get_db()
    cursor = conexao.cursor()

    cursor.execute(f"SELECT COUNT(*) FROM solicitacoes {where}", p_periodo)
    total = cursor.fetchone()[0]

    cursor.execute(f"SELECT COUNT(*) FROM solicitacoes {where + (' AND ' if where else 'WHERE ')}status='Entregue'", p_periodo)
    entregues = cursor.fetchone()[0]

    cursor.execute(f"""
        SELECT COALESCE(SUM((COALESCE(num_tentativas, 1) - 1)
                             + CASE WHEN status='Ausente' THEN 1 ELSE 0 END), 0)
        FROM solicitacoes {where}
    """, p_periodo)
    ausentes = cursor.fetchone()[0]

    cursor.execute(f"SELECT COUNT(*) FROM solicitacoes {where + (' AND ' if where else 'WHERE ')}status='Cadastrada'", p_periodo)
    pendentes = cursor.fetchone()[0]

    cursor.execute(f"SELECT COUNT(*) FROM solicitacoes {where + (' AND ' if where else 'WHERE ')}excecao_art64=TRUE", p_periodo)
    excecoes = cursor.fetchone()[0]

    _wp = f"WHERE {cond_periodo}" if cond_periodo else ""
    _andp = (_wp + " AND ") if _wp else "WHERE "
    _uc = """CASE WHEN u.perfil='creas' THEN 'CREAS'
                  WHEN u.perfil='cras_volante' THEN 'EQUIPE VOLANTE'
                  WHEN u.perfil IN ('admin','gestor') THEN 'ADMINISTRAÇÃO'
                  ELSE COALESCE(u.cras, s.cras, 'Não informado') END"""

    cursor.execute(f"""
        SELECT unidade, SUM(escutas), SUM(entregas), SUM(ausentes), SUM(escutas+entregas+ausentes)
        FROM (
            SELECT {_uc} AS unidade, COUNT(*) as escutas, 0 as entregas, 0 as ausentes
            FROM solicitacoes s LEFT JOIN usuarios u ON s.tecnico=u.usuario
            {_andp}s.status!='Cancelada' GROUP BY unidade
            UNION ALL
            SELECT {_uc} AS unidade, 0,
                   SUM(CASE WHEN s.status='Entregue' THEN 1 ELSE 0 END),
                   SUM((COALESCE(s.num_tentativas, 1) - 1)
                       + CASE WHEN s.status='Ausente' THEN 1 ELSE 0 END)
            FROM solicitacoes s LEFT JOIN usuarios u ON s.tecnico_entrega=u.usuario
            {_andp}s.tecnico_entrega IS NOT NULL AND s.status!='Cancelada' GROUP BY unidade
        ) t GROUP BY unidade ORDER BY SUM(escutas+entregas+ausentes) DESC
    """, p_periodo + p_periodo)
    por_cras = cursor.fetchall()

    cursor.execute(f"""
        SELECT nome_tecnico, SUM(escutas), SUM(entregas), SUM(ausentes), SUM(escutas+entregas+ausentes)
        FROM (
            SELECT COALESCE(u.nome, s.tecnico) as nome_tecnico,
                   COUNT(*) as escutas, 0 as entregas, 0 as ausentes
            FROM solicitacoes s LEFT JOIN usuarios u ON s.tecnico=u.usuario
            {_andp}s.status!='Cancelada' GROUP BY COALESCE(u.nome, s.tecnico)
            UNION ALL
            SELECT COALESCE(u.nome, s.tecnico_entrega) as nome_tecnico, 0,
                   SUM(CASE WHEN s.status='Entregue' THEN 1 ELSE 0 END),
                   SUM((COALESCE(s.num_tentativas, 1) - 1)
                       + CASE WHEN s.status='Ausente' THEN 1 ELSE 0 END)
            FROM solicitacoes s LEFT JOIN usuarios u ON s.tecnico_entrega=u.usuario
            {_andp}s.tecnico_entrega IS NOT NULL AND s.status!='Cancelada'
            GROUP BY COALESCE(u.nome, s.tecnico_entrega)
        ) t GROUP BY nome_tecnico ORDER BY SUM(escutas+entregas+ausentes) DESC
    """, p_periodo + p_periodo)
    por_tecnico = cursor.fetchall()

    conexao.close()
    return {
        'total': total, 'entregues': entregues, 'ausentes': ausentes,
        'pendentes': pendentes, 'excecoes': excecoes,
        'por_cras': por_cras, 'por_tecnico': por_tecnico
    }

def _nome_mes_extenso(mes):
    nomes = ['Janeiro','Fevereiro','Março','Abril','Maio','Junho',
             'Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']
    try:
        ano, num = mes.split('-')
        return f"{nomes[int(num)-1]}/{ano}"
    except Exception as e:
        logger.error(f"_nome_mes_extenso: falha ao formatar mês '{mes}': {e}")
        return mes

def _desenhar_cabecalho_relatorio(c, w, h, mes, subtitulo):
    logo_path = os.path.join(os.path.dirname(__file__), 'static', 'img', 'logo_prefeitura.png')
    logo_h = 1.8*cm
    if os.path.exists(logo_path):
        try:
            c.drawImage(ImageReader(logo_path), 2*cm, h - 2*cm - logo_h + 0.2*cm,
                        height=logo_h, width=logo_h,
                        preserveAspectRatio=True, mask='auto')
        except Exception:
            pass
    c.setFont("Helvetica-Bold", 14)
    c.setFillColorRGB(0.12, 0.24, 0.45)
    c.drawCentredString(w/2, h - 2*cm, "PREFEITURA MUNICIPAL DE JI-PARANÁ")
    c.setFont("Helvetica", 10)
    c.setFillColorRGB(0, 0, 0)
    c.drawCentredString(w/2, h - 2.6*cm, "Secretaria Municipal de Assistência Social e Família - SEMASF")
    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(w/2, h - 3.4*cm, subtitulo)
    c.setFont("Helvetica", 10)
    c.drawCentredString(w/2, h - 3.9*cm, f"Mês de referência: {_nome_mes_extenso(mes)}")
    c.setLineWidth(1)
    c.setStrokeColorRGB(0.12, 0.24, 0.45)
    c.line(2*cm, h - 4.2*cm, w - 2*cm, h - 4.2*cm)

def _desenhar_totais(c, w, y, dados):
    cards = [
        ("Total de Solicitações", dados['total'], (0.12, 0.24, 0.45)),
        ("Entregues", dados['entregues'], (0.15, 0.55, 0.25)),
        ("Ausentes", dados['ausentes'], (0.70, 0.15, 0.15)),
        ("Pendentes", dados['pendentes'], (0.80, 0.55, 0.10)),
        ("Exceções Art. 64", dados['excecoes'], (0.55, 0.15, 0.55)),
    ]
    card_w = (w - 4*cm) / len(cards)
    for i, (label, valor, cor) in enumerate(cards):
        x = 2*cm + i * card_w
        c.setStrokeColorRGB(*cor)
        c.setLineWidth(1)
        c.rect(x + 0.1*cm, y - 1.6*cm, card_w - 0.2*cm, 1.5*cm, fill=0)
        c.setFillColorRGB(*cor)
        c.setFont("Helvetica-Bold", 20)
        c.drawCentredString(x + card_w/2, y - 0.85*cm, str(valor))
        c.setFillColorRGB(0.2, 0.2, 0.2)
        c.setFont("Helvetica", 7)
        c.drawCentredString(x + card_w/2, y - 1.35*cm, label)
    return y - 2.2*cm

def _rodape_relatorio(c, w, mes):
    agora = datetime.now(FUSO_RONDONIA)
    c.setFont("Helvetica", 7)
    c.setFillColorRGB(0.4, 0.4, 0.4)
    c.drawString(2*cm, 1*cm, f"Relatório gerado em {agora.strftime('%d/%m/%Y às %H:%M:%S')} (horário de Rondônia)")
    c.drawRightString(w - 2*cm, 1*cm, "SEMASF Ji-Paraná - Sistema de Cestas Básicas")

# =====================================================
# EXPORTAÇÃO EXCEL
# =====================================================

@app.route("/relatorio/excel")
@login_required
def exportar_excel():
    if current_user.perfil not in ['admin', 'gestor', 'creas', 'cras_volante', 'cras']:
        return "Acesso negado", 403

    periodo_inicio = request.args.get('periodo_inicio', '').strip()
    periodo_fim    = request.args.get('periodo_fim', '').strip()
    cond_periodo, p_periodo = _periodo_sql(periodo_inicio, periodo_fim)
    where_excel = f"WHERE {cond_periodo}" if cond_periodo else ""

    conexao = get_db()
    cursor  = conexao.cursor()
    cursor.execute(f"""
        SELECT id, nome, cpf, bairro, cras, tecnico, status,
               data_solicitacao, data_entrega, tecnico_entrega,
               renda_per_capita, total_pessoas, excecao_art64, observacoes_entrega
        FROM solicitacoes {where_excel}
        ORDER BY id DESC
    """, p_periodo)
    linhas = cursor.fetchall()
    conexao.close()

    label = _label_periodo(periodo_inicio, periodo_fim)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Cestas {label}"[:31]

    # Estilo cabeçalho
    azul_escuro = PatternFill("solid", fgColor="1B2F5E")
    fonte_branca = Font(bold=True, color="FFFFFF")
    cabecalhos = [
        "ID", "Nome", "CPF", "Bairro", "CRAS", "Técnico",
        "Status", "Data Solicitação", "Data Entrega", "Técnico Entrega",
        "Renda Per Capita", "Nº Pessoas", "Exceção Art.64", "Observações"
    ]
    for col, titulo in enumerate(cabecalhos, 1):
        cel = ws.cell(row=1, column=col, value=titulo)
        cel.font   = fonte_branca
        cel.fill   = azul_escuro
        cel.alignment = Alignment(horizontal="center")

    # Dados
    for row_idx, row in enumerate(linhas, 2):
        row = list(row)
        # Descriptografar CPF
        if row[2]:
            row[2] = formatar_cpf(descriptografar_cpf(row[2]))
        # Booleano legível
        row[12] = "Sim" if row[12] else "Não"
        for col_idx, valor in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=valor)

    # Ajustar largura das colunas
    for col in ws.columns:
        max_len = max((len(str(cel.value)) for cel in col if cel.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    slug = (periodo_inicio or 'tudo') + ('_a_' + periodo_fim if periodo_fim and periodo_fim != periodo_inicio else '')
    nome_arquivo = f"cestas_semasf_{slug}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/relatorio/pdf/<tipo>")
@login_required
def relatorio_pdf(tipo):
    if current_user.perfil not in ['admin', 'gestor', 'creas', 'cras_volante', 'cras']:
        return "Acesso negado", 403

    periodo_inicio = request.args.get('periodo_inicio', '').strip()
    periodo_fim    = request.args.get('periodo_fim', '').strip()
    dados = _coletar_dados_relatorio(periodo_inicio, periodo_fim)
    label = _label_periodo(periodo_inicio, periodo_fim)

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    w, h = A4

    subtitulo = "RELATÓRIO RESUMIDO DE CESTAS BÁSICAS" if tipo == 'resumido' else "RELATÓRIO DE CESTAS BÁSICAS"
    _desenhar_cabecalho_relatorio(c, w, h, label, subtitulo)

    y = h - 5*cm
    c.setFont("Helvetica-Bold", 11)
    c.setFillColorRGB(0, 0.3, 0)
    c.drawString(2*cm, y, f"RESUMO — {label.upper()}")
    c.setFillColorRGB(0, 0, 0)
    y -= 0.3*cm
    y = _desenhar_totais(c, w, y, dados)

    # No PDF completo, adicionar as tabelas Por CRAS e Por Técnico
    if tipo == 'completo':
        y -= 0.5*cm
        # Tabela Por CRAS
        c.setFont("Helvetica-Bold", 11)
        c.setFillColorRGB(0, 0.3, 0)
        c.drawString(2*cm, y, "DISTRIBUIÇÃO POR UNIDADE RESPONSÁVEL")
        c.setFillColorRGB(0, 0, 0)
        y -= 0.6*cm
        c.setFont("Helvetica-Bold", 9)
        c.drawString(2.2*cm, y, "Unidade")
        c.drawString(11*cm, y, "Total")
        c.drawString(13.5*cm, y, "Entregues")
        c.drawString(16.5*cm, y, "Ausentes")
        y -= 0.1*cm
        c.line(2*cm, y, w - 2*cm, y)
        y -= 0.4*cm
        def nova_pagina_relatorio():
            _rodape_relatorio(c, w, label)
            c.showPage()
            _desenhar_cabecalho_relatorio(c, w, h, label, subtitulo)
            return h - 5*cm

        c.setFont("Helvetica", 9)
        for cras, _escutas, ent, aus, tot in dados['por_cras']:
            if y < 4*cm:
                y = nova_pagina_relatorio()
            c.drawString(2.2*cm, y, (cras or 'N/A')[:45])
            c.drawString(11*cm, y, str(tot))
            c.drawString(13.5*cm, y, str(ent or 0))
            c.drawString(16.5*cm, y, str(aus or 0))
            y -= 0.45*cm

        y -= 0.5*cm
        # Tabela Por Técnico
        if y < 6*cm:
            y = nova_pagina_relatorio()
        c.setFont("Helvetica-Bold", 11)
        c.setFillColorRGB(0, 0.3, 0)
        c.drawString(2*cm, y, "DISTRIBUIÇÃO POR TÉCNICO")
        c.setFillColorRGB(0, 0, 0)
        y -= 0.6*cm
        c.setFont("Helvetica-Bold", 9)
        c.drawString(2.2*cm, y, "Técnico")
        c.drawString(11*cm, y, "Total")
        c.drawString(13.5*cm, y, "Entregues")
        c.drawString(16.5*cm, y, "Ausentes")
        y -= 0.1*cm
        c.line(2*cm, y, w - 2*cm, y)
        y -= 0.4*cm
        c.setFont("Helvetica", 9)
        for nome, _escutas, ent, aus, tot in dados['por_tecnico']:
            if y < 3*cm:
                y = nova_pagina_relatorio()
            c.drawString(2.2*cm, y, (nome or 'N/A')[:45])
            c.drawString(11*cm, y, str(tot))
            c.drawString(13.5*cm, y, str(ent or 0))
            c.drawString(16.5*cm, y, str(aus or 0))
            y -= 0.45*cm

    _rodape_relatorio(c, w, label)
    c.save()
    buffer.seek(0)

    slug = (periodo_inicio or 'tudo') + ('_a_' + periodo_fim if periodo_fim and periodo_fim != periodo_inicio else '')
    nome_arq = f"relatorio_{tipo}_{slug}.pdf"
    logger.info(f"Relatório PDF ({tipo}) período {label} gerado por {current_user.id}")
    return send_file(buffer, as_attachment=True, download_name=nome_arq, mimetype='application/pdf')

# =====================================================
# USUÁRIOS
# =====================================================

@app.route("/usuarios")
@login_required
def listar_usuarios():
    if current_user.perfil not in ['admin', 'gestor']:
        return redirect(url_for("dashboard"))
    conexao = get_db()
    cursor = conexao.cursor()
    cursor.execute("SELECT id, usuario, nome, perfil, cras, email, acesso_atividades FROM usuarios ORDER BY id")
    usuarios = cursor.fetchall()
    conexao.close()
    lista_cras = get_lista_cras()
    lista_servicos = get_lista_servicos()
    return render_template("usuarios.html", usuarios=usuarios, current_user=current_user,
                           lista_cras=lista_cras, lista_servicos=lista_servicos)

def _eh_ultimo_admin(id_usuario):
    """True se o usuário for admin E for o único admin restante no sistema.
    Usado para bloquear exclusão/rebaixamento que travaria a prefeitura
    fora das funções administrativas (o bootstrap automático só recria um
    admin quando a tabela usuarios está totalmente vazia)."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT perfil FROM usuarios WHERE id = %s", (id_usuario,))
    row = cursor.fetchone()
    if not row or row[0] != 'admin':
        conn.close()
        return False
    cursor.execute("SELECT COUNT(*) FROM usuarios WHERE perfil = 'admin'")
    total_admins = cursor.fetchone()[0]
    conn.close()
    return total_admins <= 1


@app.route("/usuario/novo", methods=["GET", "POST"])
@login_required
def novo_usuario():
    if request.method == "POST":
        senha = request.form.get("senha", "")
        perfil = request.form.get("perfil", "")
        # CRAS/serviço só fazem sentido para esses dois perfis; demais ficam NULL
        cras = request.form.get("cras") if perfil in ("cras", "servico") else None
        acesso_atividades = request.form.get("acesso_atividades") == "1"
        if len(senha) < 6:
            flash("Mínimo 6 caracteres!", "danger")
        else:
            hash_senha = bcrypt.hashpw(senha.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            conexao = get_db()
            cursor = conexao.cursor()
            try:
                cursor.execute(
                    "INSERT INTO usuarios (usuario, nome, senha, perfil, cras, primeiro_acesso, acesso_atividades) VALUES (%s,%s,%s,%s,%s,1,%s)",
                    (request.form.get("usuario"), request.form.get("nome"), hash_senha, perfil, cras, acesso_atividades)
                )
                conexao.commit()
                flash("✅ Usuário criado!", 'success')
            except Exception as e:
                if 'unique' in str(e).lower() or 'duplicate' in str(e).lower():
                    flash("❌ Usuário já existe!", 'danger')
                else:
                    logger.error(f"Erro ao criar usuário: {e}")
                    flash("❌ Erro ao criar usuário. Tente novamente.", 'danger')
                conexao.rollback()
            finally:
                conexao.close()
    lista_cras = get_lista_cras()
    lista_servicos = get_lista_servicos()
    return render_template("novo_usuario.html", lista_cras=lista_cras, lista_servicos=lista_servicos)

# POST obrigatório: exclusão via GET dispensa a proteção CSRF e pode ser
# disparada por um simples link externo ou prefetch do navegador
@app.route("/usuario/excluir/<int:id>", methods=["POST"])
@login_required
def excluir_usuario(id):
    if current_user.perfil != 'admin':
        return "Acesso negado", 403
    if _eh_ultimo_admin(id):
        flash("❌ Não é possível excluir o único administrador do sistema.", "danger")
        return redirect(url_for("listar_usuarios"))
    conexao = get_db()
    cursor = conexao.cursor()
    cursor.execute("DELETE FROM usuarios WHERE id = %s", (id,))
    conexao.commit()
    conexao.close()
    logger.info(f"Usuário ID={id} excluído por {current_user.id}")
    flash("✅ Usuário excluído.", "success")
    return redirect(url_for("listar_usuarios"))

@app.route("/usuario/editar/<int:id>", methods=["POST"])
@login_required
def editar_usuario(id):
    if current_user.perfil != 'admin':
        return "Acesso negado", 403
    nome = request.form.get("nome", "").strip()
    if not nome:
        flash("Nome não pode ser vazio.", "danger")
        return redirect(url_for("listar_usuarios"))
    conexao = get_db()
    cursor = conexao.cursor()
    cursor.execute("UPDATE usuarios SET nome = %s WHERE id = %s", (nome, id))
    conexao.commit()
    conexao.close()
    logger.info(f"Nome do usuário ID={id} alterado por {current_user.id}")
    flash("✅ Nome atualizado!", "success")
    return redirect(url_for("listar_usuarios"))

@app.route("/usuario/editar_perfil/<int:id>", methods=["POST"])
@login_required
def editar_perfil_usuario(id):
    if current_user.perfil not in ['admin', 'gestor']:
        return "Acesso negado", 403
    perfil_novo = request.form.get("perfil", "")
    perfis_validos = ['cras', 'creas', 'cras_volante', 'servico', 'gestor', 'admin']
    if perfil_novo not in perfis_validos:
        flash("Perfil inválido.", "danger")
        return redirect(url_for("listar_usuarios"))
    if perfil_novo != 'admin' and _eh_ultimo_admin(id):
        flash("❌ Não é possível rebaixar o único administrador do sistema.", "danger")
        return redirect(url_for("listar_usuarios"))

    conexao = get_db()
    cursor = conexao.cursor()
    cursor.execute("SELECT perfil FROM usuarios WHERE id = %s", (id,))
    row = cursor.fetchone()
    perfil_atual = row[0] if row else None

    # Gestor não pode promover/rebaixar admins nem criar novos admins
    if current_user.perfil == 'gestor':
        if perfil_atual == 'admin':
            conexao.close()
            flash("❌ Gestores não podem alterar perfil de administradores.", "danger")
            return redirect(url_for("listar_usuarios"))
        if perfil_novo == 'admin':
            conexao.close()
            flash("❌ Gestores não podem criar administradores.", "danger")
            return redirect(url_for("listar_usuarios"))

    # 'cras' e 'servico' guardam o nome da unidade/serviço na mesma coluna
    # `cras`, mas com significados diferentes — ao trocar de um perfil pro
    # outro (ou pra qualquer outro perfil), o valor antigo fica sem sentido
    # e precisa ser limpo; só preserva quando o perfil não mudou de fato.
    if perfil_novo == perfil_atual:
        cursor.execute("UPDATE usuarios SET perfil = %s WHERE id = %s", (perfil_novo, id))
    else:
        cursor.execute("UPDATE usuarios SET perfil = %s, cras = NULL WHERE id = %s", (perfil_novo, id))
    conexao.commit()
    conexao.close()
    logger.info(f"Perfil do usuário ID={id} alterado para {perfil_novo} por {current_user.id}")
    flash("✅ Perfil atualizado!", "success")
    return redirect(url_for("listar_usuarios"))

@app.route("/usuario/editar_cras/<int:id>", methods=["POST"])
@login_required
def editar_cras_usuario(id):
    if current_user.perfil not in ['admin', 'gestor']:
        return "Acesso negado", 403
    cras = request.form.get("cras", "").strip() or None
    # Gestor não pode alterar admins
    if current_user.perfil == 'gestor':
        conexao = get_db()
        cursor = conexao.cursor()
        cursor.execute("SELECT perfil FROM usuarios WHERE id = %s", (id,))
        row = cursor.fetchone()
        conexao.close()
        if row and row[0] == 'admin':
            flash("❌ Gestores não podem alterar dados de administradores.", "danger")
            return redirect(url_for("listar_usuarios"))
    conexao = get_db()
    cursor = conexao.cursor()
    cursor.execute("UPDATE usuarios SET cras = %s WHERE id = %s", (cras, id))
    conexao.commit()
    conexao.close()
    logger.info(f"CRAS do usuário ID={id} alterado para {cras} por {current_user.id}")
    flash("✅ CRAS atualizado!", "success")
    return redirect(url_for("listar_usuarios"))

@app.route("/usuario/toggle_atividades/<int:id>", methods=["POST"])
@login_required
def toggle_acesso_atividades(id):
    """Liga/desliga o acesso à página de Fotos Quadrimestral para um
    usuário, sem alterar o perfil (e portanto sem mexer nos privilégios
    de solicitações que ele já tem)."""
    if current_user.perfil not in ['admin', 'gestor']:
        return "Acesso negado", 403
    conexao = get_db()
    cursor = conexao.cursor()
    if current_user.perfil == 'gestor':
        cursor.execute("SELECT perfil FROM usuarios WHERE id = %s", (id,))
        row = cursor.fetchone()
        if row and row[0] == 'admin':
            conexao.close()
            flash("❌ Gestores não podem alterar dados de administradores.", "danger")
            return redirect(url_for("listar_usuarios"))
    cursor.execute("""
        UPDATE usuarios SET acesso_atividades = NOT COALESCE(acesso_atividades, FALSE)
        WHERE id = %s RETURNING acesso_atividades
    """, (id,))
    row = cursor.fetchone()
    conexao.commit()
    conexao.close()
    if row:
        estado = "concedido" if row[0] else "removido"
        logger.info(f"Acesso a Fotos Quadrimestral {estado} para usuário ID={id} por {current_user.id}")
        flash(f"✅ Acesso a Fotos Quadrimestral {estado}!", "success")
    return redirect(url_for("listar_usuarios"))

@app.route("/usuario/editar_email/<int:id>", methods=["POST"])
@login_required
def editar_email_usuario(id):
    # Somente administradores podem cadastrar/alterar o e-mail de outro usuário
    if current_user.perfil != 'admin':
        return "Acesso negado", 403
    email = request.form.get("email", "").strip()
    if email and ('@' not in email or '.' not in email.split('@')[-1]):
        flash("❌ Informe um e-mail válido.", "danger")
        return redirect(url_for("listar_usuarios"))
    conexao = get_db()
    cursor = conexao.cursor()
    cursor.execute("UPDATE usuarios SET email = %s WHERE id = %s", (email or None, id))
    conexao.commit()
    conexao.close()
    logger.info(f"E-mail do usuário ID={id} alterado por {current_user.id}")
    flash("✅ E-mail atualizado!", "success")
    return redirect(url_for("listar_usuarios"))

@app.route("/usuario/resetar_senha/<int:id>", methods=["POST"])
@login_required
def resetar_senha_usuario(id):
    # Fallback manual enquanto o envio de e-mail não está confiável:
    # somente admin pode forçar uma nova senha para qualquer usuário.
    if current_user.perfil != 'admin':
        return "Acesso negado", 403
    nova = request.form.get("nova_senha", "")
    if len(nova) < 6:
        flash("❌ Mínimo 6 caracteres!", "danger")
    elif not any(c.isupper() for c in nova):
        flash("❌ A senha precisa ter pelo menos uma letra maiúscula!", "danger")
    elif not any(c.isdigit() for c in nova):
        flash("❌ A senha precisa ter pelo menos um número!", "danger")
    else:
        hash_nova = bcrypt.hashpw(nova.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        conexao = get_db()
        cursor = conexao.cursor()
        cursor.execute(
            "UPDATE usuarios SET senha = %s, primeiro_acesso = 1, reset_token = NULL, reset_token_expira = NULL WHERE id = %s",
            (hash_nova, id)
        )
        conexao.commit()
        conexao.close()
        logger.info(f"Senha do usuário ID={id} resetada manualmente por {current_user.id}")
        flash("✅ Senha redefinida! O usuário precisará trocá-la no próximo login.", "success")
    return redirect(url_for("listar_usuarios"))

@app.route("/usuario/alterar_senha", methods=["POST"])
@login_required
def alterar_senha_simples():
    atual = request.form.get("senha_atual", "")
    nova = request.form.get("nova_senha", "")
    confirma = request.form.get("confirmar_senha", "")
    # Verificar senha atual
    conexao = get_db()
    cursor = conexao.cursor()
    cursor.execute("SELECT senha FROM usuarios WHERE usuario = %s", (current_user.id,))
    row = cursor.fetchone()
    conexao.close()
    if not row or not bcrypt.checkpw(atual.encode('utf-8'), row[0].encode('utf-8')):
        flash("❌ Senha atual incorreta!", "danger")
        return redirect(url_for("solicitacoes"))
    if len(nova) < 6:
        flash("Mínimo 6 caracteres!", "danger")
        return redirect(url_for("solicitacoes"))
    if not any(c.isupper() for c in nova):
        flash("A nova senha precisa ter pelo menos uma letra maiúscula!", "danger")
        return redirect(url_for("solicitacoes"))
    if not any(c.isdigit() for c in nova):
        flash("A nova senha precisa ter pelo menos um número!", "danger")
        return redirect(url_for("solicitacoes"))
    if nova != confirma:
        flash("❌ As senhas não conferem!", "danger")
        return redirect(url_for("solicitacoes"))
    hash_nova = bcrypt.hashpw(nova.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    conexao = get_db()
    cursor = conexao.cursor()
    cursor.execute("UPDATE usuarios SET senha = %s WHERE usuario = %s", (hash_nova, current_user.id))
    conexao.commit()
    conexao.close()
    logger.info(f"Senha alterada por {current_user.id}")
    flash("✅ Senha alterada com sucesso!", 'success')
    return redirect(url_for("solicitacoes"))

# =====================================================
# CONFIGURAÇÕES DO SISTEMA
# =====================================================

def get_salario_minimo():
    """Busca o salário mínimo vigente do banco de dados"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT valor FROM configuracoes WHERE chave = 'salario_minimo'")
        row = cursor.fetchone()
        conn.close()
        return float(row[0]) if row else 1621.00
    except Exception as e:
        logger.error(f"get_salario_minimo: falha ao consultar salário mínimo configurado, usando fallback R$ 1621,00: {e}")
        return 1621.00

@app.route("/api/configuracoes")
@login_required
def api_configuracoes():
    """Retorna configurações públicas para uso no frontend"""
    salario = get_salario_minimo()
    return jsonify({
        'salario_minimo': salario,
        'limite_renda_per_capita': round(salario / 4, 2)
    })

@app.route("/configuracoes", methods=["GET", "POST"])
@login_required
def configuracoes():
    if current_user.perfil not in ['admin', 'gestor']:
        return redirect(url_for("dashboard"))
    erro = sucesso = None
    conn = get_db()
    cursor = conn.cursor()
    try:
        if request.method == "POST" and current_user.perfil == 'admin':
            novo_salario = request.form.get("salario_minimo", "").strip().replace(",", ".")
            try:
                valor = float(novo_salario)
                if valor <= 0:
                    raise ValueError
                cursor.execute("""
                    INSERT INTO configuracoes (chave, valor, descricao, atualizado_em)
                    VALUES ('salario_minimo', %s, 'Salário mínimo nacional vigente (R$)', %s)
                    ON CONFLICT (chave) DO UPDATE SET valor = EXCLUDED.valor, atualizado_em = EXCLUDED.atualizado_em
                """, (str(valor), datetime.now(FUSO_RONDONIA).strftime('%d/%m/%Y %H:%M:%S')))
                conn.commit()
                logger.info(f"Salário mínimo atualizado para R$ {valor} por {current_user.id}")
                sucesso = f"✅ Salário mínimo atualizado para R$ {valor:.2f}! Novo limite de renda per capita: R$ {valor/4:.2f}"
            except ValueError:
                erro = "❌ Valor inválido. Digite um número positivo (ex: 1518.00)"

        # Tudo numa só conexão
        cursor.execute("SELECT chave, valor, descricao, atualizado_em FROM configuracoes ORDER BY chave")
        configs = cursor.fetchall()

        cursor.execute("SELECT valor FROM configuracoes WHERE chave = 'salario_minimo'")
        row = cursor.fetchone()
        salario = float(row[0]) if row else 1621.00

        cursor.execute("SELECT id, cras, bairro, entrega_volante FROM cras_bairros ORDER BY cras, bairro")
        bairros_por_cras = {}
        bairros_com_id   = {}
        lista_cras_set   = []
        for bid, bcras, bbairro, bvolante in cursor.fetchall():
            bairros_por_cras.setdefault(bcras, []).append(bbairro)
            bairros_com_id.setdefault(bcras, []).append((bid, bbairro, bvolante))
            if bcras not in lista_cras_set:
                lista_cras_set.append(bcras)

        cursor.execute("SELECT id, nome FROM servicos ORDER BY nome")
        servicos = cursor.fetchall()
    finally:
        conn.close()

    return render_template("configuracoes.html",
        configs=configs,
        salario_minimo=salario,
        limite_rpc=round(salario / 4, 2),
        erro=erro,
        sucesso=sucesso,
        bairros_por_cras=bairros_por_cras,
        bairros_com_id=bairros_com_id,
        lista_cras=lista_cras_set,
        servicos=servicos,
        current_user=current_user
    )

@app.route("/configuracoes/restaurar_backup", methods=["GET", "POST"])
@login_required
def restaurar_backup():
    """Restaura o banco a partir de um arquivo de backup (.zip ou .json).
    Apaga e recria as tabelas presentes no arquivo, numa unica transacao:
    se qualquer etapa falhar, nada e alterado (ver get_db()/_PooledConn,
    que da rollback automatico ao fechar sem commit)."""
    if current_user.perfil != 'admin':
        return "Acesso negado", 403

    if request.method == "GET":
        return render_template("restaurar_backup.html")

    if not request.form.get("confirmacao"):
        flash("❌ Você precisa marcar a caixa de confirmação para restaurar o backup.", "danger")
        return redirect(url_for("restaurar_backup"))

    arquivo = request.files.get("arquivo_backup")
    if not arquivo or not arquivo.filename:
        flash("❌ Nenhum arquivo enviado.", "danger")
        return redirect(url_for("restaurar_backup"))

    try:
        conteudo_bruto = arquivo.read()
        nome = arquivo.filename.lower()
        if nome.endswith(".zip"):
            with zipfile.ZipFile(io.BytesIO(conteudo_bruto)) as zf:
                nomes_json = [n for n in zf.namelist() if n.lower().endswith(".json")]
                if not nomes_json:
                    raise ValueError("O arquivo .zip não contém nenhum .json.")
                conteudo_json = zf.read(nomes_json[0])
        elif nome.endswith(".json"):
            conteudo_json = conteudo_bruto
        else:
            raise ValueError("Envie um arquivo .zip ou .json gerado pelo backup do sistema.")

        dados = json.loads(conteudo_json)
        if "tabelas" not in dados or not isinstance(dados["tabelas"], dict):
            raise ValueError(
                "Formato de backup não reconhecido (faltando a chave 'tabelas'). "
                "Esse arquivo pode ser de uma versão antiga do backup — gere um novo antes de restaurar."
            )
    except Exception as e:
        flash(f"❌ Erro ao ler o arquivo de backup: {e}", "danger")
        return redirect(url_for("restaurar_backup"))

    tabelas_backup = dados["tabelas"]
    conn = get_db()
    cursor = conn.cursor()
    contagens = {}
    try:
        # Apaga na ordem inversa da lista (filhos antes de pais, respeitando FKs)
        for tabela in reversed(TABELAS_BACKUP):
            if tabela in tabelas_backup:
                cursor.execute(f"DELETE FROM {tabela}")

        # Insere de volta na ordem normal (pais antes de filhos)
        for tabela in TABELAS_BACKUP:
            linhas = tabelas_backup.get(tabela)
            if not linhas:
                contagens[tabela] = 0
                continue

            cursor.execute(f"SELECT * FROM {tabela} LIMIT 0")
            colunas_atuais = [d[0] for d in cursor.description]
            colunas_usadas = [c for c in colunas_atuais if c in linhas[0]]

            colunas_sql  = ", ".join(colunas_usadas)
            placeholders = ", ".join(["%s"] * len(colunas_usadas))
            valores = [tuple(linha.get(c) for c in colunas_usadas) for linha in linhas]
            cursor.executemany(
                f"INSERT INTO {tabela} ({colunas_sql}) VALUES ({placeholders})", valores
            )
            contagens[tabela] = len(linhas)

            # Todas as tabelas usam "id SERIAL PRIMARY KEY" como primeira
            # coluna, exceto configuracoes (chave TEXT) — que nao precisa
            # de reset de sequencia.
            if colunas_atuais and colunas_atuais[0] == 'id':
                cursor.execute(
                    f"SELECT setval(pg_get_serial_sequence('{tabela}', 'id'), "
                    f"COALESCE((SELECT MAX(id) FROM {tabela}), 1))"
                )

        conn.commit()
        resumo = ", ".join(f"{t}: {n}" for t, n in contagens.items())
        logger.info(f"Backup restaurado por {current_user.id}: {resumo}")
        flash(f"✅ Backup restaurado com sucesso! {resumo}", "success")
    except Exception as e:
        logger.error(f"Erro ao restaurar backup (solicitado por {current_user.id}): {e}")
        flash(f"❌ Erro ao restaurar backup — nenhuma alteração foi salva: {e}", "danger")
    finally:
        conn.close()

    return redirect(url_for("configuracoes"))

@app.route("/configuracoes/bairro/adicionar", methods=["POST"])
@login_required
def bairro_adicionar():
    if current_user.perfil not in ['admin', 'gestor']:
        return "Acesso negado", 403
    bairro = request.form.get("bairro", "").strip().upper()
    cras   = request.form.get("cras", "").strip()
    if not bairro or not cras:
        flash("❌ Preencha o nome do bairro e selecione o CRAS.", "danger")
        return redirect(url_for("configuracoes"))
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO cras_bairros (cras, bairro) VALUES (%s, %s)", (cras, bairro))
        conn.commit()
        logger.info(f"Bairro '{bairro}' adicionado ao {cras} por {current_user.id}")
        flash(f"✅ Bairro '{bairro}' adicionado ao {cras}!", "success")
    except Exception:
        conn.rollback()
        flash(f"❌ Bairro '{bairro}' já existe no sistema.", "danger")
    finally:
        conn.close()
    return redirect(url_for("configuracoes"))

@app.route("/configuracoes/bairro/mover", methods=["POST"])
@login_required
def bairro_mover():
    if current_user.perfil not in ['admin', 'gestor']:
        return "Acesso negado", 403
    bairro_id = request.form.get("bairro_id", "")
    novo_cras  = request.form.get("novo_cras", "").strip()
    if not bairro_id or not novo_cras:
        flash("❌ Dados inválidos.", "danger")
        return redirect(url_for("configuracoes"))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE cras_bairros SET cras = %s WHERE id = %s", (novo_cras, bairro_id))
    conn.commit()
    conn.close()
    flash(f"✅ Bairro movido para {novo_cras}!", "success")
    return redirect(url_for("configuracoes"))

@app.route("/configuracoes/bairro/volante", methods=["POST"])
@login_required
def bairro_volante():
    """Alterna se o bairro é atendido pela Equipe Volante na entrega (área rural)."""
    if current_user.perfil not in ['admin', 'gestor']:
        return "Acesso negado", 403
    bairro_id = request.form.get("bairro_id", "")
    if not bairro_id:
        flash("❌ Bairro não identificado.", "danger")
        return redirect(url_for("configuracoes"))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE cras_bairros SET entrega_volante = NOT entrega_volante
        WHERE id = %s RETURNING bairro, entrega_volante
    """, (bairro_id,))
    row = cursor.fetchone()
    conn.commit()
    conn.close()
    if row:
        estado = "marcado para entrega pela EQUIPE VOLANTE" if row[1] else "voltou para entrega pelo CRAS do território"
        logger.info(f"Bairro '{row[0]}' {estado} por {current_user.id}")
        flash(f"✅ Bairro '{row[0]}' {estado}.", "success")
    else:
        flash("❌ Bairro não encontrado.", "danger")
    return redirect(url_for("configuracoes"))


@app.route("/configuracoes/bairro/remover", methods=["POST"])
@login_required
def bairro_remover():
    if current_user.perfil not in ['admin', 'gestor']:
        return "Acesso negado", 403
    bairro_id = request.form.get("bairro_id", "")
    if not bairro_id:
        flash("❌ Bairro não identificado.", "danger")
        return redirect(url_for("configuracoes"))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM cras_bairros WHERE id = %s", (bairro_id,))
    conn.commit()
    conn.close()
    flash("✅ Bairro removido.", "success")
    return redirect(url_for("configuracoes"))

# =====================================================
# SERVIÇOS DA SECRETARIA (além de CRAS/CREAS/Equipe Volante)
# =====================================================

@app.route("/configuracoes/servico/adicionar", methods=["POST"])
@login_required
def servico_adicionar():
    if current_user.perfil not in ['admin', 'gestor']:
        return "Acesso negado", 403
    nome = request.form.get("nome", "").strip()
    if not nome:
        flash("❌ Informe o nome do serviço.", "danger")
        return redirect(url_for("configuracoes"))
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO servicos (nome) VALUES (%s)", (nome,))
        conn.commit()
        logger.info(f"Serviço '{nome}' adicionado por {current_user.id}")
        flash(f"✅ Serviço '{nome}' adicionado!", "success")
    except Exception:
        conn.rollback()
        flash(f"❌ Serviço '{nome}' já existe.", "danger")
    finally:
        conn.close()
    return redirect(url_for("configuracoes"))

@app.route("/configuracoes/servico/renomear", methods=["POST"])
@login_required
def servico_renomear():
    if current_user.perfil not in ['admin', 'gestor']:
        return "Acesso negado", 403
    servico_id = request.form.get("servico_id", "")
    novo_nome = request.form.get("novo_nome", "").strip()
    if not servico_id or not novo_nome:
        flash("❌ Dados inválidos.", "danger")
        return redirect(url_for("configuracoes"))
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("UPDATE servicos SET nome = %s WHERE id = %s", (novo_nome, servico_id))
        conn.commit()
        logger.info(f"Serviço ID={servico_id} renomeado para '{novo_nome}' por {current_user.id}")
        flash("✅ Serviço renomeado!", "success")
    except Exception:
        conn.rollback()
        flash(f"❌ Já existe um serviço chamado '{novo_nome}'.", "danger")
    finally:
        conn.close()
    return redirect(url_for("configuracoes"))

@app.route("/configuracoes/servico/remover", methods=["POST"])
@login_required
def servico_remover():
    if current_user.perfil not in ['admin', 'gestor']:
        return "Acesso negado", 403
    servico_id = request.form.get("servico_id", "")
    if not servico_id:
        flash("❌ Serviço não identificado.", "danger")
        return redirect(url_for("configuracoes"))
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM servicos WHERE id = %s", (servico_id,))
    conn.commit()
    conn.close()
    flash("✅ Serviço removido da lista.", "success")
    return redirect(url_for("configuracoes"))

# =====================================================
# BACKUP MANUAL (rota admin)
# =====================================================

@app.route("/api/backup")
@login_required
def backup():
    if current_user.perfil != 'admin':
        return "Acesso negado", 403
    try:
        enviar_backup_email()
        return "✅ Backup enviado por e-mail com sucesso!"
    except Exception as e:
        return f"❌ Erro ao enviar backup: {e}", 500

@app.route("/api/backup/automatico")
def backup_automatico():
    """Endpoint para cron externo (cron-job.org). Protegido por token secreto."""
    token_enviado = request.args.get('token', '')
    token_esperado = os.environ.get('BACKUP_TOKEN', '')
    if not token_esperado or token_enviado != token_esperado:
        return "Acesso negado", 403
    try:
        resultado = enviar_backup_email()
        if resultado is False:
            return "Erro: BREVO_API_KEY nao configurada", 500
        return "Backup enviado com sucesso", 200
    except Exception as e:
        logger.error(f"Erro no backup automatico via cron externo: {e}")
        return f"Erro: {e}", 500

# =====================================================
# ADMINISTRAÇÃO
# =====================================================

@app.route("/administracao")
@login_required
def administracao():
    if current_user.perfil not in ['admin', 'gestor']:
        return redirect(url_for('dashboard'))

    busca_id   = request.args.get('busca_id', '').strip()
    busca_nome = request.args.get('busca_nome', '').strip()

    solicitacao_encontrada = None
    resultados = []

    conn = get_db()
    cur  = conn.cursor()

    if busca_id:
        try:
            cur.execute("""
                SELECT s.id, s.nome, s.cpf, s.status, s.data_solicitacao,
                       COALESCE(u.nome, s.tecnico) as tecnico_nome, s.bairro, s.cras
                FROM solicitacoes s
                LEFT JOIN usuarios u ON s.tecnico = u.usuario
                WHERE s.id = %s
            """, (int(busca_id),))
            row = cur.fetchone()
            if row:
                row = list(row)
                if row[2]: row[2] = formatar_cpf(descriptografar_cpf(row[2]))
                solicitacao_encontrada = row
        except ValueError:
            pass
    elif busca_nome:
        cur.execute("""
            SELECT s.id, s.nome, s.cpf, s.status, s.data_solicitacao,
                   COALESCE(u.nome, s.tecnico) as tecnico_nome, s.bairro, s.cras
            FROM solicitacoes s
            LEFT JOIN usuarios u ON s.tecnico = u.usuario
            WHERE s.nome ILIKE %s AND s.status = 'Cadastrada'
            ORDER BY s.id DESC LIMIT 20
        """, (f"%{busca_nome}%",))
        for row in cur.fetchall():
            row = list(row)
            if row[2]: row[2] = formatar_cpf(descriptografar_cpf(row[2]))
            resultados.append(row)

    cur.execute("SELECT usuario, nome FROM usuarios ORDER BY nome")
    lista_usuarios = cur.fetchall()

    cur.execute("""
        SELECT s.id, s.nome, s.cancelado_em, COALESCE(u.nome, s.cancelado_por), s.motivo_cancelamento
        FROM solicitacoes s
        LEFT JOIN usuarios u ON s.cancelado_por = u.usuario
        WHERE s.status = 'Cancelada'
        ORDER BY s.id DESC LIMIT 15
    """)
    ultimos_cancelamentos = cur.fetchall()

    conn.close()

    return render_template('administracao.html',
        busca_id=busca_id,
        busca_nome=busca_nome,
        solicitacao_encontrada=solicitacao_encontrada,
        resultados=resultados,
        lista_usuarios=lista_usuarios,
        ultimos_cancelamentos=ultimos_cancelamentos,
        current_user=current_user
    )


@app.route("/cancelar_solicitacao", methods=["POST"])
@login_required
def cancelar_solicitacao():
    if current_user.perfil not in ['admin', 'gestor']:
        return "Acesso negado", 403

    try:
        id = int(request.form.get('solicitacao_id', 0))
    except ValueError:
        id = 0
    if not id:
        flash("ID de solicitação inválido.", "danger")
        return redirect(url_for('administracao'))

    motivo = request.form.get('motivo', '').strip()
    if not motivo:
        flash("Informe o motivo do cancelamento.", "danger")
        return redirect(url_for('administracao'))

    agora = datetime.now(FUSO_RONDONIA).strftime('%d/%m/%Y %H:%M:%S')
    conn  = get_db()
    cur   = conn.cursor()

    cur.execute("SELECT status, tecnico, nome FROM solicitacoes WHERE id = %s", (id,))
    row = cur.fetchone()
    if not row:
        flash("Solicitação não encontrada.", "danger")
        conn.close()
        return redirect(url_for('administracao'))

    status_atual, tecnico, nome_beneficiario = row
    if status_atual != 'Cadastrada':
        flash(f"Só é possível cancelar solicitações com status 'Cadastrada'. Status atual: {status_atual}", "warning")
        conn.close()
        return redirect(url_for('administracao'))

    cur.execute("""
        UPDATE solicitacoes
        SET status='Cancelada', cancelado_por=%s, cancelado_em=%s, motivo_cancelamento=%s
        WHERE id=%s
    """, (current_user.id, agora, motivo, id))

    cur.execute("""
        INSERT INTO historico_edicoes (solicitacao_id, usuario, campo, valor_antes, valor_depois, data_hora)
        VALUES (%s, %s, 'status', 'Cadastrada', 'Cancelada', %s)
    """, (id, current_user.id, agora))

    # Notifica o técnico que criou a solicitação
    if tecnico:
        msg = (f"Sua solicitação #{id} ({nome_beneficiario}) foi cancelada por "
               f"{current_user.nome or current_user.id}. Motivo: {motivo}")
        cur.execute("""
            INSERT INTO notificacoes (destinatario, remetente, mensagem, tipo, criada_em, solicitacao_id)
            VALUES (%s, %s, %s, 'cancelamento', %s, %s)
        """, (tecnico, current_user.id, msg, agora, id))

    conn.commit()
    conn.close()

    logger.info(f"Solicitação #{id} cancelada por {current_user.id}. Motivo: {motivo}")
    flash(f"Solicitação #{id} cancelada com sucesso.", "success")
    return redirect(url_for('administracao'))


@app.route("/administracao/enviar_notificacao", methods=["POST"])
@login_required
def enviar_notificacao():
    if current_user.perfil not in ['admin', 'gestor']:
        return "Acesso negado", 403

    destinatarios = request.form.getlist('destinatarios')
    mensagem      = request.form.get('mensagem', '').strip()

    if not destinatarios or not mensagem:
        flash("Selecione pelo menos um destinatário e escreva a mensagem.", "danger")
        return redirect(url_for('administracao'))

    agora = datetime.now(FUSO_RONDONIA).strftime('%d/%m/%Y %H:%M:%S')
    conn  = get_db()
    cur   = conn.cursor()

    # "todos" é um atalho para selecionar todos os usuários
    if 'todos' in destinatarios:
        cur.execute("SELECT usuario FROM usuarios")
        destinatarios = [r[0] for r in cur.fetchall()]

    for dest in destinatarios:
        cur.execute("""
            INSERT INTO notificacoes (destinatario, remetente, mensagem, tipo, criada_em)
            VALUES (%s, %s, %s, 'comunicado', %s)
        """, (dest, current_user.id, mensagem, agora))

    conn.commit()
    conn.close()
    flash(f"Notificação enviada para {len(destinatarios)} usuário(s).", "success")
    return redirect(url_for('administracao'))


# =====================================================
# NOTIFICAÇÕES
# =====================================================

@app.route("/notificacoes")
@login_required
def notificacoes():
    conn = get_db()
    cur  = conn.cursor()
    cur.execute("""
        SELECT n.id, COALESCE(u.nome, n.remetente), n.mensagem, n.tipo, n.lida, n.criada_em, n.solicitacao_id
        FROM notificacoes n
        LEFT JOIN usuarios u ON n.remetente = u.usuario
        WHERE n.destinatario = %s
        ORDER BY n.id DESC LIMIT 50
    """, (current_user.id,))
    lista = cur.fetchall()
    conn.close()
    return render_template('notificacoes.html', lista_notificacoes=lista, current_user=current_user)


@app.route("/notificacoes/marcar_lida/<int:id>", methods=["POST"])
@login_required
def marcar_notificacao_lida(id):
    conn = get_db()
    cur  = conn.cursor()
    cur.execute("UPDATE notificacoes SET lida=TRUE WHERE id=%s AND destinatario=%s", (id, current_user.id))
    conn.commit()
    conn.close()
    return redirect(url_for('notificacoes'))


@app.route("/notificacoes/marcar_todas_lidas", methods=["POST"])
@login_required
def marcar_todas_lidas():
    conn = get_db()
    cur  = conn.cursor()
    cur.execute("UPDATE notificacoes SET lida=TRUE WHERE destinatario=%s", (current_user.id,))
    conn.commit()
    conn.close()
    return redirect(url_for('notificacoes'))

# =====================================================
# EXECUÇÃO
# =====================================================

if __name__ == "__main__":
    app.run(debug=True)
